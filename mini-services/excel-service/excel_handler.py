"""
Excel Handler - Core Excel file operations using multiple libraries.
Supports .xlsx, .xls, .xlsb formats with appropriate libraries.
"""

import os
import uuid
import shutil
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
import xlrd
import pandas as pd
import numpy as np

import excel_libs
from excel_libs import (
    HAS_POLARS,
    HAS_PYXLSB,
    HAS_XLUTILS,
    HAS_XLWT,
    HAS_PYEXCELERATE,
    PREVIEW_MAX_ROWS,
    get_library_status,
    get_sheet_names_universal,
    read_dataframe,
    read_sheet_preview_pandas,
    write_dataframe,
)
from data_paths import PROJECT_ROOT, UPLOAD_DIR, migrate_legacy_upload_dir


def ensure_upload_dir():
    """Ensure the upload directory exists."""
    os.makedirs(UPLOAD_DIR, exist_ok=True)


def generate_file_id() -> str:
    """Generate a unique file ID using UUID."""
    return str(uuid.uuid4())


def get_file_extension(filename: str) -> str:
    """Get the file extension from a filename."""
    return os.path.splitext(filename)[1].lower()


def is_excel_file(filename: str) -> bool:
    """Check if a file is a supported Excel format."""
    ext = get_file_extension(filename)
    return ext in ['.xlsx', '.xls', '.xlsb', '.xlsm', '.csv', '.tsv']


def save_uploaded_file(file_content: bytes, original_filename: str) -> Dict[str, Any]:
    """Save an uploaded file and return file metadata."""
    ensure_upload_dir()
    file_id = generate_file_id()
    ext = get_file_extension(original_filename)
    stored_filename = f"{file_id}{ext}"
    file_path = os.path.join(UPLOAD_DIR, stored_filename)

    with open(file_path, 'wb') as f:
        f.write(file_content)

    file_size = os.path.getsize(file_path)
    # Крупные xlsm: только сохранить файл; листы — при загрузке в реестр
    if file_size > 12 * 1024 * 1024:
        sheets: List[str] = []
    else:
        sheets = get_sheet_names(file_path)

    return {
        "file_id": file_id,
        "original_filename": original_filename,
        "stored_filename": stored_filename,
        "file_path": file_path,
        "file_size": file_size,
        "extension": ext,
        "sheets": sheets,
        "upload_time": datetime.now().isoformat(),
    }


def get_excel_libraries() -> Dict[str, Any]:
    """Статус подключённых Excel-библиотек (для health/UI)."""
    return get_library_status()


def get_sheet_names(file_path: str) -> List[str]:
    """Get list of sheet names from an Excel file."""
    try:
        names = get_sheet_names_universal(file_path)
        if names:
            return names
    except Exception:
        pass

    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext in ['.xlsx', '.xlsm']:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        elif ext == '.xls':
            wb = xlrd.open_workbook(file_path)
            return wb.sheet_names()
        elif ext == '.xlsb':
            if HAS_PYXLSB:
                wb = pyxlsb.open_workbook(file_path)
                names = wb.sheets
                wb.close()
                return names
            return []
        elif ext == '.csv':
            return ['Sheet1']
        else:
            return []
    except Exception as e:
        return []


def read_sheet_data(
    file_path: str,
    sheet_name: str,
    cell_range: Optional[str] = None,
    max_rows: int = 10000
) -> Dict[str, Any]:
    """
    Read data from a sheet with optional range and row limit.
    Supports pagination for large files.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ['.xlsx', '.xlsm']:
        return _read_xlsx(file_path, sheet_name, cell_range, max_rows)
    elif ext == '.xls':
        return _read_xls(file_path, sheet_name, cell_range, max_rows)
    elif ext == '.xlsb':
        return _read_xlsb(file_path, sheet_name, cell_range, max_rows)
    elif ext == '.csv':
        return _read_csv(file_path, sheet_name, cell_range, max_rows)
    else:
        raise ValueError(f"Unsupported file format: {ext}")


def _read_xlsx(
    file_path: str,
    sheet_name: str,
    cell_range: Optional[str] = None,
    max_rows: int = 10000
) -> Dict[str, Any]:
    """Read data from .xlsx/.xlsm — pandas nrows для превью, иначе openpyxl."""
    if not cell_range and max_rows <= PREVIEW_MAX_ROWS:
        preview = read_sheet_preview_pandas(file_path, sheet_name, max_rows)
        if preview is not None:
            return preview

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]

    # Determine range
    min_row, max_row_val, min_col, max_col_val = 1, ws.max_row or 1, 1, ws.max_column or 1

    if cell_range:
        min_row, min_col, max_row_val, max_col_val = _parse_range(cell_range)

    # Apply max_rows limit
    actual_max_row = min(max_row_val, min_row + max_rows - 1)
    total_rows = max_row_val - min_row + 1
    has_more = max_row_val > actual_max_row

    data = []
    for row_idx, row in enumerate(ws.iter_rows(
        min_row=min_row,
        max_row=actual_max_row,
        min_col=min_col,
        max_col=max_col_val,
        values_only=False
    ), start=min_row):
        row_data = []
        for cell in row:
            cell_info = {
                "row": cell.row,
                "col": cell.column,
                "value": cell.value,
                "type": _get_cell_type(cell.value),
            }
            row_data.append(cell_info)
        data.append(row_data)

    wb.close()

    return {
        "sheet_name": sheet_name,
        "data": data,
        "range": cell_range or f"A1:{get_column_letter(max_col_val)}{max_row_val}",
        "total_rows": total_rows,
        "returned_rows": len(data),
        "has_more": has_more,
        "columns": max_col_val - min_col + 1,
    }


def _read_xls(
    file_path: str,
    sheet_name: str,
    cell_range: Optional[str] = None,
    max_rows: int = 10000
) -> Dict[str, Any]:
    """Read data from .xls file using xlrd."""
    wb = xlrd.open_workbook(file_path)

    try:
        sheet_idx = wb.sheet_names().index(sheet_name)
    except ValueError:
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb.sheet_by_index(sheet_idx)

    min_row, max_row_val, min_col, max_col_val = 0, ws.nrows, 0, ws.ncols

    if cell_range:
        r_min, c_min, r_max, c_max = _parse_range(cell_range)
        min_row = max(0, r_min - 1)
        min_col = max(0, c_min - 1)
        max_row_val = min(ws.nrows, r_max)
        max_col_val = min(ws.ncols, c_max)

    actual_max_row = min(max_row_val, min_row + max_rows)
    total_rows = max_row_val - min_row
    has_more = max_row_val > actual_max_row

    data = []
    for row_idx in range(min_row, actual_max_row):
        row_data = []
        for col_idx in range(min_col, max_col_val):
            cell = ws.cell(row_idx, col_idx)
            cell_info = {
                "row": row_idx + 1,  # 1-indexed
                "col": col_idx + 1,
                "value": cell.value,
                "type": _xlrd_cell_type(cell.ctype),
            }
            row_data.append(cell_info)
        data.append(row_data)

    return {
        "sheet_name": sheet_name,
        "data": data,
        "range": cell_range or f"A1:{get_column_letter(ws.ncols)}{ws.nrows}",
        "total_rows": total_rows,
        "returned_rows": len(data),
        "has_more": has_more,
        "columns": max_col_val - min_col,
    }


def _read_xlsb(
    file_path: str,
    sheet_name: str,
    cell_range: Optional[str] = None,
    max_rows: int = 10000
) -> Dict[str, Any]:
    """Read data from .xlsb file using pyxlsb."""
    if not HAS_PYXLSB:
        raise ImportError("pyxlsb is not installed")

    with pyxlsb.open_workbook(file_path) as wb:
        if sheet_name not in wb.sheets:
            raise ValueError(f"Sheet '{sheet_name}' not found")

        ws = wb.get_sheet(sheet_name)

        min_row, max_row_val, min_col, max_col_val = 1, 1000000, 1, 1000

        if cell_range:
            min_row, min_col, max_row_val, max_col_val = _parse_range(cell_range)

        data = []
        row_count = 0
        for row_idx, row in enumerate(ws.rows(), start=1):
            if row_idx < min_row:
                continue
            if row_idx > max_row_val or row_count >= max_rows:
                break

            row_data = []
            for cell in row:
                col_idx = cell[0] if isinstance(cell, tuple) else cell.column
                val = cell[1] if isinstance(cell, tuple) else cell.v
                if min_col <= col_idx <= max_col_val:
                    row_data.append({
                        "row": row_idx,
                        "col": col_idx,
                        "value": val,
                        "type": _get_cell_type(val),
                    })
            data.append(row_data)
            row_count += 1

    return {
        "sheet_name": sheet_name,
        "data": data,
        "range": cell_range or "A1:unknown",
        "total_rows": max_row_val - min_row + 1,
        "returned_rows": len(data),
        "has_more": False,
        "columns": max_col_val - min_col + 1,
    }


def _read_csv(
    file_path: str,
    sheet_name: str,
    cell_range: Optional[str] = None,
    max_rows: int = 10000
) -> Dict[str, Any]:
    """Read data from CSV file using pandas."""
    df = pd.read_csv(file_path, nrows=max_rows)

    min_row, max_row_val, min_col, max_col_val = 1, len(df) + 1, 1, len(df.columns) + 1

    if cell_range:
        min_row, min_col, max_row_val, max_col_val = _parse_range(cell_range)

    data = []
    for row_idx in range(max(0, min_row - 1), min(len(df), max_row_val - 1)):
        row_data = []
        for col_idx in range(max(0, min_col - 1), min(len(df.columns), max_col_val - 1)):
            val = df.iloc[row_idx, col_idx]
            row_data.append({
                "row": row_idx + 1,
                "col": col_idx + 1,
                "value": _nan_to_none(val),
                "type": _get_cell_type(_nan_to_none(val)),
            })
        data.append(row_data)

    return {
        "sheet_name": "Sheet1",
        "data": data,
        "range": cell_range or f"A1:{get_column_letter(len(df.columns))}{len(df)}",
        "total_rows": len(df),
        "returned_rows": len(data),
        "has_more": False,
        "columns": len(df.columns),
    }


def update_cells(
    file_path: str,
    sheet_name: str,
    changes: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """Update cell values in an Excel file."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ['.xlsx', '.xlsm']:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        updated = 0

        for change in changes:
            row = change.get('row')
            col = change.get('col')
            value = change.get('value')

            if row is not None and col is not None:
                ws.cell(row=row, column=col, value=value)
                updated += 1

        wb.save(file_path)
        wb.close()

        return {"updated": updated, "sheet": sheet_name}
    else:
        raise ValueError(f"Writing to {ext} format is not supported. Convert to .xlsx first.")


def create_sheet(file_path: str, sheet_name: str) -> Dict[str, Any]:
    """Create a new sheet in an Excel file."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ['.xlsx', '.xlsm']:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' already exists")

        wb.create_sheet(title=sheet_name)
        wb.save(file_path)
        wb.close()

        return {"created": sheet_name, "sheets": get_sheet_names(file_path)}
    else:
        raise ValueError(f"Creating sheets in {ext} format is not supported.")


def delete_sheet(file_path: str, sheet_name: str) -> Dict[str, Any]:
    """Delete a sheet from an Excel file."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ['.xlsx', '.xlsm']:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found")

        if len(wb.sheetnames) <= 1:
            wb.close()
            raise ValueError("Cannot delete the last sheet")

        del wb[sheet_name]
        wb.save(file_path)
        wb.close()

        return {"deleted": sheet_name, "sheets": get_sheet_names(file_path)}
    else:
        raise ValueError(f"Deleting sheets in {ext} format is not supported.")


def rename_sheet(file_path: str, old_name: str, new_name: str) -> Dict[str, Any]:
    """Rename a sheet in an Excel file."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ['.xlsx', '.xlsm']:
        wb = openpyxl.load_workbook(file_path)
        if old_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{old_name}' not found")

        if new_name in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{new_name}' already exists")

        ws = wb[old_name]
        ws.title = new_name
        wb.save(file_path)
        wb.close()

        return {"renamed": {"from": old_name, "to": new_name}, "sheets": get_sheet_names(file_path)}
    else:
        raise ValueError(f"Renaming sheets in {ext} format is not supported.")


def get_file_info(file_path: str) -> Dict[str, Any]:
    """Get detailed file metadata."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    stat = os.stat(file_path)
    sheets = get_sheet_names(file_path)

    return {
        "file_path": file_path,
        "file_size": stat.st_size,
        "created": datetime.fromtimestamp(stat.st_ctime).isoformat(),
        "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        "extension": os.path.splitext(file_path)[1].lower(),
        "sheets": sheets,
        "sheet_count": len(sheets),
    }


def get_sheet_info(file_path: str, sheet_name: str) -> Dict[str, Any]:
    """Get sheet dimensions and detailed info."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ['.xlsx', '.xlsm']:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        info = {
            "sheet_name": sheet_name,
            "min_row": ws.min_row,
            "max_row": ws.max_row,
            "min_column": ws.min_column,
            "max_column": ws.max_column,
            "rows": ws.max_row if ws.max_row else 0,
            "columns": ws.max_column if ws.max_column else 0,
        }
        wb.close()
        return info
    elif ext == '.xls':
        wb = xlrd.open_workbook(file_path)
        try:
            sheet_idx = wb.sheet_names().index(sheet_name)
        except ValueError:
            raise ValueError(f"Sheet '{sheet_name}' not found")

        ws = wb.sheet_by_index(sheet_idx)
        return {
            "sheet_name": sheet_name,
            "min_row": 0,
            "max_row": ws.nrows,
            "min_column": 0,
            "max_column": ws.ncols,
            "rows": ws.nrows,
            "columns": ws.ncols,
        }
    elif ext == '.csv':
        df = pd.read_csv(file_path, nrows=1)
        total_rows = sum(1 for _ in open(file_path)) - 1  # subtract header
        return {
            "sheet_name": "Sheet1",
            "min_row": 1,
            "max_row": total_rows,
            "min_column": 1,
            "max_column": len(df.columns),
            "rows": total_rows,
            "columns": len(df.columns),
        }
    else:
        raise ValueError(f"Unsupported format: {ext}")


def create_new_workbook(file_path: str, sheet_name: str = "Sheet1") -> str:
    """Create a new Excel workbook with one sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    wb.save(file_path)
    wb.close()
    return file_path


def convert_file(input_path: str, output_format: str) -> Dict[str, Any]:
    """Convert between Excel formats (pandas/openpyxl/xlrd/xlutils)."""
    ext = os.path.splitext(input_path)[1].lower()
    base_name = os.path.splitext(input_path)[0]
    output_path = f"{base_name}_converted.{output_format}"
    engine_used = "pandas"

    if ext == ".xls" and output_format == "xls" and HAS_XLUTILS:
        try:
            meta = excel_libs.copy_xls_with_xlutils(input_path, output_path)
            return {
                "input_path": input_path,
                "output_path": output_path,
                "output_format": output_format,
                "engine": meta.get("engine", "xlutils"),
            }
        except Exception:
            pass

    df = read_dataframe(input_path)

    if output_format == "xlsx":
        wr = write_dataframe(df, output_path)
        engine_used = wr.get("engine", "openpyxl")
    elif output_format == "xls":
        if HAS_XLWT:
            df.to_excel(output_path, index=False, engine="xlwt")
            engine_used = "xlwt"
        else:
            raise ImportError("xlwt not available for .xls output")
    elif output_format == "csv":
        df.to_csv(output_path, index=False)
        engine_used = "pandas"
    elif output_format == "tsv":
        df.to_csv(output_path, index=False, sep="\t")
        engine_used = "pandas"
    elif output_format == "json":
        df.to_json(output_path, orient="records", indent=2)
        engine_used = "pandas"
    elif output_format == "parquet":
        try:
            if HAS_POLARS:
                import polars as pl

                pl.from_pandas(df).write_parquet(output_path)
                engine_used = "polars"
            else:
                df.to_parquet(output_path, index=False)
                engine_used = "pandas"
        except Exception:
            df.to_parquet(output_path, index=False)
            engine_used = "pandas"
    elif output_format == "html":
        df.to_html(output_path, index=False)
        engine_used = "pandas"
    else:
        raise ValueError(f"Unsupported output format: {output_format}")

    return {
        "input_path": input_path,
        "output_path": output_path,
        "output_format": output_format,
        "rows": len(df),
        "columns": len(df.columns),
        "engine": engine_used,
    }


def list_uploaded_files(*, include_sheets: bool = False) -> List[Dict[str, Any]]:
    """List all files in the upload directory.

    include_sheets=False by default — reading sheet names from 18+ large xlsm blocks the server for minutes.
    """
    ensure_upload_dir()
    files = []

    for filename in os.listdir(UPLOAD_DIR):
        file_path = os.path.join(UPLOAD_DIR, filename)
        if os.path.isfile(file_path):
            stat = os.stat(file_path)
            ext = os.path.splitext(filename)[1].lower()

            file_info = {
                "file_id": os.path.splitext(filename)[0],
                "stored_filename": filename,
                "file_path": file_path,
                "file_size": stat.st_size,
                "extension": ext,
                "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            }

            if include_sheets and is_excel_file(filename):
                try:
                    file_info["sheets"] = get_sheet_names(file_path)
                except Exception:
                    file_info["sheets"] = []
            else:
                file_info["sheets"] = []

            files.append(file_info)

    return files


def delete_file(file_path: str) -> bool:
    """Delete a file from disk."""
    if os.path.exists(file_path):
        os.remove(file_path)
        return True
    return False


def find_file_by_id(file_id: str) -> Optional[str]:
    """Find a file path by its ID (UUID prefix of filename)."""
    ensure_upload_dir()
    for filename in os.listdir(UPLOAD_DIR):
        if filename.startswith(file_id):
            return os.path.join(UPLOAD_DIR, filename)
    return None


# --- Helper Functions ---

def _parse_range(cell_range: str) -> Tuple[int, int, int, int]:
    """Parse an Excel range string like 'A1:Z100' into (min_row, min_col, max_row, max_col)."""
    try:
        parts = cell_range.split(':')
        if len(parts) == 1:
            col_str, row = coordinate_from_string(parts[0].strip())
            col = column_index_from_string(col_str)
            return (row, col, row, col)
        else:
            col_str1, row1 = coordinate_from_string(parts[0].strip())
            col_str2, row2 = coordinate_from_string(parts[1].strip())
            col1 = column_index_from_string(col_str1)
            col2 = column_index_from_string(col_str2)
            return (min(row1, row2), min(col1, col2), max(row1, row2), max(col1, col2))
    except Exception:
        raise ValueError(f"Invalid range format: {cell_range}. Expected format like 'A1:Z100'")


def _get_cell_type(value) -> str:
    """Determine the type of a cell value."""
    if value is None:
        return "null"
    elif isinstance(value, bool):
        return "boolean"
    elif isinstance(value, (int, float)):
        return "number"
    elif isinstance(value, str):
        return "string"
    elif isinstance(value, datetime):
        return "datetime"
    else:
        return "unknown"


def _xlrd_cell_type(ctype: int) -> str:
    """Convert xlrd cell type to string."""
    types = {0: "empty", 1: "string", 2: "number", 3: "date", 4: "boolean", 5: "error", 6: "blank"}
    return types.get(ctype, "unknown")


def _nan_to_none(value):
    """Convert NaN values to None."""
    if isinstance(value, float) and np.isnan(value):
        return None
    return value
