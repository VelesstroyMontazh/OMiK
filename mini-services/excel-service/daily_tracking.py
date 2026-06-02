"""
Ежедневный учёт — порт логики HR-Data-Manager (лист ЕЖЕДНЕВНЫЙ, строки с 6, колонки B–Q).
Площадки (ОП) — из справочников; location_id = название площадки.
"""

from __future__ import annotations

import json
import os
import sqlite3
import uuid
from datetime import date, datetime
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl.utils.datetime import from_excel

from data_paths import UPLOAD_DIR
import references

DB_PATH = os.path.join(UPLOAD_DIR, "daily_tracking.sqlite")
HEADER_ROW = 4
DATA_START = HEADER_ROW + 1


def _conn() -> sqlite3.Connection:
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _init_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS daily_tracking (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            location_id TEXT NOT NULL,
            status TEXT,
            region TEXT,
            location_name_1c TEXT,
            tab_number TEXT,
            fio TEXT,
            birth_date_1c TEXT,
            citizenship TEXT,
            passport_series_1c TEXT,
            passport_number_1c TEXT,
            actual_position TEXT,
            section TEXT,
            visa TEXT,
            visa_type TEXT,
            region2 TEXT,
            visa_expiry TEXT,
            entry_date TEXT,
            category TEXT,
            upload_batch_id TEXT,
            created_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_daily_date ON daily_tracking(date);
        CREATE INDEX IF NOT EXISTS idx_daily_loc ON daily_tracking(location_id);
        CREATE TABLE IF NOT EXISTS daily_tracking_uploads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            location_id TEXT NOT NULL,
            tracking_date TEXT NOT NULL,
            original_name TEXT,
            row_count INTEGER,
            file_path TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS daily_combined_tracking (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            location_id TEXT NOT NULL,
            status TEXT,
            region TEXT,
            location_name_1c TEXT,
            tab_number TEXT,
            fio TEXT,
            birth_date_1c TEXT,
            citizenship TEXT,
            passport_series_1c TEXT,
            passport_number_1c TEXT,
            actual_position TEXT,
            section TEXT,
            visa TEXT,
            visa_type TEXT,
            region2 TEXT,
            visa_expiry TEXT,
            entry_date TEXT,
            category TEXT,
            built_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_daily_combined_date ON daily_combined_tracking(date);
        """
    )
    try:
        conn.execute("ALTER TABLE daily_tracking_uploads ADD COLUMN file_path TEXT")
    except sqlite3.OperationalError:
        pass


def _op_status_from_label(status: str) -> str:
    s = references._norm_key(status)
    if not s:
        return "active"
    if references._is_active_status(status):
        return "active"
    if "законч" in s or "заверш" in s:
        return "finished"
    if "приостанов" in s:
        return "paused"
    return "paused"


def list_ploshchadki_detailed(active_only: bool = True) -> List[Dict[str, str]]:
    """Площадки: в приоритете активные из Основной БД (колонка «Статус» = Актив.), иначе справочники."""
    data = references.get_cached()
    site_status: Dict[str, str] = data.get("site_status") or {}
    main_active = _sites_from_main_db_active()
    if active_only and main_active:
        names = main_active
    elif active_only:
        names = list_ploshchadki(active_only=True)
    else:
        names = list_ploshchadki(active_only=False)
    out: List[Dict[str, str]] = []
    for name in names:
        key = references._norm_key(name)
        st = site_status.get(key, "")
        op = "active" if name in main_active else _op_status_from_label(st)
        if active_only and op != "active":
            continue
        out.append({
            "name": name,
            "opStatus": op,
            "statusLabel": st or ("Актив." if name in main_active else "Активное ОП"),
        })
    rank = {"active": 0, "paused": 1, "finished": 2}
    out.sort(key=lambda x: (rank.get(x["opStatus"], 9), x["name"].lower()))
    return out


def _stored_upload_path(location_id: str, tracking_date: str) -> str:
    safe_loc = references._norm_key(location_id)[:80] or "site"
    safe_date = (tracking_date or "").replace("/", "-")[:32] or "date"
    return os.path.join(DAILY_SOURCE_FILES_DIR, safe_loc, f"{safe_date}.xlsx")


def _save_upload_file(source_path: str, location_id: str, tracking_date: str) -> str:
    dest = _stored_upload_path(location_id, tracking_date)
    os.makedirs(os.path.dirname(dest), exist_ok=True)
    import shutil
    shutil.copy2(source_path, dest)
    return dest


def _delete_stored_upload(location_id: str, tracking_date: str) -> None:
    path = _stored_upload_path(location_id, tracking_date)
    if os.path.isfile(path):
        try:
            os.remove(path)
        except OSError:
            pass
    parent = os.path.dirname(path)
    try:
        if os.path.isdir(parent) and not os.listdir(parent):
            os.rmdir(parent)
    except OSError:
        pass


def clear_site_date(
    location_id: str,
    tracking_date: str,
    *,
    role: Optional[str] = None,
    user_sites: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Удалить все строки учёта за дату по площадке, журнал загрузок и сохранённый XLSX."""
    try:
        assert_upload_access(location_id, role=role, sites=user_sites)
    except PermissionError as e:
        return {"error": str(e)}
    loc = (location_id or "").strip()
    dt = (tracking_date or "").strip()
    if not loc or not dt:
        return {"error": "Укажите площадку и дату"}

    conn = _conn()
    try:
        _init_schema(conn)
        cur = conn.execute(
            "DELETE FROM daily_tracking WHERE date = ? AND location_id = ?",
            (dt, loc),
        )
        deleted_rows = cur.rowcount
        cur_u = conn.execute(
            "DELETE FROM daily_tracking_uploads WHERE tracking_date = ? AND location_id = ?",
            (dt, loc),
        )
        deleted_uploads = cur_u.rowcount
        conn.commit()
    finally:
        conn.close()

    _delete_stored_upload(loc, dt)
    return {
        "success": True,
        "deletedRows": deleted_rows,
        "deletedUploads": deleted_uploads,
        "locationId": loc,
        "trackingDate": dt,
    }


def clear_combined_report(
    tracking_date: str,
    *,
    role: Optional[str] = None,
    user_sites: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Удалить только сформированный «Общий» отчёт за дату (данные по ОП не трогаются)."""
    _ = user_sites
    r = (role or "").strip().lower()
    if r not in ("admin", "cok"):
        return {"error": "Очистка «Общий» доступна только Admin и ЦОК"}
    dt = (tracking_date or "").strip()
    if not dt:
        return {"error": "Укажите дату"}
    conn = _conn()
    try:
        _init_schema(conn)
        cur = conn.execute("DELETE FROM daily_combined_tracking WHERE date = ?", (dt,))
        deleted_rows = cur.rowcount
        conn.commit()
    finally:
        conn.close()
    return {
        "success": True,
        "deletedRows": deleted_rows,
        "trackingDate": dt,
        "combined": True,
    }


def assert_upload_access(
    location_id: str,
    *,
    role: Optional[str] = None,
    sites: Optional[List[str]] = None,
) -> None:
    loc = (location_id or "").strip()
    if not loc:
        raise PermissionError("Не указана площадка")
    r = (role or "").strip().lower()
    if r in ("admin", "cok"):
        return
    allowed = {references._norm_key(s) for s in (sites or []) if str(s).strip()}
    if references._norm_key(loc) not in allowed:
        raise PermissionError(f"Нет прав на загрузку для площадки «{loc}»")


def list_ploshchadki(active_only: bool = True) -> List[str]:
    data = references.get_cached()
    sites: set[str] = set()
    for m in (data.get("territory_to_site") or {}).values():
        if m:
            sites.add(str(m).strip())
    for m in (data.get("podr_to_site") or {}).values():
        if m:
            sites.add(str(m).strip())
    for u in data.get("users") or []:
        for s in u.get("sites") or []:
            if s:
                sites.add(str(s).strip())

    site_status: Dict[str, str] = data.get("site_status") or {}
    out = sorted(sites)
    if not active_only or not site_status:
        return out
    active: List[str] = []
    for site in out:
        key = references._norm_key(site)
        st = site_status.get(key, "")
        if not st or references._is_active_status(st):
            active.append(site)
    return active or out


def _parse_daily_excel(file_path: str) -> List[Dict[str, Any]]:
    xl = pd.ExcelFile(file_path, engine="openpyxl")
    sheet = None
    for name in xl.sheet_names:
        u = name.upper()
        if "ЕЖЕДНЕВНЫЙ" in u or "DAILY" in u:
            sheet = name
            break
    if sheet is None:
        sheet = xl.sheet_names[0]
    df = pd.read_excel(file_path, sheet_name=sheet, header=None, engine="openpyxl")
    rows_out: List[Dict[str, Any]] = []
    for i in range(DATA_START, len(df)):
        row = df.iloc[i]
        if len(row) < 7:
            continue
        tab = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""
        fio = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ""
        citizenship = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ""
        if (not tab and not fio) or not citizenship:
            continue
        def cell(idx: int) -> str:
            if idx >= len(row) or pd.isna(row.iloc[idx]):
                return ""
            return str(row.iloc[idx]).strip()

        rows_out.append({
            "status": "Я",
            "region": cell(1),
            "location_name_1c": cell(2),
            "tab_number": tab or None,
            "fio": fio or None,
            "birth_date_1c": cell(5),
            "citizenship": citizenship,
            "passport_series_1c": cell(7),
            "passport_number_1c": cell(8),
            "actual_position": cell(9),
            "section": cell(10),
            "visa": cell(11),
            "visa_type": cell(12),
            "region2": cell(13),
            "visa_expiry": cell(14),
            "entry_date": cell(15),
            "category": "ROP",
        })
    return rows_out


def _parse_date_value(val: Any) -> Optional[date]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            return from_excel(float(val)).date()
        except (ValueError, TypeError, OverflowError):
            pass
    s = str(val).strip()
    if not s:
        return None
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").date()
        except ValueError:
            pass
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d.%m.%y", "%d/%m/%y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    mdy = s.split("/")
    if len(mdy) == 3:
        try:
            m, d, y = int(mdy[0]), int(mdy[1]), int(mdy[2])
            if y < 100:
                y += 2000
            return date(y, m, d)
        except ValueError:
            pass
    return None


def _format_date_display(val: Any) -> Any:
    d = _parse_date_value(val)
    if d:
        return d.strftime("%d.%m.%Y")
    return val


def _rows_for_api(data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in data:
        r = dict(row)
        for key in DATE_FIELD_KEYS:
            if key in r and r[key] is not None:
                r[key] = _format_date_display(r[key])
        out.append(r)
    return out


def _load_aup_position_map() -> Dict[str, str]:
    """Столбец A → должность, столбец B → статус (АУП_РОП_ИТР)."""
    if not os.path.isfile(DAILY_AUP_PATH):
        return {}
    from openpyxl import load_workbook

    out: Dict[str, str] = {}
    try:
        wb = load_workbook(DAILY_AUP_PATH, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            pos = str(row[0] or "").strip()
            status_val = str(row[1] or "").strip() if len(row) > 1 else ""
            if pos:
                out[references._norm_key(pos)] = status_val
        wb.close()
    except (OSError, ValueError, TypeError):
        return {}
    return out


def _apply_combined_row_transforms(row: Dict[str, Any], aup_map: Dict[str, str]) -> None:
    tab = str(row.get("tab_number") or "").strip()
    tab_l = tab.lower().replace("ё", "е")
    if tab_l in ("прием", "нелегал"):
        row["tab_number"] = "Кандидат"
    pos = str(row.get("actual_position") or "").strip()
    if pos:
        st = aup_map.get(references._norm_key(pos))
        if st:
            row["status"] = st
    for key in DATE_FIELD_KEYS:
        if key in row and row[key] is not None:
            row[key] = _format_date_display(row[key])


def get_aup_status() -> Dict[str, Any]:
    has = os.path.isfile(DAILY_AUP_PATH)
    meta: Dict[str, Any] = {}
    if os.path.isfile(DAILY_AUP_META_PATH):
        try:
            with open(DAILY_AUP_META_PATH, encoding="utf-8") as f:
                meta = json.load(f)
        except (OSError, json.JSONDecodeError):
            meta = {}
    return {
        "hasAup": has,
        "originalName": meta.get("original_name"),
        "uploadedAt": meta.get("uploaded_at"),
        "size": os.path.getsize(DAILY_AUP_PATH) if has else 0,
        "mappingCount": len(_load_aup_position_map()) if has else 0,
    }


def save_aup_file(file_path: str, original_name: str) -> Dict[str, Any]:
    if not os.path.isfile(file_path):
        return {"error": "Файл не найден"}
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    import shutil
    shutil.copy2(file_path, DAILY_AUP_PATH)
    meta = {
        "original_name": original_name,
        "uploaded_at": datetime.now().isoformat(),
    }
    with open(DAILY_AUP_META_PATH, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    return {"success": True, **get_aup_status()}


def build_combined_report(
    tracking_date: str,
    *,
    role: Optional[str] = None,
    user_sites: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Сформировать «Общий» отчёт из данных всех ОП за дату."""
    _ = user_sites
    r = (role or "").strip().lower()
    if r not in ("admin", "cok"):
        return {"error": "Только Admin или ЦОК"}
    dt = (tracking_date or "").strip()
    if not dt:
        return {"error": "Укажите дату"}

    raw = get_rows(dt, combined=False, limit=500_000, offset=0).get("data") or []
    if not raw:
        return {"error": "Нет данных по площадкам за эту дату для формирования «Общий»"}

    aup_map = _load_aup_position_map()
    now = datetime.now().isoformat()
    conn = _conn()
    try:
        _init_schema(conn)
        conn.execute("DELETE FROM daily_combined_tracking WHERE date = ?", (dt,))
        for row in raw:
            rec = dict(row)
            _apply_combined_row_transforms(rec, aup_map)
            conn.execute(
                """
                INSERT INTO daily_combined_tracking (
                    date, location_id, status, region, location_name_1c, tab_number, fio,
                    birth_date_1c, citizenship, passport_series_1c, passport_number_1c,
                    actual_position, section, visa, visa_type, region2, visa_expiry,
                    entry_date, category, built_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    dt,
                    rec.get("location_id"),
                    rec.get("status"),
                    rec.get("region"),
                    rec.get("location_name_1c"),
                    rec.get("tab_number"),
                    rec.get("fio"),
                    rec.get("birth_date_1c"),
                    rec.get("citizenship"),
                    rec.get("passport_series_1c"),
                    rec.get("passport_number_1c"),
                    rec.get("actual_position"),
                    rec.get("section"),
                    rec.get("visa"),
                    rec.get("visa_type"),
                    rec.get("region2"),
                    rec.get("visa_expiry"),
                    rec.get("entry_date"),
                    rec.get("category"),
                    now,
                ),
            )
        conn.commit()
    finally:
        conn.close()

    return {
        "success": True,
        "rowCount": len(raw),
        "trackingDate": dt,
        "hasCombined": True,
        "aupApplied": bool(aup_map),
    }


def get_combined_rows(
    date: str,
    limit: int = 5000,
    offset: int = 0,
) -> Dict[str, Any]:
    conn = _conn()
    try:
        _init_schema(conn)
        where = ['date = ?', 'citizenship IS NOT NULL', "citizenship != ''"]
        params: List[Any] = [date]
        sql = f"""
            SELECT * FROM daily_combined_tracking
            WHERE {' AND '.join(where)}
            ORDER BY location_id, fio
            LIMIT ? OFFSET ?
        """
        params.extend([limit, offset])
        cur = conn.execute(sql, params)
        data = _rows_for_api([dict(r) for r in cur.fetchall()])
        total = conn.execute(
            f"SELECT COUNT(*) FROM daily_combined_tracking WHERE {' AND '.join(where)}",
            params[:-2],
        ).fetchone()[0]
        return {
            "data": data,
            "total": total,
            "offset": offset,
            "limit": limit,
            "hasCombined": total > 0,
        }
    finally:
        conn.close()


def get_rows(
    date: str,
    location_id: Optional[str] = None,
    combined: bool = False,
    limit: int = 5000,
    offset: int = 0,
) -> Dict[str, Any]:
    if combined:
        return get_combined_rows(date, limit=limit, offset=offset)
    conn = _conn()
    try:
        _init_schema(conn)
        where = ['date = ?', 'citizenship IS NOT NULL', "citizenship != ''"]
        params: List[Any] = [date]
        if location_id:
            where.append("location_id = ?")
            params.append(location_id)
        sql = f"""
            SELECT * FROM daily_tracking
            WHERE {' AND '.join(where)}
            ORDER BY location_id, fio
            LIMIT ? OFFSET ?
        """
        params.extend([limit, offset])
        cur = conn.execute(sql, params)
        data = _rows_for_api([dict(r) for r in cur.fetchall()])
        count_sql = f"SELECT COUNT(*) FROM daily_tracking WHERE {' AND '.join(where)}"
        total = conn.execute(count_sql, params[:-2]).fetchone()[0]
        return {"data": data, "total": total, "offset": offset, "limit": limit}
    finally:
        conn.close()


_AUP_KW = (
    "руковод", "началь", "замест", "директор", "менедж", "координатор",
    "администратор", "инспектор", "куратор", "диспетчер", "секретарь",
    "делопроизвод", "бухгал", "экономист", "кадров", "офис", "табель",
)
_ITR_KW = (
    "инженер", "технолог", "геодез", "дефектоскоп", "пто", "прораб",
    "мастер", "механик", "энергетик", "электрик", "лаборант", "смет",
    "снабж", "технадзор", "пусконалад",
)


def _workforce_category(actual_position: Optional[str], legacy: Optional[str] = None) -> str:
    pos = (actual_position or "").lower().replace("ё", "е")
    leg = (legacy or "").upper()
    if "AUP" in leg or "АУП" in leg:
        return "AUP"
    if "ITR" in leg or "ИТР" in leg:
        return "ITR"
    for k in _AUP_KW:
        if k in pos:
            return "AUP"
    for k in _ITR_KW:
        if k in pos:
            return "ITR"
    return "ROP"


def _sites_from_main_db_active() -> List[str]:
    """Уникальные «Площадка» из Основной БД, где «Статус» = активный (Актив.)."""
    import main_db

    if not main_db.is_loaded():
        return []
    path = main_db._db_path()
    meta_path = main_db._meta_path()
    if not os.path.isfile(path) or not os.path.isfile(meta_path):
        return []
    try:
        with open(meta_path, encoding="utf-8") as f:
            meta = json.load(f)
    except (OSError, json.JSONDecodeError):
        return []
    col_map = meta.get("col_mapping") or {}
    site_col = col_map.get("Площадка", "Площадка")
    status_col = col_map.get("Статус", "Статус")
    conn = sqlite3.connect(path)
    try:
        cur = conn.execute(
            f'''
            SELECT DISTINCT "{site_col}" AS site, MAX("{status_col}") AS st
            FROM employees
            WHERE "{site_col}" IS NOT NULL AND TRIM("{site_col}") != ''
            GROUP BY "{site_col}"
            ''',
        )
        out: List[str] = []
        for row in cur.fetchall():
            site = str(row[0] or "").strip()
            st = str(row[1] or "").strip()
            if site and references._is_active_status(st):
                out.append(site)
        return sorted(out, key=lambda s: s.lower())
    finally:
        conn.close()


def get_stats(date: str, location_id: Optional[str] = None, combined: bool = False) -> Dict[str, Any]:
    rows = get_rows(date, location_id, combined=combined, limit=100_000, offset=0)["data"]
    by_cit: Dict[str, int] = {}
    by_loc: Dict[str, int] = {}
    by_cat: Dict[str, Dict[str, int]] = {}
    aup = itr = rop = 0
    for r in rows:
        c = r.get("citizenship") or "—"
        by_cit[c] = by_cit.get(c, 0) + 1
        loc = r.get("location_id") or "—"
        by_loc[loc] = by_loc.get(loc, 0) + 1
        cat = _workforce_category(r.get("actual_position"), r.get("category"))
        if cat == "AUP":
            aup += 1
        elif cat == "ITR":
            itr += 1
        else:
            rop += 1
        bucket = by_cat.setdefault(c, {"AUP": 0, "ITR": 0, "ROP": 0})
        bucket[cat] = bucket.get(cat, 0) + 1

    by_category: List[Dict[str, Any]] = []
    for citizenship, counts in sorted(by_cat.items()):
        for cat in ("AUP", "ITR", "ROP"):
            n = counts.get(cat, 0)
            if n:
                by_category.append({"citizenship": citizenship, "category": cat, "count": n})

    return {
        "total": len(rows),
        "aup": aup,
        "itr": itr,
        "rop": rop,
        "byCitizenship": [{"label": k, "count": v} for k, v in sorted(by_cit.items())],
        "byLocation": [{"label": k, "count": v} for k, v in sorted(by_loc.items())],
        "byCategory": by_category,
    }


def upload_excel(
    file_path: str,
    location_id: str,
    tracking_date: str,
    original_name: str,
    confirm: bool = False,
    *,
    role: Optional[str] = None,
    user_sites: Optional[List[str]] = None,
) -> Dict[str, Any]:
    try:
        assert_upload_access(location_id, role=role, sites=user_sites)
    except PermissionError as e:
        return {"error": str(e)}
    if not os.path.isfile(file_path):
        return {"error": "Файл не найден"}
    parsed = _parse_daily_excel(file_path)
    if not parsed:
        return {"error": "Нет строк для загрузки (проверьте лист и гражданство в колонке G)"}

    conn = _conn()
    try:
        _init_schema(conn)
        existing = conn.execute(
            "SELECT COUNT(*) FROM daily_tracking WHERE date = ? AND location_id = ?",
            (tracking_date, location_id),
        ).fetchone()[0]
        if existing and not confirm:
            return {
                "error": f"Уже есть {existing} записей за эту дату. Подтвердите замену.",
                "existingCount": existing,
                "requiresConfirm": True,
            }
        if existing:
            conn.execute(
                "DELETE FROM daily_tracking WHERE date = ? AND location_id = ?",
                (tracking_date, location_id),
            )
        batch_id = uuid.uuid4().hex[:12]
        now = datetime.now().isoformat()
        for row in parsed:
            for key in DATE_FIELD_KEYS:
                if key in row and row[key] is not None:
                    row[key] = _format_date_display(row[key])
            conn.execute(
                """
                INSERT INTO daily_tracking (
                    date, location_id, status, region, location_name_1c, tab_number, fio,
                    birth_date_1c, citizenship, passport_series_1c, passport_number_1c,
                    actual_position, section, visa, visa_type, region2, visa_expiry,
                    entry_date, category, upload_batch_id, created_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    tracking_date,
                    location_id,
                    row.get("status"),
                    row.get("region"),
                    row.get("location_name_1c"),
                    row.get("tab_number"),
                    row.get("fio"),
                    row.get("birth_date_1c"),
                    row.get("citizenship"),
                    row.get("passport_series_1c"),
                    row.get("passport_number_1c"),
                    row.get("actual_position"),
                    row.get("section"),
                    row.get("visa"),
                    row.get("visa_type"),
                    row.get("region2"),
                    row.get("visa_expiry"),
                    row.get("entry_date"),
                    row.get("category"),
                    batch_id,
                    now,
                ),
            )
        stored_path = _save_upload_file(file_path, location_id, tracking_date)
        conn.execute(
            """
            INSERT INTO daily_tracking_uploads (
                location_id, tracking_date, original_name, row_count, file_path, created_at
            )
            VALUES (?,?,?,?,?,?)
            """,
            (location_id, tracking_date, original_name, len(parsed), stored_path, now),
        )
        conn.commit()
        return {
            "success": True,
            "rowCount": len(parsed),
            "uploadBatchId": batch_id,
            "trackingDate": tracking_date,
            "locationId": location_id,
            "storedFile": stored_path,
        }
    finally:
        conn.close()
