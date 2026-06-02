import os
import json
import sqlite3
import re
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from rapidfuzz import fuzz, process

import calendar_db
import excel_handler
import main_db
import tickets_db
from data_paths import UPLOAD_DIR

CALENDAR_MERGED_DB_PATH = os.path.join(UPLOAD_DIR, "calendar_merged_db.sqlite")
CALENDAR_MERGED_META_PATH = os.path.join(UPLOAD_DIR, "calendar_merged_meta.json")

_merged_cache: Dict[str, Any] = {"loaded": False}

# Fuzzy passport match (0–100), like Excel Fuzzy Lookup
FUZZY_PASSPORT_SCORE_CUTOFF = 86
FUZZY_FIO_SCORE_CUTOFF = 86
HIGHLIGHT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

MATCH_EXACT = "Точное"
MATCH_NUM_D = "По номеру (столбец D)"
MATCH_FIO_EXACT = "Точное по ФИО"
MATCH_FIO_FUZZY = "Нечёткое по ФИО (Fuzzy)"
MATCH_FUZZY = "Нечёткое по паспорту (Fuzzy)"
MATCH_NONE = "Не найдено"

# Latin → Russian (translit for names in ticket column I)
_LATIN_DIGRAPHS: List[Tuple[str, str]] = [
    ("shch", "щ"),
    ("sch", "щ"),
    ("sh", "ш"),
    ("ch", "ч"),
    ("kh", "х"),
    ("zh", "ж"),
    ("ts", "ц"),
    ("yu", "ю"),
    ("ya", "я"),
    ("yo", "ё"),
    ("ye", "е"),
    ("iu", "ю"),
    ("ia", "я"),
    ("ii", "ий"),
    ("iy", "ий"),
    ("ey", "ей"),
    ("ay", "ай"),
    ("oy", "ой"),
    ("je", "е"),
]
_LATIN_CHAR_MAP: Dict[str, str] = {
    "a": "а",
    "b": "б",
    "c": "к",
    "d": "д",
    "e": "е",
    "f": "ф",
    "g": "г",
    "h": "х",
    "i": "и",
    "j": "й",
    "k": "к",
    "l": "л",
    "m": "м",
    "n": "н",
    "o": "о",
    "p": "п",
    "q": "к",
    "r": "р",
    "s": "с",
    "t": "т",
    "u": "у",
    "v": "в",
    "w": "в",
    "x": "кс",
    "y": "ы",
    "z": "з",
}


def load_calendar_by_path(file_path: str) -> Dict[str, Any]:
    if not file_path or not os.path.isfile(file_path):
        return {"error": f"Файл не найден: {file_path}"}
    return calendar_db.load_calendar_db(file_path)


def _normalize_passport(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().upper()
    if not s:
        return ""
    return re.sub(r"[^0-9A-ZА-ЯЁ]", "", s)


def _atomic_replace_file(src_path: str, dest_path: str) -> None:
    """Atomically replace dest_path with src_path (src must exist)."""
    dest_dir = os.path.dirname(dest_path) or "."
    os.makedirs(dest_dir, exist_ok=True)
    backup_path = f"{dest_path}.bak"
    if os.path.exists(backup_path):
        os.remove(backup_path)
    if os.path.exists(dest_path):
        os.replace(dest_path, backup_path)
    try:
        os.replace(src_path, dest_path)
        if os.path.exists(backup_path):
            os.remove(backup_path)
    except Exception:
        if os.path.exists(backup_path) and not os.path.exists(dest_path):
            os.replace(backup_path, dest_path)
        raise


def _save_merged_calendar_to_sqlite(df: pd.DataFrame) -> None:
    excel_handler.ensure_upload_dir()
    tmp_db_path = f"{CALENDAR_MERGED_DB_PATH}.tmp"
    if os.path.exists(tmp_db_path):
        os.remove(tmp_db_path)

    conn = sqlite3.connect(tmp_db_path)
    try:
        df.to_sql("calendar_merged_records", conn, index=False, if_exists="replace")
        conn.execute(
            'CREATE INDEX IF NOT EXISTS idx_merged_direction ON calendar_merged_records(direction)'
        )
        conn.execute(
            'CREATE INDEX IF NOT EXISTS idx_merged_year_month ON calendar_merged_records(year, month)'
        )
        conn.commit()
    finally:
        conn.close()

    try:
        _atomic_replace_file(tmp_db_path, CALENDAR_MERGED_DB_PATH)
    except Exception:
        if os.path.exists(tmp_db_path):
            os.remove(tmp_db_path)
        raise

    meta = {
        "loaded_at": datetime.now().isoformat(),
        "rows": int(len(df)),
        "columns": list(df.columns),
    }
    with open(CALENDAR_MERGED_META_PATH, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    global _merged_cache
    _merged_cache = {
        "loaded": True,
        "loaded_at": meta["loaded_at"],
        "rows": meta["rows"],
        "columns": meta["columns"],
    }


def _load_merged_meta_from_disk() -> bool:
    global _merged_cache
    if _merged_cache.get("loaded"):
        return True
    if os.path.exists(CALENDAR_MERGED_META_PATH) and os.path.exists(CALENDAR_MERGED_DB_PATH):
        try:
            with open(CALENDAR_MERGED_META_PATH, "r", encoding="utf-8") as f:
                meta = json.load(f)
            _merged_cache = {
                "loaded": True,
                "loaded_at": meta["loaded_at"],
                "rows": meta["rows"],
                "columns": meta.get("columns", []),
            }
            return True
        except Exception:
            return False
    return False


def is_merged_calendar_loaded() -> bool:
    if _merged_cache.get("loaded"):
        return True
    return _load_merged_meta_from_disk()


def get_merged_calendar_status() -> Dict[str, Any]:
    if not is_merged_calendar_loaded():
        return {"loaded": False}
    return {
        "loaded": True,
        "loaded_at": _merged_cache["loaded_at"],
        "rows": _merged_cache["rows"],
        "columns": _merged_cache.get("columns", []),
    }


def get_merged_calendar_data(
    direction: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    search: Optional[str] = None,
    offset: int = 0,
    limit: int = 200,
) -> Dict[str, Any]:
    if not is_merged_calendar_loaded():
        return {"error": "Объединенный календарь не построен", "data": [], "total": 0}

    conn = sqlite3.connect(CALENDAR_MERGED_DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        where_parts = []
        params: List[Any] = []

        if direction:
            where_parts.append("direction = ?")
            params.append(direction)
        if year:
            where_parts.append("year = ?")
            params.append(year)
        if month:
            where_parts.append("month = ?")
            params.append(month)
        if search:
            s = f"%{search.lower()}%"
            where_parts.append(
                '(LOWER(COALESCE(full_name, "")) LIKE ? OR '
                'LOWER(COALESCE("Табельный номер (База)", "")) LIKE ? OR '
                'LOWER(COALESCE("ФИО (База)", "")) LIKE ?)'
            )
            params.extend([s, s, s])

        where_clause = f'WHERE {" AND ".join(where_parts)}' if where_parts else ""
        total = conn.execute(
            f"SELECT COUNT(*) FROM calendar_merged_records {where_clause}",
            params,
        ).fetchone()[0]

        rows = conn.execute(
            f'SELECT * FROM calendar_merged_records {where_clause} '
            f"ORDER BY year, month, row_number LIMIT ? OFFSET ?",
            params + [limit, offset],
        ).fetchall()

        data = [{key: row[key] for key in row.keys()} for row in rows]
        return {
            "data": data,
            "total": total,
            "offset": offset,
            "limit": limit,
            "has_more": (offset + limit) < total,
        }
    finally:
        conn.close()


def _normalize_column_name(name: Any) -> str:
    return re.sub(r"[^a-zа-яё0-9]+", "", str(name or "").strip().lower())


def _pick_main_column(columns: List[str], *candidates: str) -> Optional[str]:
    normalized_columns = [(col, _normalize_column_name(col)) for col in columns]
    normalized_candidates = [_normalize_column_name(c) for c in candidates if c]

    for cand in normalized_candidates:
        for col, norm_col in normalized_columns:
            if norm_col == cand:
                return col

    for cand in normalized_candidates:
        for col, norm_col in normalized_columns:
            if cand and cand in norm_col:
                return col

    return None


def _load_main_merge_dataframe() -> Tuple[pd.DataFrame, Dict[str, Optional[str]]]:
    if not main_db.is_loaded():
        raise ValueError("Основная база не загружена")

    cache = getattr(main_db, "_cache", {}) or {}
    columns = list(cache.get("columns") or [])
    col_mapping = dict(cache.get("col_mapping") or {})
    if not columns or not col_mapping:
        raise ValueError("Не удалось получить метаданные Основной базы")

    resolved = {
        "tab_num": _pick_main_column(
            columns,
            "Таб. номер",
            "Табельный номер (с префиксами)",
            "Табельный номер",
        ),
        "fio": _pick_main_column(columns, "ФИО", "Сотрудник"),
        "pass_series": _pick_main_column(columns, "Удостоверение.Серия", "Серия"),
        "pass_number": _pick_main_column(columns, "Удостоверение.Номер", "Номер"),
        "organization": _pick_main_column(columns, "Организация"),
        "department": _pick_main_column(columns, "Подразделение", "Отдел"),
        "state": _pick_main_column(columns, "Состояние"),
    }

    if not resolved["tab_num"] or not resolved["fio"]:
        raise ValueError(
            "В Основной базе не найдены обязательные столбцы «Табельный номер» и «ФИО»"
        )
    if not resolved["pass_number"]:
        raise ValueError("В Основной базе не найден столбец номера удостоверения")

    select_parts = []
    for alias, orig_col in resolved.items():
        if not orig_col:
            continue
        sql_col = col_mapping.get(orig_col)
        if not sql_col:
            continue
        select_parts.append(f'"{sql_col}" AS "{alias}"')

    if not select_parts:
        raise ValueError("Не удалось подготовить выборку из Основной базы")

    conn = main_db._get_db_connection()
    try:
        df = pd.read_sql_query(
            f"SELECT {', '.join(select_parts)} FROM employees",
            conn,
        )
    finally:
        conn.close()

    for optional_col in ("pass_series", "organization", "department", "state"):
        if optional_col not in df.columns:
            df[optional_col] = ""

    return df, resolved


def merge_calendar_with_main_db(output_name: Optional[str] = None) -> Dict[str, Any]:
    if not calendar_db.is_loaded():
        return {
            "error": "Календарь не загружен. Сначала загрузите файл календаря.",
        }
    if not main_db.is_loaded():
        return {
            "error": "Основная база не загружена. Сначала загрузите файл Основной базы.",
        }

    try:
        main_df, _resolved = _load_main_merge_dataframe()
    except ValueError as e:
        return {"error": str(e)}

    if main_df.empty:
        return {"error": "Основная база пуста"}

    cal_conn = sqlite3.connect(calendar_db.CALENDAR_DB_PATH)
    try:
        calendar_df = pd.read_sql_query(
            "SELECT * FROM calendar_records ORDER BY year, month, row_number",
            cal_conn,
        )
    finally:
        cal_conn.close()

    if calendar_df.empty:
        return {"error": "Календарь не содержит записей"}

    main_df["passport_key"] = (
        main_df["pass_series"].fillna("").astype(str)
        + main_df["pass_number"].fillna("").astype(str)
    ).map(_normalize_passport)
    main_lookup = (
        main_df[main_df["passport_key"] != ""]
        .drop_duplicates(subset=["passport_key"])
        .loc[:, ["passport_key", "tab_num", "fio", "organization", "department", "state"]]
        .rename(
            columns={
                "tab_num": "Табельный номер (База)",
                "fio": "ФИО (База)",
                "organization": "Организация (База)",
                "department": "Подразделение (База)",
                "state": "Состояние (База)",
            }
        )
    )

    merged = calendar_df.copy()
    merged["passport_key"] = (
        merged["passport_series"].fillna("").astype(str)
        + merged["passport_number"].fillna("").astype(str)
    ).map(_normalize_passport)
    merged = merged.merge(main_lookup, on="passport_key", how="left")
    merged = merged.drop(columns=["passport_key"])

    excel_handler.ensure_upload_dir()
    base_name = (
        output_name.strip()
        if output_name and output_name.strip()
        else f"calendar_with_main_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    if not base_name.lower().endswith(".xlsx"):
        base_name = f"{base_name}.xlsx"
    stored_filename = f"{excel_handler.generate_file_id()}_{base_name}"
    out_path = os.path.join(UPLOAD_DIR, stored_filename)

    merged.to_excel(out_path, index=False, engine="openpyxl")
    _save_merged_calendar_to_sqlite(merged)

    matched_rows = int(merged["Табельный номер (База)"].notna().sum())
    total_rows = int(len(merged))
    return {
        "success": True,
        "file_path": out_path,
        "stored_filename": stored_filename,
        "file_id": os.path.splitext(stored_filename)[0],
        "rows": total_rows,
        "matched_rows": matched_rows,
        "unmatched_rows": total_rows - matched_rows,
        "columns": list(merged.columns),
    }


def _passport_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def _build_passport_indexes(
    main_map: Dict[str, Dict[str, Any]],
) -> Tuple[Dict[int, List[str]], Dict[str, List[str]]]:
    """Index main DB passport keys by length and last-6 digits for faster fuzzy lookup."""
    by_length: Dict[int, List[str]] = defaultdict(list)
    by_suffix: Dict[str, List[str]] = defaultdict(list)
    for key in main_map:
        by_length[len(key)].append(key)
        digits = _passport_digits(key)
        if len(digits) >= 6:
            by_suffix[digits[-6:]].append(key)
    return dict(by_length), dict(by_suffix)


def _fuzzy_match_passport(
    query: str,
    main_map: Dict[str, Dict[str, Any]],
    by_length: Dict[int, List[str]],
    by_suffix: Dict[str, List[str]],
    score_cutoff: int = FUZZY_PASSPORT_SCORE_CUTOFF,
) -> Tuple[Optional[Dict[str, Any]], int]:
    """
    Find best passport match when exact key lookup failed.
    Returns (employee data, score 0-100) or (None, 0).
    """
    if not query or len(query) < 4:
        return None, 0

    candidates: List[str] = []
    qlen = len(query)

    # Same-length bucket ±3 chars (OCR / missing leading zero)
    for length in range(max(4, qlen - 3), qlen + 4):
        candidates.extend(by_length.get(length, []))

    # Same last 6 digits (common when series formatting differs)
    qdigits = _passport_digits(query)
    if len(qdigits) >= 6:
        candidates.extend(by_suffix.get(qdigits[-6:], []))

    # Substring: ticket passport contained in main key or reverse
    for key in main_map:
        if len(key) >= 6 and (query in key or key in query):
            candidates.append(key)

    seen = set()
    unique: List[str] = []
    for key in candidates:
        if key not in seen:
            seen.add(key)
            unique.append(key)

    if not unique:
        unique = list(main_map.keys())

    if len(unique) > 8000:
        unique = unique[:8000]

    result = process.extractOne(
        query,
        unique,
        scorer=fuzz.ratio,
        score_cutoff=score_cutoff,
    )
    if not result:
        return None, 0

    matched_key, score, _ = result
    return main_map.get(matched_key), int(score)


def _build_column_d_maps(
    main_df: pd.DataFrame,
) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]], Dict[str, List[str]]]:
    """
    Maps for matching ticket column J only against Main DB column D (Удостоверение.Номер),
    without series (column C).
    """
    main_num_map: Dict[str, Dict[str, Any]] = {}
    main_digits_map: Dict[str, Dict[str, Any]] = {}
    by_digit_suffix: Dict[str, List[str]] = defaultdict(list)

    for _, row in main_df.iterrows():
        pass_number = row.get("pass_number")
        if pass_number is None or (isinstance(pass_number, float) and pd.isna(pass_number)):
            continue
        data = {"tab_num": row.get("tab_num"), "fio": row.get("fio")}
        num_norm = _normalize_passport(pass_number)
        digits = _passport_digits(str(pass_number))

        if num_norm and num_norm not in main_num_map:
            main_num_map[num_norm] = data
        if digits and digits not in main_digits_map:
            main_digits_map[digits] = data
            if len(digits) >= 6:
                by_digit_suffix[digits[-6:]].append(digits)

    return main_num_map, main_digits_map, dict(by_digit_suffix)


def _match_by_column_d_only(
    raw: Any,
    main_num_map: Dict[str, Dict[str, Any]],
    main_digits_map: Dict[str, Dict[str, Any]],
    by_digit_suffix: Dict[str, List[str]],
) -> Optional[Dict[str, Any]]:
    """Match ticket passport (J) against Main DB column D only (no C+D)."""
    norm_j = _normalize_passport(raw)
    if not norm_j:
        return None

    if norm_j in main_num_map:
        return main_num_map[norm_j]

    digits_j = _passport_digits(norm_j)
    if digits_j and digits_j in main_digits_map:
        return main_digits_map[digits_j]

    # J often has series+number; D is number only — compare by trailing digits
    if len(digits_j) >= 6:
        for d_key in by_digit_suffix.get(digits_j[-6:], []):
            if digits_j.endswith(d_key) or d_key.endswith(digits_j):
                return main_digits_map.get(d_key)

    return None


def _has_latin_letters(text: str) -> bool:
    return bool(re.search(r"[A-Za-z]", text))


def _translit_latin_to_russian(text: str) -> str:
    """Transliterate Latin letters in a name to Cyrillic (e.g. IVAN -> ИВАН)."""
    s = text.strip().lower()
    if not s:
        return s
    for lat, cyr in _LATIN_DIGRAPHS:
        s = s.replace(lat, cyr)
    out: List[str] = []
    for ch in s:
        if "a" <= ch <= "z":
            out.append(_LATIN_CHAR_MAP.get(ch, ch))
        else:
            out.append(ch)
    return "".join(out)


def _normalize_fio(value: Any) -> str:
    """Normalize FIO for comparison; transliterate Latin to Russian first."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s)
    if _has_latin_letters(s):
        s = _translit_latin_to_russian(s)
    s = s.upper()
    s = re.sub(r"[^А-ЯЁ\s\-]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _detect_ticket_fio_column(columns: List[str]) -> Optional[str]:
    for c in columns:
        cl = str(c).lower().replace(" ", "")
        if "ф.и.о" in cl or "фио" in cl or "f.i.o" in cl or "fio" in cl:
            return c
    if len(columns) >= 9:
        return columns[8]
    return None


def _build_fio_indexes(
    main_df: pd.DataFrame,
) -> Tuple[
    Dict[str, Dict[str, Any]],
    Dict[str, List[str]],
    Dict[str, List[str]],
    Dict[str, List[str]],
]:
    """Exact map + buckets (1-й символ, префикс 3, фамилия) для fuzzy FIO."""
    main_fio_map: Dict[str, Dict[str, Any]] = {}
    by_first_char: Dict[str, List[str]] = defaultdict(list)
    by_prefix3: Dict[str, List[str]] = defaultdict(list)
    by_surname: Dict[str, List[str]] = defaultdict(list)

    tab_key = "tab_num" if "tab_num" in main_df.columns else "tab"
    for row in main_df.itertuples(index=False):
        fio_val = getattr(row, "fio", None)
        norm = _normalize_fio(fio_val)
        if not norm or len(norm) < 3:
            continue
        if norm in main_fio_map:
            continue
        tab_num = getattr(row, tab_key, None)
        main_fio_map[norm] = {"tab_num": tab_num, "fio": fio_val}
        by_first_char[norm[0]].append(norm)
        if len(norm) >= 3:
            by_prefix3[norm[:3]].append(norm)
        parts = norm.split()
        if parts and len(parts[0]) >= 2:
            by_surname[parts[0]].append(norm)

    return main_fio_map, dict(by_first_char), dict(by_prefix3), dict(by_surname)


def _fuzzy_match_fio(
    query_raw: Any,
    main_fio_map: Dict[str, Dict[str, Any]],
    by_first_char: Dict[str, List[str]],
    score_cutoff: int = FUZZY_FIO_SCORE_CUTOFF,
    by_prefix3: Optional[Dict[str, List[str]]] = None,
    by_surname: Optional[Dict[str, List[str]]] = None,
) -> Tuple[Optional[Dict[str, Any]], int]:
    """
    Fuzzy match ticket column I (Ф.И.О.) against Main DB column B (ФИО).
    """
    query = _normalize_fio(query_raw)
    if not query or len(query) < 4:
        return None, 0

    if query in main_fio_map:
        return main_fio_map[query], 100

    seen = set()
    candidates: List[str] = []

    def _add(keys: List[str], limit: int = 400) -> None:
        for key in keys:
            if key in seen:
                continue
            seen.add(key)
            candidates.append(key)
            if len(candidates) >= limit:
                return

    _add(by_first_char.get(query[0], []))
    if len(query) >= 2 and query[1] != query[0]:
        _add(by_first_char.get(query[1], []))
    if by_prefix3 and len(query) >= 3:
        _add(by_prefix3.get(query[:3], []))
    parts = query.split()
    if by_surname and parts:
        _add(by_surname.get(parts[0], []), limit=350)

    if not candidates:
        return None, 0

    qlen = len(query)
    filtered = [c for c in candidates if abs(len(c) - qlen) <= max(6, qlen // 2)]
    if not filtered:
        filtered = candidates

    if len(filtered) > 500:
        filtered = filtered[:500]

    result = process.extractOne(
        query,
        filtered,
        scorer=fuzz.token_sort_ratio,
        score_cutoff=score_cutoff,
    )
    if not result:
        return None, 0

    matched_key, score, _ = result
    return main_fio_map.get(matched_key), int(score)


def _write_tickets_excel_highlighted(
    df: pd.DataFrame,
    out_path: str,
    sheet_name: str,
    match_status_col: str = "Сопоставление",
) -> None:
    """Write tickets merge result; highlight fuzzy / not-found rows in light yellow."""
    sheet = (sheet_name or "Данные")[:31]
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet, index=False)

    if match_status_col not in df.columns:
        return

    highlight_mask = df[match_status_col].isin({MATCH_FIO_FUZZY, MATCH_FUZZY, MATCH_NONE})
    if not highlight_mask.any():
        return

    wb = load_workbook(out_path)
    ws = wb[sheet]
    n_cols = len(df.columns)
    # On very large files, highlight key columns only (much faster than every cell).
    cols_to_fill = n_cols if len(df) <= 50000 else min(n_cols, 6)

    for i in df.index[highlight_mask]:
        row_num = int(i) + 2
        for col in range(1, cols_to_fill + 1):
            ws.cell(row=row_num, column=col).fill = HIGHLIGHT_FILL

    wb.save(out_path)


def merge_tickets_with_main_db(
    ticket_file_path: Optional[str] = None,
    output_name: Optional[str] = None,
    sheet_name: Optional[str] = None,
    passport_column: Optional[str] = None,
    use_registry: bool = False,
    registry: str = "vsm",
) -> Dict[str, Any]:
    if not main_db.is_loaded():
        return {
            "error": "Основная база не загружена. Сначала загрузите файл Основной базы.",
        }

    try:
        main_df, _resolved = _load_main_merge_dataframe()
    except ValueError as e:
        return {"error": str(e)}

    main_df["passport_key"] = (
        main_df["pass_series"].fillna("").astype(str)
        + main_df["pass_number"].fillna("").astype(str)
    ).map(_normalize_passport)
    main_map = (
        main_df[main_df["passport_key"] != ""]
        .drop_duplicates(subset=["passport_key"])
        .set_index("passport_key")[["tab_num", "fio"]]
        .to_dict(orient="index")
    )
    main_num_map, main_digits_map, by_digit_suffix = _build_column_d_maps(main_df)

    used_registry = False
    if use_registry or not ticket_file_path:
        try:
            if not tickets_db.is_loaded(registry):
                return {
                    "error": "Реестр билетов не загружен. Сначала нажмите «Загрузить Реестр по Билетам».",
                }
            df, detected_passport, target_sheet = tickets_db.read_registry_dataframe(registry)
        except ValueError as e:
            return {"error": str(e)}
        used_registry = True
    else:
        if not os.path.exists(ticket_file_path):
            return {"error": f"Файл не найден: {ticket_file_path}"}
        xls = pd.ExcelFile(ticket_file_path, engine="openpyxl")
        try:
            target_sheet = sheet_name or (xls.sheet_names[0] if xls.sheet_names else None)
        finally:
            xls.close()
        if not target_sheet:
            return {"error": "Не удалось определить лист отчета"}
        df = pd.read_excel(
            ticket_file_path,
            sheet_name=target_sheet,
            header=0,
            dtype=object,
            engine="openpyxl",
        )
        detected_passport = None

    if df.empty:
        return {"error": "Данные реестра билетов пустые"}

    src_col: Optional[str] = None
    if passport_column and passport_column in df.columns:
        src_col = passport_column
    elif detected_passport and detected_passport in df.columns:
        src_col = detected_passport
    else:
        for c in df.columns:
            if "паспорт" in str(c).lower():
                src_col = c
                break
        if src_col is None:
            if len(df.columns) >= 10:
                src_col = df.columns[9]
            else:
                return {"error": "В отчете не найден столбец ПАСПОРТ (и нет колонки J)"}

    fio_col_ticket = _detect_ticket_fio_column(list(df.columns))
    main_fio_map, fio_by_first_char, fio_by_prefix3, fio_by_surname = _build_fio_indexes(
        main_df
    )
    by_length, by_suffix = _build_passport_indexes(main_map)

    n = len(df)
    tab_nums: List[Any] = [None] * n
    fios: List[Any] = [None] * n
    match_statuses: List[str] = [MATCH_NONE] * n
    match_scores: List[Any] = [None] * n
    pending: List[int] = []

    src_values = df[src_col].tolist()
    fio_src_values = (
        df[fio_col_ticket].tolist() if fio_col_ticket and fio_col_ticket in df.columns else [None] * n
    )
    for idx, raw in enumerate(src_values):
        key = _normalize_passport(raw)
        if not key:
            continue
        exact = main_map.get(key)
        if exact:
            tab_nums[idx] = exact.get("tab_num")
            fios[idx] = exact.get("fio")
            match_statuses[idx] = MATCH_EXACT
            match_scores[idx] = 100
        else:
            pending.append(idx)

    for idx in pending:
        raw = src_values[idx]
        key = _normalize_passport(raw)
        d_hit = _match_by_column_d_only(
            raw, main_num_map, main_digits_map, by_digit_suffix
        )
        if d_hit:
            tab_nums[idx] = d_hit.get("tab_num")
            fios[idx] = d_hit.get("fio")
            match_statuses[idx] = MATCH_NUM_D
            match_scores[idx] = 100
            continue

        fio_raw = fio_src_values[idx]
        if fio_raw is not None and str(fio_raw).strip():
            fio_norm = _normalize_fio(fio_raw)
            if fio_norm and fio_norm in main_fio_map:
                hit = main_fio_map[fio_norm]
                tab_nums[idx] = hit.get("tab_num")
                fios[idx] = hit.get("fio")
                match_statuses[idx] = MATCH_FIO_EXACT
                match_scores[idx] = 100
                continue

            fio_hit, fio_score = _fuzzy_match_fio(
                fio_raw,
                main_fio_map,
                fio_by_first_char,
                by_prefix3=fio_by_prefix3,
                by_surname=fio_by_surname,
            )
            if fio_hit:
                tab_nums[idx] = fio_hit.get("tab_num")
                fios[idx] = fio_hit.get("fio")
                match_statuses[idx] = MATCH_FIO_FUZZY
                match_scores[idx] = fio_score
                continue

        fuzzy_hit, fuzzy_score = _fuzzy_match_passport(
            key, main_map, by_length, by_suffix
        )
        if fuzzy_hit:
            tab_nums[idx] = fuzzy_hit.get("tab_num")
            fios[idx] = fuzzy_hit.get("fio")
            match_statuses[idx] = MATCH_FUZZY
            match_scores[idx] = fuzzy_score

    df["Табельный номер (с префиксами)"] = tab_nums
    df["ФИО"] = fios
    df["Сопоставление"] = match_statuses
    df["Схожесть %"] = match_scores

    front_cols = [
        "Табельный номер (с префиксами)",
        "ФИО",
        "Сопоставление",
        "Схожесть %",
    ]
    rest_cols = [c for c in df.columns if c not in front_cols]
    df = df[front_cols + rest_cols]

    total = int(len(df))
    matched_exact = int((df["Сопоставление"] == MATCH_EXACT).sum())
    matched_by_d = int((df["Сопоставление"] == MATCH_NUM_D).sum())
    matched_fio_exact = int((df["Сопоставление"] == MATCH_FIO_EXACT).sum())
    matched_fio_fuzzy = int((df["Сопоставление"] == MATCH_FIO_FUZZY).sum())
    matched_fuzzy = int((df["Сопоставление"] == MATCH_FUZZY).sum())
    matched_rows = (
        matched_exact + matched_by_d + matched_fio_exact + matched_fio_fuzzy + matched_fuzzy
    )
    unmatched_rows = int((df["Сопоставление"] == MATCH_NONE).sum())
    highlight_rows = matched_fio_fuzzy + matched_fuzzy + unmatched_rows

    excel_handler.ensure_upload_dir()
    base_name = (
        output_name.strip()
        if output_name and output_name.strip()
        else f"tickets_with_main_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    if not base_name.lower().endswith(".xlsx"):
        base_name = f"{base_name}.xlsx"
    stored_filename = f"{excel_handler.generate_file_id()}_{base_name}"
    out_path = os.path.join(excel_handler.UPLOAD_DIR, stored_filename)

    _write_tickets_excel_highlighted(df, out_path, str(target_sheet or "Данные"))

    return {
        "success": True,
        "file_path": out_path,
        "stored_filename": stored_filename,
        "file_id": os.path.splitext(stored_filename)[0],
        "sheet_name": target_sheet,
        "rows": total,
        "matched_rows": matched_rows,
        "matched_exact": matched_exact,
        "matched_by_d": matched_by_d,
        "matched_fio_exact": matched_fio_exact,
        "matched_fio_fuzzy": matched_fio_fuzzy,
        "matched_fuzzy": matched_fuzzy,
        "unmatched_rows": unmatched_rows,
        "highlight_rows": highlight_rows,
        "passport_source_column": str(src_col),
        "fio_source_column": str(fio_col_ticket) if fio_col_ticket else None,
        "fuzzy_cutoff_percent": FUZZY_PASSPORT_SCORE_CUTOFF,
        "fuzzy_fio_cutoff_percent": FUZZY_FIO_SCORE_CUTOFF,
        "used_registry": used_registry,
        "registry": registry if used_registry else None,
    }
