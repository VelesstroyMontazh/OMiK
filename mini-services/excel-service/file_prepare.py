"""
Prepare Excel workbook: unhide sheets/columns, remove filters, replace formulas with values.
"""

import os
import shutil
from datetime import datetime
from typing import Any, Dict, Optional

import openpyxl
from openpyxl.utils import get_column_letter

import excel_handler
import excel_libs
from excel_libs import (
    HAS_FORMULAS_PKG,
    count_formulas_pycel,
    recalculate_workbook_formulas,
)

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def _atomic_replace_file(src_path: str, dest_path: str) -> None:
    dest_dir = os.path.dirname(dest_path) or "."
    os.makedirs(dest_dir, exist_ok=True)
    backup_path = f"{dest_path}.bak"
    if os.path.exists(backup_path):
        os.remove(backup_path)
    if os.path.exists(dest_path):
        os.replace(dest_path, backup_path)
    try:
        os.replace(src_path, dest_path)
        if os.path.exists(backup_path):
            os.remove(backup_path)
    except Exception:
        if os.path.exists(backup_path) and not os.path.exists(dest_path):
            os.replace(backup_path, dest_path)
        raise


def _cell_has_formula(cell) -> bool:
    if getattr(cell, "data_type", None) == "f":
        return True
    value = cell.value
    return isinstance(value, str) and value.startswith("=")


def _unhide_sheet_dimensions(ws) -> int:
    """Unhide hidden row/column dimensions (only defined entries, not full 1M grid)."""
    changed = 0
    for _key, dim in ws.column_dimensions.items():
        if dim.hidden:
            dim.hidden = False
            changed += 1
    for _key, dim in ws.row_dimensions.items():
        if dim.hidden:
            dim.hidden = False
            changed += 1
    return changed


def _clear_filters(ws) -> bool:
    cleared = False
    if ws.auto_filter and ws.auto_filter.ref:
        ws.auto_filter.ref = None
        cleared = True
    return cleared


def _replace_formulas_with_values(ws, ws_values) -> int:
    replaced = 0
    for row in ws.iter_rows():
        for cell in row:
            if not _cell_has_formula(cell):
                continue
            value = ws_values.cell(row=cell.row, column=cell.column).value
            cell.value = value
            replaced += 1
    return replaced


def prepare_excel_file(
    file_path: str,
    output_name: Optional[str] = None,
    save_in_place: bool = False,
) -> Dict[str, Any]:
    """
    Process Excel file:
    - all sheets visible
    - unhide columns/rows
    - remove auto-filters on each sheet
    - replace formulas with cached values on all sheets
    """
    if not file_path:
        return {"error": "Не указан путь к файлу"}
    if not os.path.exists(file_path):
        return {"error": f"Файл не найден: {file_path}"}
    if not os.path.isfile(file_path):
        return {"error": f"Это не файл: {file_path}"}

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        return {
            "error": f"Формат {ext} не поддерживается. Используйте .xlsx или .xlsm",
        }

    try:
        keep_vba = ext == ".xlsm"
        wb = openpyxl.load_workbook(
            file_path,
            data_only=False,
            keep_vba=keep_vba,
        )
        wb_values = openpyxl.load_workbook(
            file_path,
            data_only=True,
            read_only=True,
            keep_vba=False,
        )
    except Exception as e:
        return {"error": f"Не удалось открыть файл: {e}"}

    stats: Dict[str, Any] = {
        "sheets_processed": 0,
        "sheets_made_visible": 0,
        "filters_removed": 0,
        "dimensions_unhidden": 0,
        "formulas_replaced": 0,
        "sheet_details": [],
        "formula_engines": [],
    }

    if HAS_FORMULAS_PKG:
        stats["formula_engines"].append("formulas")
    pycel_info = count_formulas_pycel(file_path)
    if pycel_info.get("available"):
        stats["pycel_formula_cells"] = pycel_info.get("formula_cell_count")
        stats["formula_engines"].append("pycel")

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_stat: Dict[str, Any] = {"name": sheet_name}

            if ws.sheet_state != "visible":
                ws.sheet_state = "visible"
                stats["sheets_made_visible"] += 1
                sheet_stat["was_hidden"] = True

            if _clear_filters(ws):
                stats["filters_removed"] += 1
                sheet_stat["filter_removed"] = True

            unhidden = _unhide_sheet_dimensions(ws)
            stats["dimensions_unhidden"] += unhidden
            sheet_stat["dimensions_unhidden"] = unhidden

            ws_values = wb_values[sheet_name]
            replaced = _replace_formulas_with_values(ws, ws_values)
            stats["formulas_replaced"] += replaced
            sheet_stat["formulas_replaced"] = replaced

            stats["sheets_processed"] += 1
            if len(stats["sheet_details"]) < 50:
                stats["sheet_details"].append(sheet_stat)

        excel_handler.ensure_upload_dir()
        base = os.path.basename(file_path)
        if output_name and str(output_name).strip():
            base = str(output_name).strip()
        if not base.lower().endswith(ext):
            base = f"{os.path.splitext(base)[0]}{ext}"

        stored_filename = f"{excel_handler.generate_file_id()}_{base}"
        out_path = os.path.join(excel_handler.UPLOAD_DIR, stored_filename)
        wb.save(out_path)
        wb.close()
        wb_values.close()

        if stats["formulas_replaced"] == 0 and HAS_FORMULAS_PKG:
            recalc = recalculate_workbook_formulas(file_path)
            if recalc.get("success") and recalc.get("output_files"):
                try:
                    shutil.copy2(recalc["output_files"][0], out_path)
                    stats["formulas_recalculated"] = True
                    stats["formulas_engine"] = "formulas"
                except Exception:
                    stats["formulas_recalculated"] = False
            else:
                stats["formulas_recalculated"] = False
                if recalc.get("error"):
                    stats["formulas_recalc_error"] = recalc["error"]

        if save_in_place:
            _atomic_replace_file(out_path, file_path)
            out_path = file_path
            stored_filename = base
            file_id = os.path.splitext(base)[0]
        else:
            file_id = os.path.splitext(stored_filename)[0]

        return {
            "success": True,
            "file_path": out_path,
            "stored_filename": stored_filename,
            "file_id": file_id,
            "source_file": file_path,
            "saved_in_place": save_in_place,
            "processed_at": datetime.now().isoformat(),
            **stats,
        }
    except Exception as e:
        try:
            wb_values.close()
        except Exception:
            pass
        return {"error": f"Ошибка обработки файла: {e}"}
