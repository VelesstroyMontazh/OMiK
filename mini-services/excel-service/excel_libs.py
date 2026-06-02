"""
Мультибиблиотечный слой Excel: выбор движка по задаче и формату файла.
Polars, pyexcelerate, xlsxwriter, xlutils, formulas; опционально pycel и xlwings.
Чтение/превью — напрямую через pandas/openpyxl/xlrd/pyxlsb (без pyexcel/pylightxl).
"""

from __future__ import annotations

import os
import shutil
import tempfile
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

# --- Обязательные / основные ---
import openpyxl
import xlrd
import xlsxwriter

# --- Опциональные движки ---
try:
    import polars as pl

    HAS_POLARS = True
except ImportError:
    pl = None  # type: ignore
    HAS_POLARS = False

try:
    import pyxlsb

    HAS_PYXLSB = True
except ImportError:
    pyxlsb = None  # type: ignore
    HAS_PYXLSB = False

try:
    import xlwt

    HAS_XLWT = True
except ImportError:
    HAS_XLWT = False

try:
    from pyexcelerate import Workbook as PyExWorkbook

    HAS_PYEXCELERATE = True
except ImportError:
    PyExWorkbook = None  # type: ignore
    HAS_PYEXCELERATE = False

try:
    from xlutils.copy import copy as xlutils_copy

    HAS_XLUTILS = True
except ImportError:
    xlutils_copy = None  # type: ignore
    HAS_XLUTILS = False

try:
    import formulas as formulas_pkg

    HAS_FORMULAS_PKG = True
except ImportError:
    formulas_pkg = None  # type: ignore
    HAS_FORMULAS_PKG = False

try:
    import pycel

    HAS_PYCEL = True
except ImportError:
    pycel = None  # type: ignore
    HAS_PYCEL = False

try:
    import xlwings as xw

    HAS_XLWINGS = True
except ImportError:
    xw = None  # type: ignore
    HAS_XLWINGS = False

# PyXLL — только на машине с установленным Excel-addin, не pip-пакет для сервера
HAS_PYXLL = False

POLARS_ROW_THRESHOLD = 8_000
PYEXCELERATE_ROW_THRESHOLD = 15_000
PREVIEW_MAX_ROWS = 5_000


def get_library_status() -> Dict[str, Any]:
    """Статус библиотек для /api/health и UI."""
    return {
        "core": {
            "pandas": True,
            "numpy": True,
            "openpyxl": True,
            "xlrd": True,
            "xlsxwriter": True,
        },
        "optional": {
            "polars": HAS_POLARS,
            "pyxlsb": HAS_PYXLSB,
            "xlwt": HAS_XLWT,
            "pyexcelerate": HAS_PYEXCELERATE,
            "xlutils": HAS_XLUTILS,
            "formulas": HAS_FORMULAS_PKG,
            "pycel": HAS_PYCEL,
            "xlwings": HAS_XLWINGS,
            "pyxll": HAS_PYXLL,
        },
        "routing": {
            "large_read": "polars" if HAS_POLARS else "pandas",
            "large_write": (
                "pyexcelerate"
                if HAS_PYEXCELERATE
                else ("xlsxwriter" if True else "openpyxl")
            ),
            "format_convert": "pandas",
            "fast_preview_xlsx": f"pandas_nrows_{PREVIEW_MAX_ROWS}",
            "formula_recalc": "formulas" if HAS_FORMULAS_PKG else "openpyxl_cached_only",
            "xls_edit_copy": "xlutils" if HAS_XLUTILS else None,
            "excel_desktop": "xlwings" if HAS_XLWINGS else None,
        },
    }


def _ext(path: str) -> str:
    return os.path.splitext(path)[1].lower()


def _estimate_rows(file_path: str, sheet_name: str) -> int:
    ext = _ext(file_path)
    try:
        if ext in (".xlsx", ".xlsm"):
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return 0
            ws = wb[sheet_name]
            rows = ws.max_row or 0
            wb.close()
            return int(rows)
        if ext == ".xls":
            book = xlrd.open_workbook(file_path)
            idx = book.sheet_names().index(sheet_name)
            return book.sheet_by_index(idx).nrows
        if ext == ".csv":
            return sum(1 for _ in open(file_path, encoding="utf-8", errors="replace")) - 1
    except Exception:
        return 0
    return 0


def read_dataframe(
    file_path: str,
    sheet_name: Optional[str] = None,
    header: int = 0,
) -> pd.DataFrame:
    """
    Универсальное чтение в DataFrame.
    Большие .xlsx — Polars (если установлен), иначе Pandas.
    """
    ext = _ext(file_path)
    sheet = sheet_name or 0
    est = _estimate_rows(file_path, sheet if isinstance(sheet, str) else "Sheet1")

    if HAS_POLARS and ext in (".xlsx", ".xlsm") and est >= POLARS_ROW_THRESHOLD and header == 0:
        try:
            kwargs: Dict[str, Any] = {
                "engine": "openpyxl",
                "raise_if_empty": False,
            }
            if isinstance(sheet, str):
                kwargs["sheet_name"] = sheet
            else:
                kwargs["sheet_id"] = int(sheet)
            pl_df = pl.read_excel(file_path, **kwargs)  # type: ignore
            if isinstance(pl_df, dict):
                pl_df = next(iter(pl_df.values()))
            pdf = pl_df.to_pandas()
            # Polars can collapse wide sheets with merged title rows into a single column.
            if pdf.shape[1] <= 1:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                try:
                    ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[int(sheet)]
                    max_col = ws.max_column or 0
                finally:
                    wb.close()
                if max_col > 1:
                    raise ValueError("polars returned a single column for a wide sheet")
            return pdf
        except Exception:
            pass

    if ext in (".xlsx", ".xlsm"):
        return pd.read_excel(file_path, sheet_name=sheet, header=header, engine="openpyxl")
    if ext == ".xls":
        return pd.read_excel(file_path, sheet_name=sheet, header=header, engine="xlrd")
    if ext == ".xlsb":
        return pd.read_excel(file_path, sheet_name=sheet, header=header, engine="pyxlsb")
    if ext == ".csv":
        return pd.read_csv(file_path, header=header)
    raise ValueError(f"Unsupported format for read_dataframe: {ext}")


def get_sheet_names_universal(file_path: str) -> List[str]:
    """Имена листов через нативные движки по расширению файла."""
    ext = _ext(file_path)
    if ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    if ext == ".xls":
        return xlrd.open_workbook(file_path).sheet_names()
    if ext == ".xlsb" and HAS_PYXLSB:
        with pyxlsb.open_workbook(file_path) as wb:
            return list(wb.sheets)
    if ext == ".csv":
        return ["Sheet1"]
    return []


def read_sheet_preview_pandas(
    file_path: str,
    sheet_name: str,
    max_rows: int,
) -> Optional[Dict[str, Any]]:
    """Превью .xlsx/.xlsm: первые N строк через pandas (nrows), без pylightxl."""
    ext = _ext(file_path)
    if ext not in (".xlsx", ".xlsm"):
        return None
    nrows = min(max_rows, PREVIEW_MAX_ROWS)
    try:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            nrows=nrows,
            header=None,
            engine="openpyxl",
        )
        grid = df.where(pd.notna(df), None).values.tolist()
        return grid_to_sheet_payload(grid, sheet_name, max_rows, read_engine="pandas")
    except Exception:
        return None


def polars_sort_dataframe(df: pd.DataFrame, column: str, ascending: bool) -> pd.DataFrame:
    if not HAS_POLARS:
        return df.sort_values(by=column, ascending=ascending, na_position="last")
    pl_df = pl.from_pandas(df)  # type: ignore
    sorted_pl = pl_df.sort(column, descending=not ascending, nulls_last=True)
    return sorted_pl.to_pandas()


def polars_filter_dataframe(df: pd.DataFrame, column: str, mask: pd.Series) -> pd.DataFrame:
    if not HAS_POLARS:
        return df[mask]
    pl_df = pl.from_pandas(df)  # type: ignore
    pl_mask = pl.from_pandas(mask.to_frame("_m"))["_m"]  # type: ignore
    return pl_df.filter(pl_mask).to_pandas()


def polars_analyze_numeric(df: pd.DataFrame, operations: List[str]) -> Dict[str, Any]:
    if not HAS_POLARS or df.empty:
        return {}
    numeric = df.select_dtypes(include=[np.number])
    if numeric.empty:
        return {}
    pl_df = pl.from_pandas(numeric)  # type: ignore
    results: Dict[str, Any] = {}
    def _scalar_dict(frame_result) -> Dict[str, Any]:
        raw = frame_result.to_dict(as_series=False)
        out: Dict[str, Any] = {}
        for col, vals in raw.items():
            if isinstance(vals, list) and len(vals) == 1:
                out[col] = vals[0]
            else:
                out[col] = vals
        return out

    for op in operations:
        op_l = op.lower()
        if op_l in ("sum",):
            results[op_l] = _scalar_dict(pl_df.sum())
        elif op_l in ("avg", "mean"):
            results[op_l] = _scalar_dict(pl_df.mean())
        elif op_l in ("min",):
            results[op_l] = _scalar_dict(pl_df.min())
        elif op_l in ("max",):
            results[op_l] = _scalar_dict(pl_df.max())
        elif op_l in ("std",):
            results[op_l] = _scalar_dict(pl_df.std())
        elif op_l in ("median",):
            results[op_l] = _scalar_dict(pl_df.median())
    return results


def write_dataframe(
    df: pd.DataFrame,
    output_path: str,
    sheet_name: str = "Sheet1",
) -> Dict[str, str]:
    """
    Запись DataFrame: pyexcelerate (очень большие) → xlsxwriter → openpyxl.
    """
    ext = _ext(output_path)
    rows, cols = df.shape
    engine = "openpyxl"

    if ext not in (".xlsx", ".xlsm", ".xls"):
        df.to_csv(output_path, index=False)
        return {"engine": "pandas_csv", "rows": str(rows)}

    if rows >= PYEXCELERATE_ROW_THRESHOLD and HAS_PYEXCELERATE and ext in (".xlsx",):
        data = [df.columns.tolist()] + df.fillna("").astype(object).values.tolist()
        wb = PyExWorkbook()  # type: ignore
        wb.new_sheet(sheet_name, data)
        wb.save(output_path)
        return {"engine": "pyexcelerate", "rows": str(rows)}

    if rows >= 3_000:
        with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return {"engine": "xlsxwriter", "rows": str(rows)}

    df.to_excel(output_path, sheet_name=sheet_name, index=False, engine="openpyxl")
    return {"engine": "openpyxl", "rows": str(rows)}


def copy_xls_with_xlutils(input_path: str, output_path: str) -> Dict[str, Any]:
    """Копия .xls с сохранением форматирования (xlutils + xlwt)."""
    if not HAS_XLUTILS or not HAS_XLWT:
        raise ImportError("xlutils/xlwt не установлены")
    rb = xlrd.open_workbook(input_path, formatting_info=True)
    wb = xlutils_copy(rb)  # type: ignore
    wb.save(output_path)
    return {"engine": "xlutils", "sheets": rb.sheet_names()}


def recalculate_workbook_formulas(
    file_path: str,
    output_dir: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Пересчёт формул пакетом formulas (без Excel COM).
    Возвращает путь к первому сгенерированному xlsx или ошибку.
    """
    if not HAS_FORMULAS_PKG:
        return {"skipped": True, "reason": "formulas not installed"}
    out_dir = output_dir or tempfile.mkdtemp(prefix="formulas_out_")
    try:
        model = formulas_pkg.ExcelModel().loads(file_path).finish()  # type: ignore
        model.calculate()
        written = model.write(dirpath=out_dir)
        out_files = []
        for _book, books in (written or {}).items():
            if isinstance(books, dict):
                for name, book in books.items():
                    candidate = os.path.join(out_dir, name)
                    if os.path.isfile(candidate):
                        out_files.append(candidate)
        return {
            "success": True,
            "engine": "formulas",
            "output_dir": out_dir,
            "output_files": out_files,
        }
    except Exception as e:
        return {"success": False, "engine": "formulas", "error": str(e)}


def count_formulas_pycel(file_path: str) -> Dict[str, Any]:
    """Подсчёт ячеек с формулами через Pycel (если доступен)."""
    if not HAS_PYCEL:
        return {"available": False}
    try:
        from pycel import ExcelCompiler

        compiler = ExcelCompiler(filename=file_path)
        formula_cells = [
            addr
            for addr, cell in compiler.formula_cells.items()
            if cell is not None
        ]
        return {"available": True, "formula_cell_count": len(formula_cells)}
    except Exception as e:
        return {"available": True, "error": str(e)}


def read_range_xlwings(
    file_path: str,
    sheet_name: str,
    cell_range: str,
) -> Dict[str, Any]:
    """
    Чтение диапазона через Excel COM (xlwings). Только Windows с установленным Excel.
    """
    if not HAS_XLWINGS:
        return {"available": False, "error": "xlwings not installed"}
    try:
        app = xw.App(visible=False, add_book=False)  # type: ignore
        try:
            book = app.books.open(file_path, read_only=True)
            sheet = book.sheets[sheet_name]
            values = sheet.range(cell_range).value
            book.close()
            return {"available": True, "engine": "xlwings", "values": values}
        finally:
            app.quit()
    except Exception as e:
        return {"available": True, "engine": "xlwings", "error": str(e)}


def grid_to_sheet_payload(
    grid: List[List[Any]],
    sheet_name: str,
    max_rows: int,
    read_engine: str = "pandas",
) -> Dict[str, Any]:
    """Преобразование таблицы значений в формат read_sheet_data."""
    data = []
    for row_idx, row in enumerate(grid[:max_rows], start=1):
        row_data = []
        for col_idx, val in enumerate(row, start=1):
            row_data.append(
                {
                    "row": row_idx,
                    "col": col_idx,
                    "value": val,
                    "type": _cell_type(val),
                }
            )
        data.append(row_data)
    cols = max((len(r) for r in grid), default=0)
    return {
        "sheet_name": sheet_name,
        "data": data,
        "range": f"A1",
        "total_rows": len(grid),
        "returned_rows": len(data),
        "has_more": len(grid) > max_rows,
        "columns": cols,
        "read_engine": read_engine,
    }


def _cell_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float, np.integer, np.floating)):
        return "number"
    if isinstance(value, str):
        return "string"
    return "unknown"
