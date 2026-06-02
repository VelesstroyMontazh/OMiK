"""
Data Operations - Data manipulation and analysis operations.
Implements sort, filter, find/replace, pivot, merge, format, insert/delete, and analysis.
"""

import os
from typing import Optional, List, Dict, Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
import pandas as pd
import numpy as np

import excel_libs
from excel_libs import (
    HAS_POLARS,
    POLARS_ROW_THRESHOLD,
    polars_analyze_numeric,
    polars_filter_dataframe,
    polars_sort_dataframe,
    read_dataframe,
)

from excel_handler import _parse_range, _nan_to_none


# =============================================================================
# Sort Operations
# =============================================================================

def sort_data(
    file_path: str,
    sheet_name: str,
    column: str,
    ascending: bool = True,
    cell_range: Optional[str] = None
) -> Dict[str, Any]:
    """Sort data in a sheet (Polars для больших таблиц, иначе Pandas)."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ['.xlsx', '.xlsm', '.xls', '.xlsb']:
        df = read_dataframe(file_path, sheet_name=sheet_name)
    elif ext == '.csv':
        df = pd.read_csv(file_path)
    else:
        raise ValueError(f"Sorting only supported for Excel/CSV files. Got: {ext}")

    if cell_range:
        min_row, min_col, max_row, max_col = _parse_range(cell_range)
        df = df.iloc[max(0, min_row - 2):max_row - 1, max(0, min_col - 1):max_col]

    if df.empty:
        return {"sorted": False, "message": "No data to sort"}

    # Resolve column - could be a header name or a column letter
    sort_col = column
    if column not in df.columns:
        # Try as a column letter (A, B, C, ...)
        try:
            col_idx = column_index_from_string(column) - 1
            if 0 <= col_idx < len(df.columns):
                sort_col = df.columns[col_idx]
            else:
                raise ValueError(f"Column '{column}' not found")
        except Exception:
            raise ValueError(f"Column '{column}' not found in data. Available: {list(df.columns)}")

    use_polars = HAS_POLARS and len(df) >= POLARS_ROW_THRESHOLD
    if use_polars:
        df_sorted = polars_sort_dataframe(df, str(sort_col), ascending)
    else:
        df_sorted = df.sort_values(by=sort_col, ascending=ascending, na_position='last')

    # Write back sorted data to the file
    if ext in ['.xlsx', '.xlsm']:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        # Write headers (row 1)
        for col_idx, col_name in enumerate(df_sorted.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data (starting from row 2)
        for row_idx, row_data in enumerate(df_sorted.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=_nan_to_none(value))

        wb.save(file_path)
        wb.close()
    else:
        # For CSV, overwrite the file
        df_sorted.to_csv(file_path, index=False)

    return {
        "sorted": True,
        "column": sort_col,
        "ascending": ascending,
        "rows_sorted": len(df_sorted),
        "engine": "polars" if use_polars else "pandas",
    }


# =============================================================================
# Filter Operations
# =============================================================================

def filter_data(
    file_path: str,
    sheet_name: str,
    column: str,
    condition: str,
    value: Any,
    cell_range: Optional[str] = None
) -> Dict[str, Any]:
    """
    Filter data based on conditions.
    Conditions: equals, not_equals, contains, starts_with, ends_with,
                greater_than, less_than, greater_equal, less_equal, between
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ['.xlsx', '.xlsm', '.xls', '.xlsb']:
        df = read_dataframe(file_path, sheet_name=sheet_name)
    elif ext == '.csv':
        df = pd.read_csv(file_path)
    else:
        raise ValueError(f"Unsupported format for filtering: {ext}")

    if column not in df.columns:
        # Try column index
        try:
            col_idx = column_index_from_string(column) - 1
            column = df.columns[col_idx]
        except Exception:
            raise ValueError(f"Column '{column}' not found")

    # Apply filter
    condition = condition.lower().replace('_', '').replace('-', '')

    if condition in ['equals', 'eq', '==']:
        mask = df[column] == value
    elif condition in ['notequals', 'ne', '!=']:
        mask = df[column] != value
    elif condition in ['contains', 'like']:
        mask = df[column].astype(str).str.contains(str(value), na=False)
    elif condition in ['startswith']:
        mask = df[column].astype(str).str.startswith(str(value), na=False)
    elif condition in ['endswith']:
        mask = df[column].astype(str).str.endswith(str(value), na=False)
    elif condition in ['greaterthan', 'gt', '>']:
        mask = df[column] > value
    elif condition in ['lessthan', 'lt', '<']:
        mask = df[column] < value
    elif condition in ['greaterequal', 'gte', '>=']:
        mask = df[column] >= value
    elif condition in ['lessequal', 'lte', '<=']:
        mask = df[column] <= value
    elif condition in ['between']:
        if isinstance(value, (list, tuple)) and len(value) == 2:
            mask = (df[column] >= value[0]) & (df[column] <= value[1])
        else:
            raise ValueError("Between condition requires a list of two values [min, max]")
    elif condition in ['isnull', 'isna']:
        mask = df[column].isna()
    elif condition in ['notnull', 'notna']:
        mask = df[column].notna()
    else:
        raise ValueError(f"Unknown condition: {condition}")

    use_polars = HAS_POLARS and len(df) >= POLARS_ROW_THRESHOLD
    if use_polars:
        filtered = polars_filter_dataframe(df, column, mask)
    else:
        filtered = df[mask]

    result_data = filtered.to_dict(orient='records')

    for row in result_data:
        for k, v in row.items():
            row[k] = _nan_to_none(v)

    return {
        "filtered": True,
        "column": column,
        "condition": condition,
        "value": value,
        "total_rows": len(df),
        "matched_rows": len(filtered),
        "data": result_data,
        "engine": "polars" if use_polars else "pandas",
    }


# =============================================================================
# Find and Replace
# =============================================================================

def find_replace(
    file_path: str,
    sheet_name: str,
    find: str,
    replace: str,
    cell_range: Optional[str] = None
) -> Dict[str, Any]:
    """Find and replace values in a sheet."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext not in ['.xlsx', '.xlsm']:
        raise ValueError(f"Find/replace only supported for .xlsx files. Got: {ext}")

    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]

    min_row, max_row = ws.min_row or 1, ws.max_row or 1
    min_col, max_col = ws.min_column or 1, ws.max_column or 1

    if cell_range:
        min_row, min_col, max_row, max_col = _parse_range(cell_range)

    replacements = 0
    cells_modified = []

    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col
    ):
        for cell in row:
            if cell.value is not None and str(find) in str(cell.value):
                old_value = cell.value
                cell.value = str(cell.value).replace(str(find), str(replace))
                replacements += 1
                cells_modified.append({
                    "cell": cell.coordinate,
                    "old_value": old_value,
                    "new_value": cell.value,
                })

    wb.save(file_path)
    wb.close()

    return {
        "replacements": replacements,
        "cells_modified": cells_modified[:100],  # Limit response size
        "find": find,
        "replace": replace,
    }


# =============================================================================
# Pivot Table
# =============================================================================

def create_pivot(
    file_path: str,
    sheet_name: str,
    rows: List[str],
    columns: Optional[List[str]] = None,
    values: Optional[List[str]] = None,
    agg_func: str = "sum"
) -> Dict[str, Any]:
    """Create a pivot table from sheet data."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ['.xlsx', '.xlsm']:
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    elif ext == '.xls':
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')
    elif ext == '.csv':
        df = pd.read_csv(file_path)
    else:
        raise ValueError(f"Unsupported format: {ext}")

    # Validate columns
    for col in rows:
        if col not in df.columns:
            raise ValueError(f"Row column '{col}' not found in data")

    if values:
        for col in values:
            if col not in df.columns:
                raise ValueError(f"Value column '{col}' not found in data")

    # Create pivot table
    agg_funcs = {
        'sum': 'sum', 'avg': 'mean', 'mean': 'mean', 'count': 'count',
        'min': 'min', 'max': 'max', 'std': 'std', 'median': 'median',
        'first': 'first', 'last': 'last'
    }

    func = agg_funcs.get(agg_func.lower(), agg_func)

    try:
        pivot_df = pd.pivot_table(
            df,
            index=rows,
            columns=columns,
            values=values,
            aggfunc=func,
            fill_value=0
        )
    except Exception as e:
        raise ValueError(f"Pivot table creation failed: {str(e)}")

    # Convert to serializable format
    pivot_df_reset = pivot_df.reset_index()

    # Save pivot table to a new sheet
    if ext in ['.xlsx', '.xlsm']:
        wb = openpyxl.load_workbook(file_path)
        pivot_sheet_name = f"Pivot_{sheet_name}"
        if pivot_sheet_name in wb.sheetnames:
            del wb[pivot_sheet_name]

        ws = wb.create_sheet(title=pivot_sheet_name)

        # Write headers
        for col_idx, col_name in enumerate(pivot_df_reset.columns, start=1):
            ws.cell(row=1, column=col_idx, value=str(col_name))

        # Write data
        for row_idx, row_data in enumerate(pivot_df_reset.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=_nan_to_none(value))

        wb.save(file_path)
        wb.close()

    result_data = pivot_df_reset.to_dict(orient='records')
    for row in result_data:
        for k, v in row.items():
            row[k] = _nan_to_none(v)

    return {
        "pivot_created": True,
        "pivot_sheet": pivot_sheet_name if ext in ['.xlsx', '.xlsm'] else None,
        "rows": rows,
        "columns": columns,
        "values": values,
        "agg_func": agg_func,
        "data": result_data,
        "total_rows": len(pivot_df_reset),
    }


# =============================================================================
# Merge/Unmerge Cells
# =============================================================================

def merge_unmerge_cells(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    action: str = "merge"
) -> Dict[str, Any]:
    """Merge or unmerge cells."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext not in ['.xlsx', '.xlsm']:
        raise ValueError(f"Merge/unmerge only supported for .xlsx files. Got: {ext}")

    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]

    if action.lower() == "merge":
        ws.merge_cells(cell_range)
    elif action.lower() == "unmerge":
        ws.unmerge_cells(cell_range)
    else:
        wb.close()
        raise ValueError(f"Unknown action: {action}. Use 'merge' or 'unmerge'")

    wb.save(file_path)
    wb.close()

    return {
        "action": action,
        "range": cell_range,
        "sheet": sheet_name,
    }


# =============================================================================
# Format Cells
# =============================================================================

def format_cells(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    format_type: str,
    format_value: Any
) -> Dict[str, Any]:
    """
    Format cells with various styles.
    format_type: number_format, font, fill, border, alignment
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext not in ['.xlsx', '.xlsm']:
        raise ValueError(f"Formatting only supported for .xlsx files. Got: {ext}")

    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]
    min_row, min_col, max_row, max_col = _parse_range(cell_range)

    cells_formatted = 0

    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col
    ):
        for cell in row:
            _apply_format(cell, format_type, format_value)
            cells_formatted += 1

    wb.save(file_path)
    wb.close()

    return {
        "formatted": True,
        "range": cell_range,
        "format_type": format_type,
        "cells_affected": cells_formatted,
    }


def _apply_format(cell, format_type: str, format_value: Any):
    """Apply a specific format to a cell."""
    format_type = format_type.lower()

    if format_type == "number_format":
        cell.number_format = format_value

    elif format_type == "font":
        if isinstance(format_value, dict):
            cell.font = Font(
                name=format_value.get("name"),
                size=format_value.get("size"),
                bold=format_value.get("bold", False),
                italic=format_value.get("italic", False),
                underline=format_value.get("underline"),
                color=format_value.get("color"),
            )
        else:
            # Simple preset
            presets = {
                "bold": Font(bold=True),
                "italic": Font(italic=True),
                "header": Font(bold=True, size=12),
                "title": Font(bold=True, size=14),
            }
            if format_value in presets:
                cell.font = presets[format_value]

    elif format_type == "fill":
        if isinstance(format_value, dict):
            cell.fill = PatternFill(
                start_color=format_value.get("color", format_value.get("start_color")),
                end_color=format_value.get("end_color"),
                fill_type=format_value.get("type", "solid"),
            )
        else:
            # Simple color fill
            cell.fill = PatternFill(start_color=format_value, end_color=format_value, fill_type="solid")

    elif format_type == "border":
        if isinstance(format_value, dict):
            side = Side(
                border_style=format_value.get("style", "thin"),
                color=format_value.get("color", "000000"),
            )
            cell.border = Border(left=side, right=side, top=side, bottom=side)
        else:
            presets = {
                "thin": Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin")
                ),
                "thick": Border(
                    left=Side(style="thick"), right=Side(style="thick"),
                    top=Side(style="thick"), bottom=Side(style="thick")
                ),
            }
            if format_value in presets:
                cell.border = presets[format_value]

    elif format_type == "alignment":
        if isinstance(format_value, dict):
            cell.alignment = Alignment(
                horizontal=format_value.get("horizontal"),
                vertical=format_value.get("vertical"),
                wrap_text=format_value.get("wrap_text", False),
            )
        else:
            presets = {
                "center": Alignment(horizontal="center", vertical="center"),
                "left": Alignment(horizontal="left"),
                "right": Alignment(horizontal="right"),
            }
            if format_value in presets:
                cell.alignment = presets[format_value]


# =============================================================================
# Insert/Delete Rows and Columns
# =============================================================================

def insert_rows_cols(
    file_path: str,
    sheet_name: str,
    position: int,
    count: int = 1,
    direction: str = "rows"
) -> Dict[str, Any]:
    """Insert rows or columns at a position."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext not in ['.xlsx', '.xlsm']:
        raise ValueError(f"Insert only supported for .xlsx files. Got: {ext}")

    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]

    if direction.lower() in ["rows", "row"]:
        ws.insert_rows(position, amount=count)
    elif direction.lower() in ["cols", "columns", "col", "column"]:
        ws.insert_cols(position, amount=count)
    else:
        wb.close()
        raise ValueError(f"Invalid direction: {direction}. Use 'rows' or 'cols'")

    wb.save(file_path)
    wb.close()

    return {
        "inserted": True,
        "position": position,
        "count": count,
        "direction": direction,
    }


def delete_rows_cols(
    file_path: str,
    sheet_name: str,
    position: int,
    count: int = 1,
    direction: str = "rows"
) -> Dict[str, Any]:
    """Delete rows or columns at a position."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext not in ['.xlsx', '.xlsm']:
        raise ValueError(f"Delete only supported for .xlsx files. Got: {ext}")

    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]

    if direction.lower() in ["rows", "row"]:
        ws.delete_rows(position, amount=count)
    elif direction.lower() in ["cols", "columns", "col", "column"]:
        ws.delete_cols(position, amount=count)
    else:
        wb.close()
        raise ValueError(f"Invalid direction: {direction}. Use 'rows' or 'cols'")

    wb.save(file_path)
    wb.close()

    return {
        "deleted": True,
        "position": position,
        "count": count,
        "direction": direction,
    }


# =============================================================================
# Statistical Analysis
# =============================================================================

def analyze_data(
    file_path: str,
    sheet_name: str,
    cell_range: Optional[str] = None,
    operations: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    Perform statistical analysis on data.
    Operations: sum, avg, count, min, max, std, median, var, mode, skew, kurt
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ['.xlsx', '.xlsm', '.xls', '.xlsb']:
        df = read_dataframe(file_path, sheet_name=sheet_name)
    elif ext == '.csv':
        df = pd.read_csv(file_path)
    else:
        raise ValueError(f"Unsupported format for analysis: {ext}")

    if cell_range:
        min_row, min_col, max_row, max_col = _parse_range(cell_range)
        # Convert to 0-indexed
        df = df.iloc[max(0, min_row - 2):max_row - 1, max(0, min_col - 1):max_col]

    if operations is None:
        operations = ["sum", "avg", "count", "min", "max", "std", "median"]

    # Select only numeric columns for analysis
    numeric_df = df.select_dtypes(include=[np.number])

    if numeric_df.empty:
        return {
            "analysis": {},
            "operations": operations,
            "message": "No numeric columns found for analysis",
        }

    engine = "pandas"
    fast: Dict[str, Any] = {}
    polars_ops = {o.lower() for o in operations} & {"sum", "avg", "mean", "min", "max", "std", "median"}
    if HAS_POLARS and len(df) >= POLARS_ROW_THRESHOLD and polars_ops:
        fast = polars_analyze_numeric(numeric_df, list(polars_ops))
        if fast:
            engine = "polars"
            for op_key, vals in fast.items():
                for col_key in vals:
                    if isinstance(vals[col_key], list) and len(vals[col_key]) == 1:
                        vals[col_key] = vals[col_key][0]
                    vals[col_key] = _nan_to_none(vals[col_key])
            remaining = [o for o in operations if o.lower() not in fast]
            if not remaining:
                return {
                    "analysis": fast,
                    "operations": operations,
                    "numeric_columns": list(numeric_df.columns),
                    "total_rows": len(df),
                    "total_columns": len(df.columns),
                    "engine": engine,
                }

    results = {}

    for op in operations:
        op_lower = op.lower()

        if op_lower in ["sum"]:
            results[op_lower] = numeric_df.sum().to_dict()
        elif op_lower in ["avg", "mean"]:
            results[op_lower] = numeric_df.mean().to_dict()
        elif op_lower in ["count"]:
            results[op_lower] = numeric_df.count().to_dict()
        elif op_lower in ["min"]:
            results[op_lower] = numeric_df.min().to_dict()
        elif op_lower in ["max"]:
            results[op_lower] = numeric_df.max().to_dict()
        elif op_lower in ["std"]:
            results[op_lower] = numeric_df.std().to_dict()
        elif op_lower in ["median"]:
            results[op_lower] = numeric_df.median().to_dict()
        elif op_lower in ["var", "variance"]:
            results[op_lower] = numeric_df.var().to_dict()
        elif op_lower in ["mode"]:
            results[op_lower] = {}
            for col in numeric_df.columns:
                mode_val = numeric_df[col].mode()
                results[op_lower][col] = mode_val.tolist() if not mode_val.empty else None
        elif op_lower in ["skew"]:
            results[op_lower] = numeric_df.skew().to_dict()
        elif op_lower in ["kurt", "kurtosis"]:
            results[op_lower] = numeric_df.kurtosis().to_dict()
        elif op_lower in ["describe"]:
            desc = numeric_df.describe().to_dict()
            results[op_lower] = {}
            for col, stats in desc.items():
                results[op_lower][col] = {k: _nan_to_none(v) for k, v in stats.items()}

    # Convert NaN to None in results
    for op_key in results:
        if isinstance(results[op_key], dict):
            for col_key in results[op_key]:
                val = results[op_key][col_key]
                if isinstance(val, dict):
                    for k, v in val.items():
                        val[k] = _nan_to_none(v)
                else:
                    results[op_key][col_key] = _nan_to_none(val)

    if engine == "polars" and fast:
        results = {**fast, **results}

    return {
        "analysis": results,
        "operations": operations,
        "numeric_columns": list(numeric_df.columns),
        "total_rows": len(df),
        "total_columns": len(df.columns),
        "engine": engine,
    }
