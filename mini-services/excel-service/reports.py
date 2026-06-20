"""
Reporting Engine - Generates various statistical reports from Main DB and Calendar DB.
"""

import os
import sqlite3
import json
import uuid
from typing import Optional, Dict, Any, List
from datetime import datetime, date, timedelta

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from calendar_db import CALENDAR_EXPORT_COLUMNS
from data_paths import UPLOAD_DIR
import main_db

MAIN_DB_PATH = main_db.DB_PATH
MAIN_META_PATH = main_db.META_PATH
CALENDAR_DB_PATH = os.path.join(UPLOAD_DIR, "calendar_db.sqlite")

# Allowed column names for SQL queries to prevent SQL injection
ALLOWED_MAIN_DB_COLUMNS = {
    "Дата приема", "Дата_приема", "Дата увольнения", "Дата_увольнения",
    "Состояние", "Страна гражданства", "Страна_гражданства",
    "Площадка", "Итого", "Территория", "Организация",
    "Подразделение", "Должность", "ФИО", "rowid"
}

ALLOWED_CALENDAR_COLUMNS = {
    "year", "month", "direction", "citizenship", "arrival_status",
    "justification", "department", "worker_type", "row_number",
    "date", "merged_row_id"
}

def _safe_column_name(name: str, allowed: set) -> str:
    """Validate column name against whitelist to prevent SQL injection."""
    if name not in allowed:
        raise ValueError(f"Недопустимое имя колонки: {name}")
    return name


def _safe_int(val, default=None):
    """Safely convert a value to int."""
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def _get_main_db_conn() -> Optional[sqlite3.Connection]:
    if not os.path.exists(MAIN_DB_PATH):
        return None
    conn = sqlite3.connect(MAIN_DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.create_function("LOWER", 1, lambda x: x.lower() if x else None)
    return conn


def _get_calendar_db_conn() -> Optional[sqlite3.Connection]:
    if not os.path.exists(CALENDAR_DB_PATH):
        return None
    conn = sqlite3.connect(CALENDAR_DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.create_function("LOWER", 1, lambda x: x.lower() if x else None)
    return conn


def _load_col_mapping() -> Dict[str, str]:
    """Load column mapping from metadata."""
    if os.path.exists(MAIN_META_PATH):
        with open(MAIN_META_PATH, 'r', encoding='utf-8') as f:
            meta = json.load(f)
        return meta.get("col_mapping", {})
    return {}


# ==================== MAIN DB REPORTS ====================

def report_employment_by_period(
    year: Optional[int] = None,
    month: Optional[int] = None,
    citizenship: Optional[str] = None,
    territory: Optional[str] = None,
    organization: Optional[str] = None,
    status: Optional[str] = None,
) -> Dict[str, Any]:
    """Report on employees hired (Дата приема) by period with filters."""
    conn = _get_main_db_conn()
    if conn is None:
        return {"error": "Main database not loaded"}

    try:
        col_mapping = _load_col_mapping()
        date_col = col_mapping.get("Дата приема", "Дата_приема")
        dismissal_col = col_mapping.get("Дата увольнения", "Дата_увольнения")
        status_col = col_mapping.get("Состояние", "Состояние")
        citizenship_col = col_mapping.get("Страна гражданства", "Страна_гражданства")
        territory_col = col_mapping.get("Площадка") or col_mapping.get("Итого") or col_mapping.get("Территория", "Территория")
        org_col = col_mapping.get("Организация", "Организация")

        # Build WHERE clause
        where_parts = [f'"{date_col}" IS NOT NULL AND "{date_col}" != ""']
        params = []

        if citizenship:
            where_parts.append(f'LOWER("{citizenship_col}") LIKE ?')
            params.append(f'%{citizenship.lower()}%')
        if territory:
            where_parts.append(f'LOWER("{territory_col}") LIKE ?')
            params.append(f'%{territory.lower()}%')
        if organization:
            where_parts.append(f'LOWER("{org_col}") LIKE ?')
            params.append(f'%{organization.lower()}%')
        if status:
            where_parts.append(f'LOWER("{status_col}") LIKE ?')
            params.append(f'%{status.lower()}%')

        where_clause = f'WHERE {" AND ".join(where_parts)}'

        # By year
        by_year = conn.execute(
            f'''
            SELECT SUBSTR("{date_col}", 1, 4) as yr, COUNT(*) as cnt
            FROM employees {where_clause}
            AND LENGTH(SUBSTR("{date_col}", 1, 4)) = 4
            AND SUBSTR("{date_col}", 1, 4) GLOB '[0-9][0-9][0-9][0-9]'
            GROUP BY yr ORDER BY yr
            ''', params
        ).fetchall()

        # By year-month
        by_month = conn.execute(
            f'''
            SELECT SUBSTR("{date_col}", 1, 4) as yr,
                   SUBSTR("{date_col}", 4, 2) as mn,
                   SUBSTR("{date_col}", 1, 7) as period,
                   COUNT(*) as cnt
            FROM employees {where_clause}
            AND LENGTH(SUBSTR("{date_col}", 1, 4)) = 4
            AND SUBSTR("{date_col}", 1, 4) GLOB '[0-9][0-9][0-9][0-9]'
            GROUP BY yr, mn ORDER BY yr, mn
            ''', params
        ).fetchall()

        # Also try with different date formats (DD.MM.YYYY)
        by_year_v2 = conn.execute(
            f'''
            SELECT SUBSTR("{date_col}", 7, 4) as yr, COUNT(*) as cnt
            FROM employees {where_clause}
            AND SUBSTR("{date_col}", 3, 1) = '.'
            AND LENGTH("{date_col}") >= 10
            AND SUBSTR("{date_col}", 7, 4) GLOB '[0-9][0-9][0-9][0-9]'
            GROUP BY yr ORDER BY yr
            ''', params
        ).fetchall()

        by_month_v2 = conn.execute(
            f'''
            SELECT SUBSTR("{date_col}", 7, 4) as yr,
                   SUBSTR("{date_col}", 4, 2) as mn,
                   SUBSTR("{date_col}", 7, 4) || '-' || SUBSTR("{date_col}", 4, 2) as period,
                   COUNT(*) as cnt
            FROM employees {where_clause}
            AND SUBSTR("{date_col}", 3, 1) = '.'
            AND LENGTH("{date_col}") >= 10
            AND SUBSTR("{date_col}", 7, 4) GLOB '[0-9][0-9][0-9][0-9]'
            GROUP BY yr, mn ORDER BY yr, mn
            ''', params
        ).fetchall()

        # Use whichever format gives better results
        if len(by_year_v2) > len(by_year):
            by_year_final = [{"year": _safe_int(r[0]), "count": r[1]} for r in by_year_v2 if r[0] and _safe_int(r[0])]
            by_month_final = [{"year": _safe_int(r[0]), "month": _safe_int(r[1]), "period": r[2], "count": r[3]} for r in by_month_v2 if r[0] and _safe_int(r[0])]
        else:
            by_year_final = [{"year": _safe_int(r[0]), "count": r[1]} for r in by_year if r[0] and _safe_int(r[0])]
            by_month_final = [{"year": _safe_int(r[0]), "month": _safe_int(r[1]), "period": r[2], "count": r[3]} for r in by_month if r[0] and _safe_int(r[0])]

        # Filter by requested year/month if specified
        filtered_by_month = by_month_final
        if year:
            filtered_by_month = [r for r in filtered_by_month if r.get("year") == year]
        if month:
            filtered_by_month = [r for r in filtered_by_month if r.get("month") == month]

        # By citizenship for the filtered data
        citizenship_filter = list(params)
        citizenship_where = f'WHERE {" AND ".join(where_parts)}'
        by_citizenship = conn.execute(
            f'''
            SELECT "{citizenship_col}", COUNT(*) as cnt
            FROM employees {citizenship_where}
            AND "{date_col}" IS NOT NULL AND "{date_col}" != ""
            GROUP BY "{citizenship_col}" ORDER BY cnt DESC LIMIT 20
            ''', citizenship_filter
        ).fetchall()

        # By territory
        by_territory = conn.execute(
            f'''
            SELECT "{territory_col}", COUNT(*) as cnt
            FROM employees {citizenship_where}
            AND "{date_col}" IS NOT NULL AND "{date_col}" != ""
            GROUP BY "{territory_col}" ORDER BY cnt DESC LIMIT 20
            ''', citizenship_filter
        ).fetchall()

        # Total count
        total = conn.execute(
            f'SELECT COUNT(*) FROM employees {where_clause}', params
        ).fetchone()[0]

        return {
            "report_type": "employment_by_period",
            "title": "Трудоустройство по периодам",
            "total": total,
            "by_year": by_year_final,
            "by_month": filtered_by_month if (year or month) else by_month_final,
            "by_citizenship": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_citizenship],
            "by_territory": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_territory],
            "filters_applied": {
                "year": year,
                "month": month,
                "citizenship": citizenship,
                "territory": territory,
                "organization": organization,
                "status": status,
            }
        }
    finally:
        conn.close()


def report_dismissal_by_period(
    year: Optional[int] = None,
    month: Optional[int] = None,
    citizenship: Optional[str] = None,
    territory: Optional[str] = None,
    organization: Optional[str] = None,
) -> Dict[str, Any]:
    """Report on employees dismissed (Дата увольнения) by period with filters."""
    conn = _get_main_db_conn()
    if conn is None:
        return {"error": "Main database not loaded"}

    try:
        col_mapping = _load_col_mapping()
        dismissal_col = col_mapping.get("Дата увольнения", "Дата_увольнения")
        citizenship_col = col_mapping.get("Страна гражданства", "Страна_гражданства")
        territory_col = col_mapping.get("Площадка") or col_mapping.get("Итого") or col_mapping.get("Территория", "Территория")
        org_col = col_mapping.get("Организация", "Организация")
        status_col = col_mapping.get("Состояние", "Состояние")

        where_parts = [f'"{dismissal_col}" IS NOT NULL AND "{dismissal_col}" != ""']
        params = []

        if citizenship:
            where_parts.append(f'LOWER("{citizenship_col}") LIKE ?')
            params.append(f'%{citizenship.lower()}%')
        if territory:
            where_parts.append(f'LOWER("{territory_col}") LIKE ?')
            params.append(f'%{territory.lower()}%')
        if organization:
            where_parts.append(f'LOWER("{org_col}") LIKE ?')
            params.append(f'%{organization.lower()}%')

        where_clause = f'WHERE {" AND ".join(where_parts)}'

        # By year (DD.MM.YYYY format)
        by_year = conn.execute(
            f'''
            SELECT SUBSTR("{dismissal_col}", 7, 4) as yr, COUNT(*) as cnt
            FROM employees {where_clause}
            AND SUBSTR("{dismissal_col}", 3, 1) = '.'
            AND LENGTH("{dismissal_col}") >= 10
            AND SUBSTR("{dismissal_col}", 7, 4) GLOB '[0-9][0-9][0-9][0-9]'
            GROUP BY yr ORDER BY yr
            ''', params
        ).fetchall()

        # Try YYYY format
        by_year_v1 = conn.execute(
            f'''
            SELECT SUBSTR("{dismissal_col}", 1, 4) as yr, COUNT(*) as cnt
            FROM employees {where_clause}
            AND SUBSTR("{dismissal_col}", 5, 1) = '-'
            AND SUBSTR("{dismissal_col}", 1, 4) GLOB '[0-9][0-9][0-9][0-9]'
            GROUP BY yr ORDER BY yr
            ''', params
        ).fetchall()

        if len(by_year_v1) > len(by_year):
            by_year = by_year_v1
            is_ddmm = False
        else:
            is_ddmm = True

        # By month
        if is_ddmm:
            by_month = conn.execute(
                f'''
                SELECT SUBSTR("{dismissal_col}", 7, 4) as yr,
                       SUBSTR("{dismissal_col}", 4, 2) as mn,
                       SUBSTR("{dismissal_col}", 7, 4) || '-' || SUBSTR("{dismissal_col}", 4, 2) as period,
                       COUNT(*) as cnt
                FROM employees {where_clause}
                AND SUBSTR("{dismissal_col}", 3, 1) = '.'
                AND LENGTH("{dismissal_col}") >= 10
                AND SUBSTR("{dismissal_col}", 7, 4) GLOB '[0-9][0-9][0-9][0-9]'
                GROUP BY yr, mn ORDER BY yr, mn
                ''', params
            ).fetchall()
        else:
            by_month = conn.execute(
                f'''
                SELECT SUBSTR("{dismissal_col}", 1, 4) as yr,
                       SUBSTR("{dismissal_col}", 6, 2) as mn,
                       SUBSTR("{dismissal_col}", 1, 7) as period,
                       COUNT(*) as cnt
                FROM employees {where_clause}
                AND SUBSTR("{dismissal_col}", 5, 1) = '-'
                AND SUBSTR("{dismissal_col}", 1, 4) GLOB '[0-9][0-9][0-9][0-9]'
                GROUP BY yr, mn ORDER BY yr, mn
                ''', params
            ).fetchall()

        by_year_final = [{"year": _safe_int(r[0]), "count": r[1]} for r in by_year if r[0] and _safe_int(r[0])]
        by_month_final = [{"year": _safe_int(r[0]), "month": _safe_int(r[1]), "period": r[2], "count": r[3]} for r in by_month if r[0] and _safe_int(r[0])]

        filtered_by_month = by_month_final
        if year:
            filtered_by_month = [r for r in filtered_by_month if r.get("year") == year]
        if month:
            filtered_by_month = [r for r in filtered_by_month if r.get("month") == month]

        # By citizenship
        by_citizenship = conn.execute(
            f'''
            SELECT "{citizenship_col}", COUNT(*) as cnt
            FROM employees {where_clause}
            GROUP BY "{citizenship_col}" ORDER BY cnt DESC LIMIT 20
            ''', params
        ).fetchall()

        # By territory
        by_territory = conn.execute(
            f'''
            SELECT "{territory_col}", COUNT(*) as cnt
            FROM employees {where_clause}
            GROUP BY "{territory_col}" ORDER BY cnt DESC LIMIT 20
            ''', params
        ).fetchall()

        # By status
        by_status = conn.execute(
            f'''
            SELECT "{status_col}", COUNT(*) as cnt
            FROM employees {where_clause}
            GROUP BY "{status_col}" ORDER BY cnt DESC
            ''', params
        ).fetchall()

        total = conn.execute(
            f'SELECT COUNT(*) FROM employees {where_clause}', params
        ).fetchone()[0]

        return {
            "report_type": "dismissal_by_period",
            "title": "Увольнения по периодам",
            "total": total,
            "by_year": by_year_final,
            "by_month": filtered_by_month if (year or month) else by_month_final,
            "by_citizenship": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_citizenship],
            "by_territory": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_territory],
            "by_status": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_status],
            "filters_applied": {
                "year": year,
                "month": month,
                "citizenship": citizenship,
                "territory": territory,
                "organization": organization,
            }
        }
    finally:
        conn.close()


def report_current_composition(
    status: Optional[str] = None,
    citizenship: Optional[str] = None,
    territory: Optional[str] = None,
    organization: Optional[str] = None,
) -> Dict[str, Any]:
    """Report on current employee composition by status, citizenship, territory."""
    conn = _get_main_db_conn()
    if conn is None:
        return {"error": "Main database not loaded"}

    try:
        col_mapping = _load_col_mapping()
        status_col = col_mapping.get("Состояние", "Состояние")
        citizenship_col = col_mapping.get("Страна гражданства", "Страна_гражданства")
        territory_col = col_mapping.get("Площадка") or col_mapping.get("Итого") or col_mapping.get("Территория", "Территория")
        org_col = col_mapping.get("Организация", "Организация")
        dept_col = col_mapping.get("Подразделение", "Подразделение")
        pos_col = col_mapping.get("Должность", "Должность")

        where_parts = []
        params = []

        if status:
            where_parts.append(f'LOWER("{status_col}") LIKE ?')
            params.append(f'%{status.lower()}%')
        if citizenship:
            where_parts.append(f'LOWER("{citizenship_col}") LIKE ?')
            params.append(f'%{citizenship.lower()}%')
        if territory:
            where_parts.append(f'LOWER("{territory_col}") LIKE ?')
            params.append(f'%{territory.lower()}%')
        if organization:
            where_parts.append(f'LOWER("{org_col}") LIKE ?')
            params.append(f'%{organization.lower()}%')

        where_clause = f'WHERE {" AND ".join(where_parts)}' if where_parts else ''

        # By status
        by_status = conn.execute(
            f'SELECT "{status_col}", COUNT(*) as cnt FROM employees {where_clause} GROUP BY "{status_col}" ORDER BY cnt DESC',
            params
        ).fetchall()

        # By citizenship
        by_citizenship = conn.execute(
            f'SELECT "{citizenship_col}", COUNT(*) as cnt FROM employees {where_clause} GROUP BY "{citizenship_col}" ORDER BY cnt DESC LIMIT 20',
            params
        ).fetchall()

        # By territory
        by_territory = conn.execute(
            f'SELECT "{territory_col}", COUNT(*) as cnt FROM employees {where_clause} GROUP BY "{territory_col}" ORDER BY cnt DESC LIMIT 20',
            params
        ).fetchall()

        # By organization
        by_organization = conn.execute(
            f'SELECT "{org_col}", COUNT(*) as cnt FROM employees {where_clause} GROUP BY "{org_col}" ORDER BY cnt DESC LIMIT 20',
            params
        ).fetchall()

        # By department (top 20)
        by_department = conn.execute(
            f'SELECT "{dept_col}", COUNT(*) as cnt FROM employees {where_clause} GROUP BY "{dept_col}" ORDER BY cnt DESC LIMIT 20',
            params
        ).fetchall()

        # By position (top 20)
        by_position = conn.execute(
            f'SELECT "{pos_col}", COUNT(*) as cnt FROM employees {where_clause} GROUP BY "{pos_col}" ORDER BY cnt DESC LIMIT 20',
            params
        ).fetchall()

        # Cross-tab: status x citizenship
        cross_status_citizenship = conn.execute(
            f'''
            SELECT "{status_col}", "{citizenship_col}", COUNT(*) as cnt
            FROM employees {where_clause}
            GROUP BY "{status_col}", "{citizenship_col}"
            ORDER BY cnt DESC LIMIT 50
            ''', params
        ).fetchall()

        total = conn.execute(f'SELECT COUNT(*) FROM employees {where_clause}', params).fetchone()[0]

        return {
            "report_type": "current_composition",
            "title": "Текущий состав сотрудников",
            "total": total,
            "by_status": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_status],
            "by_citizenship": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_citizenship],
            "by_territory": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_territory],
            "by_organization": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_organization],
            "by_department": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_department],
            "by_position": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_position],
            "cross_status_citizenship": [
                {"status": r[0] or "Не указано", "citizenship": r[1] or "Не указано", "count": r[2]}
                for r in cross_status_citizenship
            ],
            "filters_applied": {
                "status": status,
                "citizenship": citizenship,
                "territory": territory,
                "organization": organization,
            }
        }
    finally:
        conn.close()


def report_calendar_summary(
    direction: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    citizenship: Optional[str] = None,
    justification: Optional[str] = None,
    justification_contains: Optional[str] = None,
    arrival_status: Optional[str] = None,
    worker_type: Optional[str] = None,
    department: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> Dict[str, Any]:
    """Generate calendar report (Прилет/Вылет) with filters."""
    from calendar_db import is_loaded as cal_loaded, build_calendar_filter_clause

    if not cal_loaded():
        return {"error": "Calendar database not loaded. Please load the calendar file first."}

    conn = _get_calendar_db_conn()
    if conn is None:
        return {"error": "Calendar database not found"}

    try:
        where_parts, params = build_calendar_filter_clause(
            direction=direction,
            year=year,
            month=month,
            date_from=date_from,
            date_to=date_to,
            citizenship=citizenship,
            justification=justification,
            justification_contains=justification_contains,
            arrival_status=arrival_status,
            worker_type=worker_type,
            department=department,
        )
        where_clause = f'WHERE {" AND ".join(where_parts)}' if where_parts else ''

        total = conn.execute(f'SELECT COUNT(*) FROM calendar_records {where_clause}', params).fetchone()[0]

        by_justification = conn.execute(
            f'SELECT justification, COUNT(*) as cnt FROM calendar_records {where_clause} '
            f'GROUP BY justification ORDER BY cnt DESC LIMIT 30',
            params,
        ).fetchall()

        by_citizenship = conn.execute(
            f'SELECT citizenship, COUNT(*) as cnt FROM calendar_records {where_clause} '
            f'GROUP BY citizenship ORDER BY cnt DESC LIMIT 20',
            params,
        ).fetchall()

        by_arrival_status = conn.execute(
            f'SELECT arrival_status, COUNT(*) as cnt FROM calendar_records {where_clause} '
            f'GROUP BY arrival_status ORDER BY cnt DESC',
            params,
        ).fetchall()

        by_month = conn.execute(
            f'SELECT year, month, direction, COUNT(*) as cnt FROM calendar_records {where_clause} '
            f'GROUP BY year, month, direction ORDER BY year, month',
            params,
        ).fetchall()

        by_department = conn.execute(
            f'SELECT department, COUNT(*) as cnt FROM calendar_records {where_clause} '
            f'GROUP BY department ORDER BY cnt DESC LIMIT 20',
            params,
        ).fetchall()

        title_parts = ["Календарь"]
        if direction:
            title_parts.append(direction)
        if justification:
            title_parts.append(f"— {justification}")
        elif justification_contains:
            title_parts.append(f"— обоснование «{justification_contains}»")
        if date_from or date_to:
            title_parts.append(f"({date_from or '…'} — {date_to or '…'})")

        return {
            "report_type": "calendar_summary",
            "title": " ".join(title_parts),
            "total": total,
            "by_justification": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_justification],
            "by_citizenship": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_citizenship],
            "by_arrival_status": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_arrival_status],
            "by_month": [{"year": r[0], "month": r[1], "direction": r[2], "count": r[3]} for r in by_month],
            "by_department": [{"name": r[0] or "Не указано", "count": r[1]} for r in by_department],
            "filters_applied": {
                "direction": direction,
                "year": year,
                "month": month,
                "date_from": date_from,
                "date_to": date_to,
                "citizenship": citizenship,
                "justification": justification,
                "justification_contains": justification_contains,
                "arrival_status": arrival_status,
                "worker_type": worker_type,
                "department": department,
            },
        }
    finally:
        conn.close()


def report_calendar_conditional(
    direction: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    citizenship: Optional[str] = None,
    justification: Optional[str] = None,
    justification_contains: Optional[str] = None,
    arrival_status: Optional[str] = None,
    worker_type: Optional[str] = None,
    department: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    output_name: Optional[str] = None,
) -> Dict[str, Any]:
    """Conditional calendar report with Excel export and in-app preview."""
    from calendar_db import (
        is_loaded as cal_loaded,
        build_calendar_filter_clause,
        CALENDAR_EXPORT_COLUMNS,
    )
    import excel_handler

    if not cal_loaded():
        return {"error": "Календарь не загружен. Сначала загрузите файл календаря."}

    summary = report_calendar_summary(
        direction=direction,
        year=year,
        month=month,
        citizenship=citizenship,
        justification=justification,
        justification_contains=justification_contains,
        arrival_status=arrival_status,
        worker_type=worker_type,
        department=department,
        date_from=date_from,
        date_to=date_to,
    )
    if "error" in summary:
        return summary

    conn = _get_calendar_db_conn()
    if conn is None:
        return {"error": "Calendar database not found"}

    try:
        where_parts, params = build_calendar_filter_clause(
            direction=direction,
            year=year,
            month=month,
            date_from=date_from,
            date_to=date_to,
            citizenship=citizenship,
            justification=justification,
            justification_contains=justification_contains,
            arrival_status=arrival_status,
            worker_type=worker_type,
            department=department,
        )
        where_clause = f'WHERE {" AND ".join(where_parts)}' if where_parts else ''

        export_cols = list(CALENDAR_EXPORT_COLUMNS.keys())
        col_sql = ", ".join(export_cols)
        query = (
            f"SELECT {col_sql} FROM calendar_records {where_clause} "
            f"ORDER BY year, month, row_number"
        )
        df = pd.read_sql_query(query, conn, params=params)

        preview_limit = 200
        preview_rows = df.head(preview_limit).to_dict(orient="records")

        excel_handler.ensure_upload_dir()
        base_name = output_name.strip() if output_name else (
            f"calendar_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not base_name.lower().endswith(".xlsx"):
            base_name = f"{base_name}.xlsx"
        stored_filename = f"{uuid.uuid4()}_{base_name}"
        output_path = os.path.join(UPLOAD_DIR, stored_filename)

        export_df = df.rename(columns=CALENDAR_EXPORT_COLUMNS)
        summary_rows = [
            ("Показатель", "Значение"),
            ("Всего записей", summary["total"]),
            ("Направление", direction or "Все"),
            ("Год", year or "Все"),
            ("Месяц", month or "Все"),
            ("Дата с", date_from or "—"),
            ("Дата по", date_to or "—"),
            ("Обоснование (точное)", justification or "—"),
            ("Обоснование (содержит)", justification_contains or "—"),
            ("Гражданство", citizenship or "—"),
            ("Статус прибытия", arrival_status or "—"),
            ("Рабочий/ИТР", worker_type or "—"),
            ("Отдел", department or "—"),
        ]

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            export_df.to_excel(writer, sheet_name="Данные", index=False)
            pd.DataFrame(summary_rows[1:], columns=summary_rows[0]).to_excel(
                writer, sheet_name="Сводка", index=False
            )
            if summary.get("by_justification"):
                pd.DataFrame(summary["by_justification"]).to_excel(
                    writer, sheet_name="По обоснованию", index=False
                )

        return {
            **summary,
            "report_type": "calendar_conditional",
            "preview_rows": preview_rows,
            "preview_limit": preview_limit,
            "file_path": output_path,
            "stored_filename": stored_filename,
            "file_id": os.path.splitext(stored_filename)[0],
        }
    finally:
        conn.close()


from calendar_db import CALENDAR_EXPORT_COLUMNS

MERGED_EXPORT_COLUMNS: Dict[str, str] = {
    **CALENDAR_EXPORT_COLUMNS,
    "Табельный номер (База)": "Табельный номер (База)",
    "ФИО (База)": "ФИО (База)",
    "Организация (База)": "Организация (База)",
    "Подразделение (База)": "Подразделение (База)",
    "Состояние (База)": "Состояние (База)",
}


def _get_merged_calendar_db_conn():
    from integration_ops import CALENDAR_MERGED_DB_PATH, is_merged_calendar_loaded

    if not is_merged_calendar_loaded():
        return None
    if not os.path.exists(CALENDAR_MERGED_DB_PATH):
        return None
    return sqlite3.connect(CALENDAR_MERGED_DB_PATH)


def report_calendar_merged_conditional(
    direction: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    citizenship: Optional[str] = None,
    justification: Optional[str] = None,
    justification_contains: Optional[str] = None,
    arrival_status: Optional[str] = None,
    worker_type: Optional[str] = None,
    department: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    output_name: Optional[str] = None,
) -> Dict[str, Any]:
    """Conditional report on calendar+main merged dataset."""
    from calendar_db import build_calendar_filter_clause
    from integration_ops import is_merged_calendar_loaded
    import excel_handler

    if not is_merged_calendar_loaded():
        return {
            "error": "Объединённый календарь не построен. Сначала выполните «Объединить с Базой» на вкладке «Календарь + База».",
        }

    summary = report_calendar_summary(
        direction=direction,
        year=year,
        month=month,
        citizenship=citizenship,
        justification=justification,
        justification_contains=justification_contains,
        arrival_status=arrival_status,
        worker_type=worker_type,
        department=department,
        date_from=date_from,
        date_to=date_to,
    )
    if "error" in summary:
        return summary

    conn = _get_merged_calendar_db_conn()
    if conn is None:
        return {"error": "База объединённого календаря не найдена"}

    try:
        where_parts, params = build_calendar_filter_clause(
            direction=direction,
            year=year,
            month=month,
            date_from=date_from,
            date_to=date_to,
            citizenship=citizenship,
            justification=justification,
            justification_contains=justification_contains,
            arrival_status=arrival_status,
            worker_type=worker_type,
            department=department,
        )
        where_clause = f'WHERE {" AND ".join(where_parts)}' if where_parts else ''

        export_cols = list(MERGED_EXPORT_COLUMNS.keys())
        col_sql = ", ".join(f'"{c}"' if " " in c or "(" in c else c for c in export_cols)
        query = (
            f"SELECT {col_sql} FROM calendar_merged_records {where_clause} "
            f"ORDER BY year, month, row_number"
        )
        df = pd.read_sql_query(query, conn, params=params)

        matched_base = int(df["Табельный номер (База)"].notna().sum()) if "Табельный номер (База)" in df.columns else 0
        preview_limit = 200
        preview_rows = df.head(preview_limit).to_dict(orient="records")

        excel_handler.ensure_upload_dir()
        base_name = output_name.strip() if output_name else (
            f"calendar_merged_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not base_name.lower().endswith(".xlsx"):
            base_name = f"{base_name}.xlsx"
        stored_filename = f"{uuid.uuid4()}_{base_name}"
        output_path = os.path.join(UPLOAD_DIR, stored_filename)

        export_df = df.rename(columns=MERGED_EXPORT_COLUMNS)
        title = summary.get("title", "Календарь").replace("Календарь", "Календарь + База", 1)
        summary_rows = [
            ("Показатель", "Значение"),
            ("Всего записей", len(df)),
            ("Сопоставлено с Базой", matched_base),
            ("Направление", direction or "Все"),
            ("Год", year or "Все"),
            ("Месяц", month or "Все"),
        ]

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            export_df.to_excel(writer, sheet_name="Данные", index=False)
            pd.DataFrame(summary_rows[1:], columns=summary_rows[0]).to_excel(
                writer, sheet_name="Сводка", index=False
            )

        return {
            **summary,
            "title": title,
            "report_type": "calendar_merged_conditional",
            "total": int(len(df)),
            "matched_with_base": matched_base,
            "preview_rows": preview_rows,
            "preview_limit": preview_limit,
            "file_path": output_path,
            "stored_filename": stored_filename,
            "file_id": os.path.splitext(stored_filename)[0],
        }
    finally:
        conn.close()


def get_main_db_unique_values(column: str) -> List[str]:
    """Get unique values for a main DB column."""
    conn = _get_main_db_conn()
    if conn is None:
        return []

    try:
        col_mapping = _load_col_mapping()
        s_col = col_mapping.get(column)
        if not s_col:
            return []

        rows = conn.execute(
            f'SELECT DISTINCT "{s_col}" FROM employees WHERE "{s_col}" IS NOT NULL AND "{s_col}" != "" ORDER BY "{s_col}"'
        ).fetchall()
        return [r[0] for r in rows if r[0]]
    finally:
        conn.close()


def get_available_report_filters() -> Dict[str, Any]:
    """Get available filter options for all report types."""
    result = {
        "main_db": {},
        "calendar": {},
    }

    # Main DB filters
    conn = _get_main_db_conn()
    if conn:
        try:
            col_mapping = _load_col_mapping()
            for col_name, key in [
                ("Состояние", "statuses"),
                ("Страна гражданства", "citizenships"),
                ("Территория", "territories"),
                ("Организация", "organizations"),
            ]:
                s_col = col_mapping.get(col_name)
                if s_col:
                    rows = conn.execute(
                        f'SELECT DISTINCT "{s_col}" FROM employees WHERE "{s_col}" IS NOT NULL AND "{s_col}" != "" ORDER BY "{s_col}"'
                    ).fetchall()
                    result["main_db"][key] = [r[0] for r in rows if r[0]]

            # Available years from employment dates
            date_col = col_mapping.get("Дата приема", "Дата_приема")
            rows = conn.execute(
                f'SELECT DISTINCT SUBSTR("{date_col}", 7, 4) FROM employees '
                f'WHERE SUBSTR("{date_col}", 3, 1) = \'.\' '
                f'AND SUBSTR("{date_col}", 7, 4) GLOB \'[0-9][0-9][0-9][0-9]\' '
                f'ORDER BY SUBSTR("{date_col}", 7, 4)'
            ).fetchall()
            result["main_db"]["employment_years"] = [int(r[0]) for r in rows if r[0]]

            dismissal_col = col_mapping.get("Дата увольнения", "Дата_увольнения")
            rows = conn.execute(
                f'SELECT DISTINCT SUBSTR("{dismissal_col}", 7, 4) FROM employees '
                f'WHERE SUBSTR("{dismissal_col}", 3, 1) = \'.\' '
                f'AND SUBSTR("{dismissal_col}", 7, 4) GLOB \'[0-9][0-9][0-9][0-9]\' '
                f'ORDER BY SUBSTR("{dismissal_col}", 7, 4)'
            ).fetchall()
            result["main_db"]["dismissal_years"] = [int(r[0]) for r in rows if r[0]]

        finally:
            conn.close()

    # Calendar filters
    from calendar_db import is_loaded as cal_loaded, get_unique_values
    if cal_loaded():
        result["calendar"]["citizenships"] = get_unique_values("citizenship")
        result["calendar"]["justifications"] = get_unique_values("justification")
        result["calendar"]["arrival_statuses"] = get_unique_values("arrival_status")
        result["calendar"]["directions"] = ["Прилет", "Вылет"]
        result["calendar"]["worker_types"] = get_unique_values("worker_type")
        result["calendar"]["departments"] = get_unique_values("department")

        conn2 = _get_calendar_db_conn()
        if conn2:
            try:
                rows = conn2.execute('SELECT DISTINCT year FROM calendar_records WHERE year IS NOT NULL ORDER BY year').fetchall()
                result["calendar"]["years"] = [r[0] for r in rows]
                rows = conn2.execute('SELECT DISTINCT month FROM calendar_records WHERE month IS NOT NULL ORDER BY month').fetchall()
                result["calendar"]["months"] = [r[0] for r in rows]
            finally:
                conn2.close()

    return result


def _parse_date_series(series: pd.Series) -> pd.Series:
    """
    Parse mixed-format date series.
    First pass: default parser (ISO friendly)
    Second pass: day-first parser for unresolved values.
    """
    first = pd.to_datetime(series, errors="coerce")
    need_second = first.isna() & series.notna() & (series.astype(str).str.strip() != "")
    if need_second.any():
        second = pd.to_datetime(series[need_second], errors="coerce", dayfirst=True)
        first.loc[need_second] = second
    return first


def report_base_presence_matrix(
    start_date_str: Optional[str] = None,
    end_date_str: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Build custom day-by-day presence matrix report from Main DB.

    Output columns:
    Табельный номер (с префиксами), ФИО, Страна гражданства, Организация,
    Подразделение, Территория, Состояние, Дата приема, Дата увольнения,
    then daily columns from start_date to end_date with formula:
    =ЕСЛИ(И(J$1>=$H2; ИЛИ($I2=""; J$1<=$I2)); 1; 0)
    """
    conn = _get_main_db_conn()
    if conn is None:
        return {"error": "Main database not loaded"}

    try:
        col_mapping = _load_col_mapping()
        required = [
            "Табельный номер (с префиксами)",
            "ФИО",
            "Страна гражданства",
            "Организация",
            "Подразделение",
            "Территория",
            "Состояние",
            "Дата приема",
            "Дата увольнения",
        ]
        sql_cols = [col_mapping.get(c) for c in required]
        if any(not c for c in sql_cols):
            return {"error": "Не найдены обязательные столбцы в основной БД"}

        quoted = ", ".join(f'"{c}"' for c in sql_cols)
        df = pd.read_sql_query(f"SELECT {quoted} FROM employees", conn)
        df.columns = required

        # Preserve raw strings for audit
        raw_hire = df["Дата приема"].astype(str).str.strip()
        raw_fire = df["Дата увольнения"].astype(str).str.strip()

        # Parse dates
        df["Дата приема"] = _parse_date_series(df["Дата приема"])
        df["Дата увольнения"] = _parse_date_series(df["Дата увольнения"])

        invalid_hire = int(((raw_hire != "") & df["Дата приема"].isna()).sum())
        invalid_fire = int(((raw_fire != "") & df["Дата увольнения"].isna()).sum())
        missing_hire = int(df["Дата приема"].isna().sum())

        fire_before_hire = (
            df["Дата приема"].notna()
            & df["Дата увольнения"].notna()
            & (df["Дата увольнения"] < df["Дата приема"])
        )
        fire_before_hire_count = int(fire_before_hire.sum())
        if fire_before_hire_count > 0:
            # Keep hire date and clear invalid dismissal date.
            df.loc[fire_before_hire, "Дата увольнения"] = pd.NaT

        start_date_value = (
            datetime.strptime(start_date_str, "%d.%m.%Y").date()
            if start_date_str
            else date(2025, 1, 1)
        )
        end_date_value = (
            datetime.strptime(end_date_str, "%d.%m.%Y").date()
            if end_date_str
            else date.today()
        )
        if end_date_value < start_date_value:
            return {"error": "Дата окончания меньше даты начала"}

        start_ts = pd.Timestamp(start_date_value)
        end_ts = pd.Timestamp(end_date_value)

        # Keep only rows that intersect the report period
        relevant_mask = (
            df["Дата приема"].notna()
            & (df["Дата приема"] <= end_ts)
            & (df["Дата увольнения"].isna() | (df["Дата увольнения"] >= start_ts))
        )
        report_df = df.loc[relevant_mask].copy()

        # Build date columns
        day_columns: List[date] = []
        cursor = start_date_value
        while cursor <= end_date_value:
            day_columns.append(cursor)
            cursor += timedelta(days=1)

        # Write workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"

        headers = required + day_columns
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            if isinstance(header, date):
                cell.value = header
                cell.number_format = "DD.MM.YYYY"
                cell.fill = PatternFill("solid", fgColor="00FFFF")
            else:
                cell.value = header
                cell.fill = PatternFill("solid", fgColor="FFFF00")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, row in enumerate(report_df.itertuples(index=False), start=2):
            for j, val in enumerate(row, start=1):
                c = ws.cell(row=i, column=j)
                if j in (8, 9):  # date columns
                    if pd.notna(val):
                        c.value = val.date()
                        c.number_format = "DD.MM.YYYY"
                    else:
                        c.value = None
                else:
                    c.value = None if pd.isna(val) else str(val)

            # Date columns start at J (10)
            for k in range(10, 10 + len(day_columns)):
                letter = ws.cell(row=1, column=k).column_letter
                ws.cell(row=i, column=k).value = (
                    f'=ЕСЛИ(И({letter}$1>=$H{i}; ИЛИ($I{i}=""; {letter}$1<=$I{i})); 1; 0)'
                )

        ws.freeze_panes = "J2"
        widths = [22, 32, 20, 28, 28, 28, 20, 14, 14]
        for cidx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(1, cidx).column_letter].width = width
        for cidx in range(10, 10 + len(day_columns)):
            ws.column_dimensions[ws.cell(1, cidx).column_letter].width = 12

        audit = wb.create_sheet("Аудит")
        audit_rows = [
            ("Показатель", "Значение"),
            ("Всего записей в Базе", int(len(df))),
            ("Записей в отчете (пересечение с периодом)", int(len(report_df))),
            ("Пустая дата приема", missing_hire),
            ("Некорректный формат даты приема", invalid_hire),
            ("Некорректный формат даты увольнения", invalid_fire),
            (
                "Дата увольнения раньше даты приема (исправлено очисткой даты увольнения)",
                fire_before_hire_count,
            ),
            ("Начало периода", start_date_value.strftime("%d.%m.%Y")),
            ("Конец периода", end_date_value.strftime("%d.%m.%Y")),
            ("Количество дневных колонок", len(day_columns)),
        ]
        for r_idx, (k, v) in enumerate(audit_rows, start=1):
            audit.cell(r_idx, 1, k)
            audit.cell(r_idx, 2, v)
            if r_idx == 1:
                audit.cell(r_idx, 1).font = Font(bold=True)
                audit.cell(r_idx, 2).font = Font(bold=True)
        audit.column_dimensions["A"].width = 78
        audit.column_dimensions["B"].width = 24

        os.makedirs(UPLOAD_DIR, exist_ok=True)
        output_name = f"{uuid.uuid4()}_base_report_{start_date_value.strftime('%Y%m%d')}_{end_date_value.strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(UPLOAD_DIR, output_name)
        wb.save(output_path)
        wb.close()

        return {
            "report_type": "base_presence_matrix",
            "title": "Матрица присутствия (База)",
            "total": int(len(report_df)),
            "total_rows_in_db": int(len(df)),
            "date_columns": len(day_columns),
            "start_date": start_date_value.strftime("%d.%m.%Y"),
            "end_date": end_date_value.strftime("%d.%m.%Y"),
            "file_path": output_path,
            "stored_filename": output_name,
            "file_id": os.path.splitext(output_name)[0],
            "audit": {
                "missing_hire": missing_hire,
                "invalid_hire": invalid_hire,
                "invalid_fire": invalid_fire,
                "fire_before_hire_fixed": fire_before_hire_count,
            },
        }
    finally:
        conn.close()
