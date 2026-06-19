"""
Main Database Caching System - Uses SQLite for reliable, memory-efficient storage.
Loads Excel data into SQLite, then uses SQL for all queries.

Key columns (0-based indices): 0, 1, 2, 3, 4, 5, 6, 8, 10, 11, 12, 13

IMPROVEMENTS:
- LRU Cache with max size limit to prevent memory leaks
- Async-safe operations with asyncio.to_thread()
- Cross-platform path handling (Windows/Linux/macOS)
- Integration with OMiK/download/ reference files
- Magic bytes validation for file uploads
"""

import os
import json
import re
import sqlite3
import threading
import time
import gc
import hashlib
import mimetypes
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime
from collections import OrderedDict
from pathlib import Path
import asyncio

import pandas as pd
import numpy as np
import openpyxl

from data_paths import MAIN_DB_DIR

# Cross-platform project root detection
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
UPLOAD_DIR = os.path.join(PROJECT_ROOT, "upload")
DB_PATH = os.path.join(UPLOAD_DIR, "main_db.sqlite")
META_PATH = os.path.join(UPLOAD_DIR, "main_db_meta.json")

# Reference files directory (OMiK/download/)
REFERENCE_DIR = os.path.join(PROJECT_ROOT, "OMiK", "download")

KEY_COLUMN_INDICES = [0, 1, 2, 3, 4, 5, 6, 8, 10, 11, 12, 13]

# LRU Cache configuration
MAX_CACHE_SIZE = 100  # Maximum number of cached items
CACHE_TTL_SECONDS = 3600  # Cache TTL: 1 hour

_conn_lock = threading.RLock()
_open_connections: List[sqlite3.Connection] = []

# LRU Cache implementation with size limit and TTL
class LRUCache:
    """Thread-safe LRU Cache with max size and TTL."""
    
    def __init__(self, max_size: int = MAX_CACHE_SIZE, ttl_seconds: int = CACHE_TTL_SECONDS):
        self._cache: OrderedDict = OrderedDict()
        self._timestamps: Dict[str, float] = {}
        self._max_size = max_size
        self._ttl_seconds = ttl_seconds
        self._lock = threading.RLock()
    
    def get(self, key: str, default: Any = None) -> Any:
        with self._lock:
            if key not in self._cache:
                return default
            
            # Check TTL
            if time.time() - self._timestamps[key] > self._ttl_seconds:
                self._remove(key)
                return default
            
            # Move to end (most recently used)
            self._cache.move_to_end(key)
            return self._get_cache()[key]
    
    def set(self, key: str, value: Any) -> None:
        with self._lock:
            if key in self._cache:
                self._cache.move_to_end(key)
            self._get_cache()[key] = value
            self._timestamps[key] = time.time()
            
            # Evict oldest if over max size
            while len(self._cache) > self._max_size:
                oldest_key = next(iter(self._cache))
                self._remove(oldest_key)
    
    def _remove(self, key: str) -> None:
        if key in self._cache:
            del self._get_cache()[key]
        if key in self._timestamps:
            del self._timestamps[key]
    
    def clear(self) -> None:
        with self._lock:
            self._cache.clear()
            self._timestamps.clear()
    
    def invalidate(self, key: str) -> None:
        with self._lock:
            self._remove(key)
    
    def size(self) -> int:
        with self._lock:
            return len(self._cache)
    
    def cleanup_expired(self) -> int:
        """Remove expired entries, return count of removed items."""
        with self._lock:
            now = time.time()
            expired_keys = [
                k for k, ts in self._timestamps.items()
                if now - ts > self._ttl_seconds
            ]
            for key in expired_keys:
                self._remove(key)
            return len(expired_keys)


# Global LRU cache instead of simple dict
_cache_store = LRUCache(max_size=MAX_CACHE_SIZE, ttl_seconds=CACHE_TTL_SECONDS)

# Backward compatibility wrapper
def _get_cache() -> Dict[str, Any]:
    """Get current cache state as dict for backward compatibility."""
    cached = _cache_store.get("main_db_state", {})
    if not cached:
        return {
            "loaded": False,
            "file_path": None,
            "sheet_name": None,
            "columns": [],
            "key_columns": [],
            "col_mapping": {},
            "loaded_at": None,
            "row_count": 0,
            "col_count": 0,
        }
    return cached

def _set_cache(data: Dict[str, Any]) -> None:
    """Set cache state."""
    _cache_store.set("main_db_state", data)

def _invalidate_cache() -> None:
    """Invalidate the main DB cache."""
    _cache_store.invalidate("main_db_state")


def _nan_to_none(value):
    if value is None:
        return None
    if isinstance(value, float) and np.isnan(value):
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _convert_value_for_json(value):
    if value is None:
        return None
    if isinstance(value, (pd.Timestamp, datetime)):
        return str(value)
    if isinstance(value, float) and np.isnan(value):
        return None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        if np.isnan(value):
            return None
        return float(value)
    if isinstance(value, np.bool_):
        return bool(value)
    return value


def _get_cell_type(value) -> str:
    if value is None:
        return "null"
    elif isinstance(value, bool):
        return "boolean"
    elif isinstance(value, (int, float)):
        return "number"
    elif isinstance(value, str):
        return "string"
    else:
        return "string"


def main_db_upload_root() -> str:
    """Единственный каталог для Excel и SQLite Основной Базы."""
    return os.path.normcase(MAIN_DB_DIR)


def main_db_search_dirs() -> List[str]:
    """Только upload проекта (C:\\...\\OMiK_VSM\\upload)."""
    if os.path.isdir(MAIN_DB_DIR):
        return [MAIN_DB_DIR]
    return []


def _path_under_main_db_upload(file_path: str) -> bool:
    root = main_db_upload_root()
    norm = os.path.normcase(os.path.abspath(file_path))
    return norm == root or norm.startswith(root + os.sep)


def _detect_main_db_file() -> Optional[str]:
    xlsx_files: List[tuple[str, int, str]] = []
    for folder in main_db_search_dirs():
        try:
            names = os.listdir(folder)
        except OSError:
            continue
        for filename in names:
            if not filename.lower().endswith((".xlsx", ".xlsm")):
                continue
            file_path = os.path.join(folder, filename)
            if os.path.isfile(file_path):
                xlsx_files.append((file_path, os.path.getsize(file_path), filename))
    if not xlsx_files:
        return None
    for file_path, _, filename in xlsx_files:
        name_lower = filename.lower()
        if "1с" in name_lower or "1c" in name_lower:
            return file_path
    xlsx_files.sort(key=lambda x: x[1], reverse=True)
    return xlsx_files[0][0]


def _match_key_columns(columns: List[str]) -> List[str]:
    matched = []
    for idx in KEY_COLUMN_INDICES:
        if idx < len(columns):
            col_name = columns[idx]
            if col_name not in matched:
                matched.append(col_name)
    return matched


def _cell_as_text(value: Any) -> str:
    """Текст ячейки без потери ведущих нулей (строка или целое из Excel)."""
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    if isinstance(value, str):
        s = value.strip()
        return "" if s.lower() in ("nan", "none", "nat") else s
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)):
        if float(value) == int(value):
            return str(int(value))
        return str(value).strip()
    s = str(value).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


def _is_passport_series_col(col_name: str) -> bool:
    cl = col_name.lower().replace(" ", "")
    return "удостоверение" in cl and "серия" in cl


def _is_passport_number_col(col_name: str) -> bool:
    cl = col_name.lower().replace(" ", "")
    return "удостоверение" in cl and "номер" in cl and "серия" not in cl


# Загранпаспорт: серии 82 и 83 — две цифры, без «00» впереди (не 0082/0083).
_PASSPORT_SERIES_NO_ZFILL = frozenset({"82", "83"})


def _format_passport_series_digits(digits: str) -> str:
    core = digits.lstrip("0") or "0"
    if core in _PASSPORT_SERIES_NO_ZFILL:
        return core
    if len(digits) <= 4:
        return digits.zfill(4)
    return digits


def _format_passport_value(value: Any, col_name: str) -> str:
    """Серия — 4 цифры (кроме 82/83), номер — 6 цифр (ведущие нули)."""
    s = _cell_as_text(value)
    if not s:
        return ""
    digits = re.sub(r"\D", "", s)
    if not digits:
        return s
    if _is_passport_series_col(col_name) and len(digits) <= 4:
        return _format_passport_series_digits(digits)
    if _is_passport_number_col(col_name) and len(digits) <= 6:
        return digits.zfill(6)
    return digits if digits == re.sub(r"\s", "", s) else s


def _overlay_passport_columns_from_workbook(
    df: pd.DataFrame,
    columns: List[str],
    file_path: str,
    sheet_name: str,
) -> pd.DataFrame:
    """Один проход по листу: Серия/Номер как в Excel (тип «текст» сохраняет нули)."""
    targets = [c for c in columns if _is_passport_series_col(c) or _is_passport_number_col(c)]
    if not targets:
        return df
    out = df.copy()
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        col_index: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell.value is not None:
                col_index[str(cell.value).strip()] = idx
        target_ci: List[tuple[str, int]] = []
        for col_name in targets:
            ci = col_index.get(col_name)
            if ci is not None:
                target_ci.append((col_name, ci))
        if not target_ci:
            wb.close()
            return out
        buffers: Dict[str, List[str]] = {name: [] for name, _ in target_ci}
        for row in ws.iter_rows(min_row=2, values_only=False):
            cells = list(row)
            for col_name, ci in target_ci:
                if ci >= len(cells):
                    buffers[col_name].append("")
                    continue
                cell = cells[ci]
                if getattr(cell, "data_type", None) == "s" or isinstance(cell.value, str):
                    buffers[col_name].append(_cell_as_text(cell.value))
                else:
                    buffers[col_name].append(_format_passport_value(cell.value, col_name))
        wb.close()
        for col_name, _ in target_ci:
            if len(buffers[col_name]) == len(out):
                out[col_name] = buffers[col_name]
    except Exception:
        return df
    return out


def _fix_passport_columns_in_dataframe(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if _is_passport_series_col(col) or _is_passport_number_col(col):
            out[col] = out[col].map(lambda v, c=col: _format_passport_value(v, c))
    return out


def _cell_as_text(value: Any) -> str:
    """Текст ячейки без потери ведущих нулей (строка или целое из Excel)."""
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    if isinstance(value, str):
        s = value.strip()
        return "" if s.lower() in ("nan", "none", "nat") else s
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)):
        if float(value) == int(value):
            return str(int(value))
        return str(value).strip()
    s = str(value).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


def _is_passport_series_col(col_name: str) -> bool:
    cl = col_name.lower().replace(" ", "")
    return "удостоверение" in cl and "серия" in cl


def _is_passport_number_col(col_name: str) -> bool:
    cl = col_name.lower().replace(" ", "")
    return "удостоверение" in cl and "номер" in cl and "серия" not in cl


# Загранпаспорт: серии 82 и 83 — две цифры, без «00» впереди (не 0082/0083).
_PASSPORT_SERIES_NO_ZFILL = frozenset({"82", "83"})


def _format_passport_series_digits(digits: str) -> str:
    core = digits.lstrip("0") or "0"
    if core in _PASSPORT_SERIES_NO_ZFILL:
        return core
    if len(digits) <= 4:
        return digits.zfill(4)
    return digits


def _format_passport_value(value: Any, col_name: str) -> str:
    """Серия — 4 цифры (кроме 82/83), номер — 6 цифр (ведущие нули)."""
    s = _cell_as_text(value)
    if not s:
        return ""
    digits = re.sub(r"\D", "", s)
    if not digits:
        return s
    if _is_passport_series_col(col_name) and len(digits) <= 4:
        return _format_passport_series_digits(digits)
    if _is_passport_number_col(col_name) and len(digits) <= 6:
        return digits.zfill(6)
    return digits if digits == re.sub(r"\s", "", s) else s


def _overlay_passport_columns_from_workbook(
    df: pd.DataFrame,
    columns: List[str],
    file_path: str,
    sheet_name: str,
) -> pd.DataFrame:
    """Один проход по листу: Серия/Номер как в Excel (тип «текст» сохраняет нули)."""
    targets = [c for c in columns if _is_passport_series_col(c) or _is_passport_number_col(c)]
    if not targets:
        return df
    out = df.copy()
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        col_index: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell.value is not None:
                col_index[str(cell.value).strip()] = idx
        target_ci: List[tuple[str, int]] = []
        for col_name in targets:
            ci = col_index.get(col_name)
            if ci is not None:
                target_ci.append((col_name, ci))
        if not target_ci:
            wb.close()
            return out
        buffers: Dict[str, List[str]] = {name: [] for name, _ in target_ci}
        for row in ws.iter_rows(min_row=2, values_only=False):
            cells = list(row)
            for col_name, ci in target_ci:
                if ci >= len(cells):
                    buffers[col_name].append("")
                    continue
                cell = cells[ci]
                if getattr(cell, "data_type", None) == "s" or isinstance(cell.value, str):
                    buffers[col_name].append(_cell_as_text(cell.value))
                else:
                    buffers[col_name].append(_format_passport_value(cell.value, col_name))
        wb.close()
        for col_name, _ in target_ci:
            if len(buffers[col_name]) == len(out):
                out[col_name] = buffers[col_name]
    except Exception:
        return df
    return out


def _fix_passport_columns_in_dataframe(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if _is_passport_series_col(col) or _is_passport_number_col(col):
            out[col] = out[col].map(lambda v, c=col: _format_passport_value(v, c))
    return out


def _cell_as_text(value: Any) -> str:
    """Текст ячейки без потери ведущих нулей (строка или целое из Excel)."""
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    if isinstance(value, str):
        s = value.strip()
        return "" if s.lower() in ("nan", "none", "nat") else s
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)):
        if float(value) == int(value):
            return str(int(value))
        return str(value).strip()
    s = str(value).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


def _is_passport_series_col(col_name: str) -> bool:
    cl = col_name.lower().replace(" ", "")
    return "удостоверение" in cl and "серия" in cl


def _is_passport_number_col(col_name: str) -> bool:
    cl = col_name.lower().replace(" ", "")
    return "удостоверение" in cl and "номер" in cl and "серия" not in cl


# Загранпаспорт: серии 82 и 83 — две цифры, без «00» впереди (не 0082/0083).
_PASSPORT_SERIES_NO_ZFILL = frozenset({"82", "83"})


def _format_passport_series_digits(digits: str) -> str:
    core = digits.lstrip("0") or "0"
    if core in _PASSPORT_SERIES_NO_ZFILL:
        return core
    if len(digits) <= 4:
        return digits.zfill(4)
    return digits


def _format_passport_value(value: Any, col_name: str) -> str:
    """Серия — 4 цифры (кроме 82/83), номер — 6 цифр (ведущие нули)."""
    s = _cell_as_text(value)
    if not s:
        return ""
    digits = re.sub(r"\D", "", s)
    if not digits:
        return s
    if _is_passport_series_col(col_name) and len(digits) <= 4:
        return _format_passport_series_digits(digits)
    if _is_passport_number_col(col_name) and len(digits) <= 6:
        return digits.zfill(6)
    return digits if digits == re.sub(r"\s", "", s) else s


def _overlay_passport_columns_from_workbook(
    df: pd.DataFrame,
    columns: List[str],
    file_path: str,
    sheet_name: str,
) -> pd.DataFrame:
    """Один проход по листу: Серия/Номер как в Excel (тип «текст» сохраняет нули)."""
    targets = [c for c in columns if _is_passport_series_col(c) or _is_passport_number_col(c)]
    if not targets:
        return df
    out = df.copy()
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        col_index: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell.value is not None:
                col_index[str(cell.value).strip()] = idx
        target_ci: List[tuple[str, int]] = []
        for col_name in targets:
            ci = col_index.get(col_name)
            if ci is not None:
                target_ci.append((col_name, ci))
        if not target_ci:
            wb.close()
            return out
        buffers: Dict[str, List[str]] = {name: [] for name, _ in target_ci}
        for row in ws.iter_rows(min_row=2, values_only=False):
            cells = list(row)
            for col_name, ci in target_ci:
                if ci >= len(cells):
                    buffers[col_name].append("")
                    continue
                cell = cells[ci]
                if getattr(cell, "data_type", None) == "s" or isinstance(cell.value, str):
                    buffers[col_name].append(_cell_as_text(cell.value))
                else:
                    buffers[col_name].append(_format_passport_value(cell.value, col_name))
        wb.close()
        for col_name, _ in target_ci:
            if len(buffers[col_name]) == len(out):
                out[col_name] = buffers[col_name]
    except Exception:
        return df
    return out


def _fix_passport_columns_in_dataframe(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if _is_passport_series_col(col) or _is_passport_number_col(col):
            out[col] = out[col].map(lambda v, c=col: _format_passport_value(v, c))
    return out


def _cell_as_text(value: Any) -> str:
    """Текст ячейки без потери ведущих нулей (строка или целое из Excel)."""
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    if isinstance(value, str):
        s = value.strip()
        return "" if s.lower() in ("nan", "none", "nat") else s
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)):
        if float(value) == int(value):
            return str(int(value))
        return str(value).strip()
    s = str(value).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


def _is_passport_series_col(col_name: str) -> bool:
    cl = col_name.lower().replace(" ", "")
    return "удостоверение" in cl and "серия" in cl


def _is_passport_number_col(col_name: str) -> bool:
    cl = col_name.lower().replace(" ", "")
    return "удостоверение" in cl and "номер" in cl and "серия" not in cl


# Загранпаспорт: серии 82 и 83 — две цифры, без «00» впереди (не 0082/0083).
_PASSPORT_SERIES_NO_ZFILL = frozenset({"82", "83"})


def _format_passport_series_digits(digits: str) -> str:
    core = digits.lstrip("0") or "0"
    if core in _PASSPORT_SERIES_NO_ZFILL:
        return core
    if len(digits) <= 4:
        return digits.zfill(4)
    return digits


def _format_passport_value(value: Any, col_name: str) -> str:
    """Серия — 4 цифры (кроме 82/83), номер — 6 цифр (ведущие нули)."""
    s = _cell_as_text(value)
    if not s:
        return ""
    digits = re.sub(r"\D", "", s)
    if not digits:
        return s
    if _is_passport_series_col(col_name) and len(digits) <= 4:
        return _format_passport_series_digits(digits)
    if _is_passport_number_col(col_name) and len(digits) <= 6:
        return digits.zfill(6)
    return digits if digits == re.sub(r"\s", "", s) else s


def _overlay_passport_columns_from_workbook(
    df: pd.DataFrame,
    columns: List[str],
    file_path: str,
    sheet_name: str,
) -> pd.DataFrame:
    """Один проход по листу: Серия/Номер как в Excel (тип «текст» сохраняет нули)."""
    targets = [c for c in columns if _is_passport_series_col(c) or _is_passport_number_col(c)]
    if not targets:
        return df
    out = df.copy()
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        col_index: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell.value is not None:
                col_index[str(cell.value).strip()] = idx
        target_ci: List[tuple[str, int]] = []
        for col_name in targets:
            ci = col_index.get(col_name)
            if ci is not None:
                target_ci.append((col_name, ci))
        if not target_ci:
            wb.close()
            return out
        buffers: Dict[str, List[str]] = {name: [] for name, _ in target_ci}
        for row in ws.iter_rows(min_row=2, values_only=False):
            cells = list(row)
            for col_name, ci in target_ci:
                if ci >= len(cells):
                    buffers[col_name].append("")
                    continue
                cell = cells[ci]
                if getattr(cell, "data_type", None) == "s" or isinstance(cell.value, str):
                    buffers[col_name].append(_cell_as_text(cell.value))
                else:
                    buffers[col_name].append(_format_passport_value(cell.value, col_name))
        wb.close()
        for col_name, _ in target_ci:
            if len(buffers[col_name]) == len(out):
                out[col_name] = buffers[col_name]
    except Exception:
        return df
    return out


def _fix_passport_columns_in_dataframe(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if _is_passport_series_col(col) or _is_passport_number_col(col):
            out[col] = out[col].map(lambda v, c=col: _format_passport_value(v, c))
    return out


def _cell_as_text(value: Any) -> str:
    """Текст ячейки без потери ведущих нулей (строка или целое из Excel)."""
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    if isinstance(value, str):
        s = value.strip()
        return "" if s.lower() in ("nan", "none", "nat") else s
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)):
        if float(value) == int(value):
            return str(int(value))
        return str(value).strip()
    s = str(value).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


def _is_passport_series_col(col_name: str) -> bool:
    cl = col_name.lower().replace(" ", "")
    return "удостоверение" in cl and "серия" in cl


def _is_passport_number_col(col_name: str) -> bool:
    cl = col_name.lower().replace(" ", "")
    return "удостоверение" in cl and "номер" in cl and "серия" not in cl


# Загранпаспорт: серии 82 и 83 — две цифры, без «00» впереди (не 0082/0083).
_PASSPORT_SERIES_NO_ZFILL = frozenset({"82", "83"})


def _format_passport_series_digits(digits: str) -> str:
    core = digits.lstrip("0") or "0"
    if core in _PASSPORT_SERIES_NO_ZFILL:
        return core
    if len(digits) <= 4:
        return digits.zfill(4)
    return digits


def _format_passport_value(value: Any, col_name: str) -> str:
    """Серия — 4 цифры (кроме 82/83), номер — 6 цифр (ведущие нули)."""
    s = _cell_as_text(value)
    if not s:
        return ""
    digits = re.sub(r"\D", "", s)
    if not digits:
        return s
    if _is_passport_series_col(col_name) and len(digits) <= 4:
        return _format_passport_series_digits(digits)
    if _is_passport_number_col(col_name) and len(digits) <= 6:
        return digits.zfill(6)
    return digits if digits == re.sub(r"\s", "", s) else s


def _overlay_passport_columns_from_workbook(
    df: pd.DataFrame,
    columns: List[str],
    file_path: str,
    sheet_name: str,
) -> pd.DataFrame:
    """Один проход по листу: Серия/Номер как в Excel (тип «текст» сохраняет нули)."""
    targets = [c for c in columns if _is_passport_series_col(c) or _is_passport_number_col(c)]
    if not targets:
        return df
    out = df.copy()
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        col_index: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell.value is not None:
                col_index[str(cell.value).strip()] = idx
        target_ci: List[tuple[str, int]] = []
        for col_name in targets:
            ci = col_index.get(col_name)
            if ci is not None:
                target_ci.append((col_name, ci))
        if not target_ci:
            wb.close()
            return out
        buffers: Dict[str, List[str]] = {name: [] for name, _ in target_ci}
        for row in ws.iter_rows(min_row=2, values_only=False):
            cells = list(row)
            for col_name, ci in target_ci:
                if ci >= len(cells):
                    buffers[col_name].append("")
                    continue
                cell = cells[ci]
                if getattr(cell, "data_type", None) == "s" or isinstance(cell.value, str):
                    buffers[col_name].append(_cell_as_text(cell.value))
                else:
                    buffers[col_name].append(_format_passport_value(cell.value, col_name))
        wb.close()
        for col_name, _ in target_ci:
            if len(buffers[col_name]) == len(out):
                out[col_name] = buffers[col_name]
    except Exception:
        return df
    return out


def _fix_passport_columns_in_dataframe(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if _is_passport_series_col(col) or _is_passport_number_col(col):
            out[col] = out[col].map(lambda v, c=col: _format_passport_value(v, c))
    return out


def _sanitize_col_name(name: str) -> str:
    """Sanitize column name for use as SQLite column name."""
    # Replace dots and special chars with underscore
    return name.replace('.', '_').replace(' ', '_').replace('(', '').replace(')', '').replace('-', '_')


def _register_connection(conn: sqlite3.Connection) -> None:
    with _conn_lock:
        _open_connections.append(conn)


def close_all_connections() -> None:
    """Закрыть все соединения с main_db.sqlite (перед пересборкой файла)."""
    with _conn_lock:
        for conn in list(_open_connections):
            try:
                conn.close()
            except Exception:
                pass
        _open_connections.clear()
    gc.collect()


def _remove_path_with_retry(path: str, attempts: int = 8) -> None:
    if not os.path.isfile(path):
        return
    last_err: Optional[Exception] = None
    for _ in range(attempts):
        try:
            os.remove(path)
            return
        except OSError as exc:
            last_err = exc
            time.sleep(0.35)
    if last_err:
        raise last_err


def _get_db_connection() -> sqlite3.Connection:
    """Get a connection to the main database with custom Unicode LOWER function."""
    conn = sqlite3.connect(_db_path(), timeout=60.0)
    conn.row_factory = sqlite3.Row
    conn.create_function("LOWER", 1, lambda x: x.lower() if x else None)
    _register_connection(conn)
    return conn


def load_main_db(
    file_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    *,
    set_active: bool = False,
) -> Dict[str, Any]:
    """Load an Excel file into a new SQLite instance (не перезаписывает предыдущие)."""
    
    if not file_path or not str(file_path).strip():
        return {
            "loaded": False,
            "error": "Укажите путь к файлу Excel (выгрузка 1С). Автопоиск отключён — выберите файл вручную.",
        }

    file_path = os.path.abspath(str(file_path).strip())

    os.makedirs(MAIN_DB_DIR, exist_ok=True)

    if not _path_under_main_db_upload(file_path):
        return {
            "loaded": False,
            "error": (
                "Основная База загружается только из папки upload проекта: "
                f"{MAIN_DB_DIR}"
            ),
        }

    if not os.path.exists(file_path):
        return {"loaded": False, "error": f"Файл не найден: {file_path}"}

    if not file_path.lower().endswith((".xlsx", ".xlsm")):
        return {"loaded": False, "error": "Нужен файл Excel (.xlsx или .xlsm)"}

    close_all_connections()

    registry.migrate_legacy_if_needed()
    instance_id = registry.new_instance_id()
    target_db, build_path, target_meta = registry.paths_for_load(instance_id)

    try:
        if sheet_name is None:
            xl = pd.ExcelFile(file_path, engine='openpyxl')
            actual_sheet = xl.sheet_names[0]
            xl.close()
        else:
            actual_sheet = sheet_name

        df = pd.read_excel(
            file_path,
            sheet_name=actual_sheet,
            engine='openpyxl',
            dtype=object,
        )

        columns = list(df.columns)
        df = _overlay_passport_columns_from_workbook(df, columns, file_path, actual_sheet)
        df = _fix_passport_columns_in_dataframe(df, columns)
        key_columns = _match_key_columns(columns)
        col_count = len(columns)

        # Sanitize column names for SQLite
        col_mapping = {}  # original -> sanitized
        for col in columns:
            sanitized = _sanitize_col_name(col)
            # Ensure uniqueness
            base = sanitized
            counter = 1
            while sanitized in col_mapping.values():
                sanitized = f"{base}_{counter}"
                counter += 1
            col_mapping[col] = sanitized

        if os.path.exists(build_path):
            _remove_path_with_retry(build_path)

        conn = sqlite3.connect(build_path)

        # Sanitize column names for SQLite
        col_mapping = {}
        sanitized_names = []
        for col in columns:
            sanitized = _sanitize_col_name(col)
            base = sanitized
            counter = 1
            while sanitized in sanitized_names:
                sanitized = f"{base}_{counter}"
                counter += 1
            col_mapping[col] = sanitized
            sanitized_names.append(sanitized)

        # Rename DataFrame columns to sanitized names
        df_renamed = df.rename(columns=col_mapping)

        # Write to SQLite using pandas to_sql (very fast)
        df_renamed.to_sql('employees', conn, if_exists='replace', index=False, chunksize=10000)

        # Free the DataFrame memory
        del df
        del df_renamed

        row_count = conn.execute('SELECT COUNT(*) FROM employees').fetchone()[0]

        # Create indexes on key columns for fast search
        for kc in key_columns:
            if kc in col_mapping:
                idx_name = f"idx_{col_mapping[kc]}"
                conn.execute(f'CREATE INDEX IF NOT EXISTS "{idx_name}" ON employees ("{col_mapping[kc]}")')
        conn.commit()
        conn.close()

        close_all_connections()
        if os.path.exists(target_db):
            _remove_path_with_retry(target_db)
        os.replace(build_path, target_db)

        # Save metadata
        meta = {
            "instance_id": instance_id,
            "source_excel": file_path,
            "file_path": file_path,
            "sheet_name": actual_sheet,
            "columns": columns,
            "key_columns": key_columns,
            "col_mapping": col_mapping,
            "loaded_at": datetime.now().isoformat(),
            "row_count": row_count,
            "col_count": col_count,
        }
        with open(target_meta, 'w', encoding='utf-8') as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)

        reg_info = registry.register_instance(
            instance_id,
            source_excel=file_path,
            meta=meta,
            set_active=set_active or registry.get_active_id() is None,
        )

        if reg_info.get("is_active"):
            _set_cache({
                "loaded": True,
                "file_path": file_path,
                "sheet_name": actual_sheet,
                "columns": columns,
                "key_columns": key_columns,
                "col_mapping": col_mapping,
                "loaded_at": meta["loaded_at"],
                "row_count": row_count,
                "col_count": col_count,
            })
            try:
                import tickets_costs

                tickets_costs._main_employees_cache["raw_df"] = None
                tickets_costs._main_employees_cache["mtime"] = 0.0
            except Exception:
                pass
        else:
            current = _get_cache()
            current["loaded"] = current.get("loaded", False)
            _set_cache(current)

        return {
            "loaded": True,
            "instance_id": instance_id,
            "is_active": reg_info.get("is_active", False),
            "source_excel": file_path,
            "file_path": file_path,
            "sheet_name": actual_sheet,
            "columns": columns,
            "key_columns": key_columns,
            "row_count": row_count,
            "col_count": col_count,
            "loaded_at": meta["loaded_at"],
            "message": (
                f"Создана база «{os.path.basename(file_path)}» ({row_count:,} строк). "
                + ("Она активна." if reg_info.get("is_active") else "Нажмите «Задействовать» в настройках.")
            ).replace(",", " "),
        }

    except Exception as e:
        close_all_connections()
        if os.path.exists(build_path):
            try:
                _remove_path_with_retry(build_path)
            except Exception:
                pass
        return {"loaded": False, "error": f"Failed to load file: {str(e)}"}


def _load_meta_from_disk():
    """Load metadata from disk if cache is empty but database exists."""
    if _get_cache()["loaded"]:
        return True
    meta_path = _meta_path()
    db_path = _db_path()
    if os.path.exists(meta_path) and os.path.exists(db_path):
        try:
            with open(meta_path, 'r', encoding='utf-8') as f:
                meta = json.load(f)
            _set_cache({
                "loaded": True,
                "file_path": meta.get("source_excel") or meta["file_path"],
                "sheet_name": meta["sheet_name"],
                "columns": meta["columns"],
                "key_columns": meta["key_columns"],
                "col_mapping": meta["col_mapping"],
                "loaded_at": meta["loaded_at"],
                "row_count": meta["row_count"],
                "col_count": meta["col_count"],
            })
            return True
        except Exception:
            return False
    return False


def is_loaded() -> bool:
    if _get_cache()["loaded"]:
        return True
    return _load_meta_from_disk()


def invalidate_cache() -> None:
    """Сбросить кэш после изменения SQLite/meta (например, справочники)."""
    _invalidate_cache()


def get_status() -> Dict[str, Any]:
    registry.migrate_legacy_if_needed()
    detected = _detect_main_db_file()
    active_id = registry.get_active_id()
    base = {
        "data_dir": MAIN_DB_DIR,
        "upload_dir": MAIN_DB_DIR,
        "search_dirs": main_db_search_dirs(),
        "detected_excel": detected,
        "active_instance_id": active_id,
        "instances": registry.list_instances(),
    }
    if not is_loaded():
        hint = detected or ""
        return {
            "loaded": False,
            "message": (
                f"Загрузите базу в Настройках → БАЗА или положите .xlsx в {MAIN_DB_DIR}."
                + (f" Найден в upload: {os.path.basename(hint)}" if hint else "")
            ),
            **base,
        }
    src = _get_cache()["file_path"]
    return {
        "loaded": True,
        "file_path": src,
        "source_excel": src,
        "file_name": os.path.basename(src) if src else "",
        "sheet_name": _get_cache()["sheet_name"],
        "columns": _get_cache()["columns"],
        "key_columns": _get_cache()["key_columns"],
        "row_count": _get_cache()["row_count"],
        "col_count": _get_cache()["col_count"],
        "loaded_at": _get_cache()["loaded_at"],
        "active_instance_id": active_id,
        "message": f"Активная база: {os.path.basename(src)} ({_get_cache()['row_count']:,} строк)".replace(",", " "),
        **base,
    }


def get_columns() -> Dict[str, Any]:
    if not is_loaded():
        return {"error": "Main database not loaded", "columns": []}

    key_col_set = set(_get_cache()["key_columns"])
    columns_info = []
    for idx, col_name in enumerate(_get_cache()["columns"]):
        columns_info.append({
            "name": col_name,
            "index": idx,
            "is_key": col_name in key_col_set,
        })

    return {
        "columns": columns_info,
        "total_columns": len(columns_info),
        "key_column_count": len(_get_cache()["key_columns"]),
    }


def get_data(
    offset: int = 0,
    limit: int = 100,
    search: Optional[str] = None,
    filters: Optional[Dict[str, str]] = None,
    sort_column: Optional[str] = None,
    sort_ascending: bool = True,
    key_columns_only: bool = False,
) -> Dict[str, Any]:
    if not is_loaded():
        return {"error": "Main database not loaded", "data": [], "total_rows": 0}

    col_mapping = _get_cache()["col_mapping"]

    # Build column list
    if key_columns_only:
        display_cols = [c for c in _get_cache()["key_columns"]]
    else:
        display_cols = list(_get_cache()["columns"])

    select_cols = [f'"{col_mapping[c]}"' for c in display_cols]
    select_clause = ", ".join(select_cols)

    # Build WHERE clause
    where_parts = []
    params = []

    if search:
        search_lower = search.lower()
        search_conditions = []
        for kc in _get_cache()["key_columns"]:
            search_conditions.append(f'LOWER("{col_mapping[kc]}") LIKE ?')
            params.append(f'%{search_lower}%')
        where_parts.append(f'({" OR ".join(search_conditions)})')

    if filters:
        for col_name, filter_value in filters.items():
            if col_name in col_mapping:
                where_parts.append(f'LOWER("{col_mapping[col_name]}") LIKE ?')
                params.append(f'%{filter_value.lower()}%')

    where_clause = f'WHERE {" AND ".join(where_parts)}' if where_parts else ''

    # Build ORDER BY
    order_clause = ''
    if sort_column is not None:
        if sort_column in col_mapping:
            order_col = col_mapping[sort_column]
        else:
            try:
                col_idx = int(sort_column)
                if 0 <= col_idx < len(display_cols):
                    order_col = col_mapping[display_cols[col_idx]]
                else:
                    order_col = None
            except (ValueError, TypeError):
                order_col = None

        if order_col:
            direction = 'ASC' if sort_ascending else 'DESC'
            order_clause = f'ORDER BY "{order_col}" {direction}'

    # Get total count with filters
    conn = _get_db_connection()
    try:
        count_sql = f'SELECT COUNT(*) FROM employees {where_clause}'
        total_filtered = conn.execute(count_sql, params).fetchone()[0]

        # Get data
        data_sql = f'SELECT {select_clause} FROM employees {where_clause} {order_clause} LIMIT ? OFFSET ?'
        data_params = params + [limit, offset]
        rows = conn.execute(data_sql, data_params).fetchall()

        # Convert to cell format
        data = []
        for row_idx, row in enumerate(rows):
            row_data = []
            display_row = offset + row_idx + 2
            for col_idx, col_name in enumerate(display_cols):
                value = row[col_idx]
                json_value = _convert_value_for_json(value) if value is not None else None
                if _is_passport_series_col(col_name) or _is_passport_number_col(col_name):
                    json_value = _format_passport_value(value, col_name) or None
                elif json_value is not None and isinstance(json_value, str):
                    try:
                        if '.' in json_value:
                            json_value = float(json_value)
                        else:
                            json_value = int(json_value)
                    except (ValueError, TypeError):
                        pass
                cell_info = {
                    "row": display_row,
                    "col": col_idx + 1,
                    "value": json_value,
                    "type": _get_cell_type(json_value),
                    "column": col_name,
                }
                row_data.append(cell_info)
            data.append(row_data)

        has_more = (offset + limit) < total_filtered

        return {
            "data": data,
            "total_rows": total_filtered,
            "total_unfiltered_rows": _get_cache()["row_count"],
            "offset": offset,
            "limit": limit,
            "returned_rows": len(data),
            "has_more": has_more,
            "columns": display_cols,
            "key_columns_only": key_columns_only,
        }
    finally:
        conn.close()


def get_stats() -> Dict[str, Any]:
    if not is_loaded():
        return {"error": "Main database not loaded"}

    conn = _get_db_connection()
    try:
        stats = {
            "total_rows": _get_cache()["row_count"],
            "total_columns": _get_cache()["col_count"],
            "key_column_count": len(_get_cache()["key_columns"]),
            "file_path": _get_cache()["file_path"],
            "sheet_name": _get_cache()["sheet_name"],
            "loaded_at": _get_cache()["loaded_at"],
        }

        col_mapping = _get_cache()["col_mapping"]
        key_col_stats = {}

        for col_name in _get_cache()["key_columns"]:
            if col_name in col_mapping:
                s_col = col_mapping[col_name]
                row_count = conn.execute(f'SELECT COUNT(*) FROM employees WHERE "{s_col}" IS NOT NULL').fetchone()[0]
                null_count = _get_cache()["row_count"] - row_count
                unique_count = conn.execute(f'SELECT COUNT(DISTINCT "{s_col}") FROM employees').fetchone()[0]

                top_values = conn.execute(
                    f'SELECT "{s_col}", COUNT(*) as cnt FROM employees WHERE "{s_col}" IS NOT NULL GROUP BY "{s_col}" ORDER BY cnt DESC LIMIT 10'
                ).fetchall()

                key_col_stats[col_name] = {
                    "unique_count": unique_count,
                    "null_count": null_count,
                    "non_null_count": row_count,
                    "top_values": [{"value": r[0], "count": r[1]} for r in top_values],
                }

        stats["key_column_stats"] = key_col_stats

        # All columns overview
        all_col_stats = []
        for col_name in _get_cache()["columns"]:
            if col_name in col_mapping:
                s_col = col_mapping[col_name]
                null_count = conn.execute(f'SELECT COUNT(*) FROM employees WHERE "{s_col}" IS NULL').fetchone()[0]
                unique_count = conn.execute(f'SELECT COUNT(DISTINCT "{s_col}") FROM employees').fetchone()[0]
                all_col_stats.append({
                    "name": col_name,
                    "dtype": "text",
                    "unique_count": unique_count,
                    "null_count": null_count,
                    "is_key": col_name in set(_get_cache()["key_columns"]),
                })

        stats["all_columns"] = all_col_stats
        db_path = _db_path()
        if os.path.isfile(db_path):
            stats["memory_usage_mb"] = round(os.path.getsize(db_path) / (1024 * 1024), 2)

        return stats
    finally:
        conn.close()


def search_advanced(
    query: Optional[str] = None,
    key_columns_only: bool = False,
    exact_match: bool = False,
    offset: int = 0,
    limit: int = 100,
) -> Dict[str, Any]:
    """Search using SQL LIKE on key columns."""
    if not is_loaded():
        return {"error": "Main database not loaded", "results": [], "total_rows": 0}

    if query is None:
        return {"error": "Query parameter is required", "results": [], "total_rows": 0}

    # Just use get_data with search - same implementation
    result = get_data(
        offset=offset,
        limit=limit,
        search=query,
        key_columns_only=key_columns_only,
    )

    # Rename 'data' to 'results' for compatibility
    result["results"] = result.pop("data", [])
    result["total_matched"] = result.get("total_rows", 0)
    result["query"] = query
    result["exact_match"] = exact_match

    return result


def clear_cache() -> Dict[str, Any]:
    """Сброс только кэша в памяти (файлы instances не удаляются)."""
    file_path = _get_cache().get("file_path")
    close_all_connections()
    _set_cache({
        "loaded": False,
        "file_path": None,
        "sheet_name": None,
        "columns": [],
        "key_columns": [],
        "col_mapping": {},
        "loaded_at": None,
        "row_count": 0,
        "col_count": 0,
    })
    return {"cleared": True, "previous_file": file_path}


def activate_instance(instance_id: str) -> Dict[str, Any]:
    result = registry.activate_instance(instance_id)
    if not result.get("ok"):
        return result
    close_all_connections()
    _invalidate_cache()
    if _load_meta_from_disk():
        return {**get_status(), "activated": instance_id}
    return {"ok": True, "activated": instance_id, "loaded": False}


def list_instances() -> Dict[str, Any]:
    registry.migrate_legacy_if_needed()
    return {"instances": registry.list_instances(), "active_id": registry.get_active_id()}


def delete_instance(instance_id: str) -> Dict[str, Any]:
    active = registry.get_active_id()
    result = registry.delete_instance(instance_id)
    close_all_connections()
    if active == instance_id:
        _invalidate_cache()
        _load_meta_from_disk()
    return result


def verify_instance(instance_id: str) -> Dict[str, Any]:
    db = registry.instance_db_path(instance_id)
    meta_path = registry.instance_meta_path(instance_id)
    if not os.path.isfile(db):
        return {"ok": False, "error": "База не найдена"}
    meta = {}
    if os.path.isfile(meta_path):
        with open(meta_path, "r", encoding="utf-8") as f:
            meta = json.load(f)
    conn = sqlite3.connect(db, timeout=30.0)
    try:
        row_count = conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
        col_count = len(meta.get("columns") or [])
        return {
            "ok": True,
            "instance_id": instance_id,
            "row_count": row_count,
            "col_count": col_count,
            "file_name": os.path.basename(meta.get("source_excel") or ""),
            "loaded_at": meta.get("loaded_at"),
            "size_mb": round(os.path.getsize(db) / (1024 * 1024), 2),
        }
    finally:
        conn.close()


def export_instance_to_excel(instance_id: str) -> Dict[str, Any]:
    db = registry.instance_db_path(instance_id)
    if not os.path.isfile(db):
        return {"ok": False, "error": "База не найдена"}
    meta = {}
    meta_path = registry.instance_meta_path(instance_id)
    if os.path.isfile(meta_path):
        with open(meta_path, "r", encoding="utf-8") as f:
            meta = json.load(f)
    base_name = os.path.basename(meta.get("source_excel") or instance_id)
    if base_name.lower().endswith((".xlsx", ".xlsm")):
        base_name = os.path.splitext(base_name)[0]
    out_name = f"{base_name}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    os.makedirs(registry.EXPORTS_DIR, exist_ok=True)
    out_path = os.path.join(registry.EXPORTS_DIR, out_name)
    conn = sqlite3.connect(db, timeout=60.0)
    try:
        df = pd.read_sql("SELECT * FROM employees", conn)
        df.to_excel(out_path, index=False, engine="openpyxl")
    finally:
        conn.close()
    return {"ok": True, "export_path": out_path, "file_name": out_name}
