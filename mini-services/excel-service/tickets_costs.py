"""
Затраты по билетам — загрузка (строка 4 заголовков), обработка, дедупликация, обогащение из Базы, дашборд.
"""

from __future__ import annotations

import json
import os
import re
import shutil
import sqlite3
import threading
import time
import uuid
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd

import excel_handler
import excel_libs
from excel_libs import get_sheet_names_universal, read_dataframe, write_dataframe
import integration_ops
import main_db
import tickets_db
from data_paths import UPLOAD_DIR

HEADER_ROW_INDEX = 3  # 4-я строка Excel
HEADER_ROW_1BASED = HEADER_ROW_INDEX + 1
PREVIEW_ROW_LIMIT = 100
# Листы отчётов (Геленджик и др.), не исходники «Затраты по билетам»
TICKETS_SHEET_BLOCKLIST = frozenset(
    {
        "Путь сотрудника",
        "График присутствия",
        "Сводка",
        "Периоды в базе",
        "Данные",
    }
)

_registry_db_locks: Dict[str, threading.RLock] = {}
_registry_db_locks_guard = threading.Lock()
_open_registry_connections: Dict[str, set] = {}
_registry_clearing: set[str] = set()


class _LockedSqliteConnection:
    """sqlite3.Connection wrapper: close() releases per-registry lock."""

    __slots__ = ("_conn", "_lock", "_registry", "_closed")

    def __init__(self, conn: sqlite3.Connection, lock: threading.RLock, registry: str) -> None:
        self._conn = conn
        self._lock = lock
        self._registry = registry
        self._closed = False
        with _registry_db_locks_guard:
            _open_registry_connections.setdefault(registry, set()).add(self)

    def close(self) -> None:
        if self._closed:
            return
        self._closed = True
        try:
            with _registry_db_locks_guard:
                bucket = _open_registry_connections.get(self._registry)
                if bucket is not None:
                    bucket.discard(self)
            self._conn.close()
        finally:
            self._lock.release()

    def __del__(self) -> None:
        try:
            self.close()
        except Exception:
            pass

    def __getattr__(self, name: str) -> Any:
        return getattr(self._conn, name)


def _force_close_registry_connections(registry: str) -> int:
    """Закрыть все открытые соединения реестра (перед удалением файла БД)."""
    reg = tickets_db.normalize_registry(registry)
    with _registry_db_locks_guard:
        conns = list(_open_registry_connections.get(reg, set()))
    closed = 0
    for wrapper in conns:
        try:
            wrapper.close()
            closed += 1
        except Exception:
            pass
    if closed:
        time.sleep(0.15)
    return closed


def _db_lock(registry: str) -> threading.RLock:
    reg = tickets_db.normalize_registry(registry)
    with _registry_db_locks_guard:
        if reg not in _registry_db_locks:
            _registry_db_locks[reg] = threading.RLock()
        return _registry_db_locks[reg]


REGISTRY_VSM = "vsm"
REGISTRY_SK = "sk"
REGISTRY_LABELS = {
    REGISTRY_VSM: "ВелесстройМонтаж",
    REGISTRY_SK: "Стройконстракшен",
}

# Индексы колонок исходного Excel (A=0, заголовки — 4-я строка). QWER2–QWER18.
RAW_COL_MAP = {
    0: "nakladnaya",  # A — не в таблице отображения
    1: "vid_uslugi",  # B QWER2
    2: "podrazdelenie",  # C QWER3
    3: "obosnovanie_pereleta",  # D QWER4
    4: "organizaciya",  # E QWER5
    5: "klassifikaciya",  # F QWER6
    6: "operaciya",  # G QWER7
    7: "fio",  # H QWER8
    8: "tab_nomer",  # I QWER9
    9: "pasport",  # J QWER10
    10: "napravlenie_gorod",  # K QWER11
    11: "marshrut",  # L QWER12
    12: "data_vyleta_plan",  # M QWER13
    13: "data_prileta_plan",  # N QWER14
    14: "data_vypiski",  # O QWER15
    15: "nomer_bileta",  # P QWER16
    16: "aviaperevozchik",  # Q QWER17
    23: "summa_bilet_ag",  # X QWER18
}

# Только нужные колонки A–Q и X (без лишних колонок Excel → в разы быстрее чтение)
RAW_COL_INDICES: Tuple[int, ...] = tuple(sorted(RAW_COL_MAP.keys()))
RAW_COL_NAMES: Tuple[str, ...] = tuple(RAW_COL_MAP[i] for i in RAW_COL_INDICES)

PROCESSED_COLUMNS = [
    ("nomer_bileta", "Номер билета"),
    ("vid_uslugi", "Вид услуги"),
    ("tabelyny_1c", "Табельный_1С"),
    ("fio_1c", "ФИО_1С"),
    ("ploshchadka", "Площадка"),
    ("podrazdelenie", "Подразделение"),
    ("obosnovanie_pereleta", "Обоснование перелета"),
    ("organizaciya", "Организация"),
    ("klassifikaciya", "Классификация сотрудников"),
    ("operaciya", "Операция"),
    ("fio", "Ф.И.О."),
    ("tab_nomer", "Табельный номер"),
    ("pasport", "Паспорт"),
    ("marshrut", "Маршрут"),
    ("napravlenie_gorod", "Направление город"),
    ("marshrut_data_vyleta", "Маршрут + Дата вылета"),
    ("data_vypiski", "Дата выписки билета"),
    ("summa_pokupka", "Сумма за Покупку"),
    ("summa_obmen", "Сумма за Обмен"),
    ("summa_vozvrat_sbor", "Сумма за Возврат+Сбор"),
    ("summa_total", "Общая сумма затрат"),
    ("data_vyleta_plan", "Планируемая дата вылета"),
    ("data_prileta_plan", "Планируемая дата прилета"),
    ("aviaperevozchik", "Авиаперевозчик"),
]

TEXT_MERGE_COLS = {
    "vid_uslugi",
    "ploshchadka",
    "podrazdelenie",
    "obosnovanie_pereleta",
    "organizaciya",
    "klassifikaciya",
    "operaciya",
    "fio",
    "tab_nomer",
    "pasport",
    "marshrut",
    "napravlenie_gorod",
    "marshrut_data_vyleta",
    "aviaperevozchik",
}
SUM_COLS = {"summa_pokupka", "summa_obmen", "summa_vozvrat_sbor", "summa_total"}
DATE_COLS = {"data_vypiski", "data_vyleta_plan", "data_prileta_plan"}
OPERATION_COL = "operaciya"
OPERATION_COL_IDX = 6
RETURN_OPS = frozenset({"возврат", "сбор поставщика (возврат)"})
YELLOW_FLAG = "_conflict"
EDITABLE_COL_KEYS = {c[0] for c in PROCESSED_COLUMNS}


def _assign_row_ids(df: pd.DataFrame, registry: str) -> pd.DataFrame:
    """Стабильный идентификатор строки для редактирования в UI."""
    out = df.copy()
    ids: List[str] = []
    for i, row in out.iterrows():
        a = _norm_ticket_num(row.get("nomer_bileta"))
        b = _norm_ticket_num(row.get("nomer_bileta_obmen"))
        ids.append(f"{registry}|{a}|{b}|{i}")
    out["_row_id"] = ids
    return out


def _ensure_row_ids(df: pd.DataFrame, registry: str) -> pd.DataFrame:
    if df.empty:
        return df
    if "_row_id" not in df.columns:
        return _assign_row_ids(df, registry)
    return df


def _paths(registry: str) -> Dict[str, str]:
    reg = tickets_db.normalize_registry(registry)
    return {
        "db": os.path.join(UPLOAD_DIR, f"tickets_costs_{reg}.sqlite"),
        "meta": os.path.join(UPLOAD_DIR, f"tickets_costs_{reg}_meta.json"),
    }


def _sources_dir(registry: str) -> str:
    reg = tickets_db.normalize_registry(registry)
    path = os.path.join(UPLOAD_DIR, f"tickets_costs_{reg}", "sources")
    os.makedirs(path, exist_ok=True)
    return path


def _ensure_schema(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS source_files (
            file_id TEXT PRIMARY KEY,
            original_name TEXT NOT NULL,
            stored_path TEXT NOT NULL,
            uploaded_at TEXT NOT NULL,
            row_count INTEGER DEFAULT 0
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS processing_runs (
            run_id TEXT PRIMARY KEY,
            run_type TEXT NOT NULL,
            label TEXT NOT NULL,
            created_at TEXT NOT NULL,
            row_count INTEGER DEFAULT 0,
            active INTEGER DEFAULT 0
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS upload_queue (
            id TEXT PRIMARY KEY,
            original_name TEXT NOT NULL,
            file_path TEXT NOT NULL,
            file_id TEXT,
            added_at TEXT NOT NULL
        )
        """
    )


def _remove_path(path: str, retries: int = 12) -> bool:
    if not path or not os.path.exists(path):
        return True
    for attempt in range(retries):
        try:
            if os.path.isdir(path):
                shutil.rmtree(path, ignore_errors=False)
            else:
                os.remove(path)
            return True
        except OSError:
            time.sleep(min(2.0, 0.25 * (attempt + 1)))
    return False


def _wipe_registry_db(db_path: str) -> None:
    """Сброс WAL перед удалением файлов БД (без DROP — файлы удаляются в clear_registry)."""
    if not os.path.isfile(db_path):
        return
    last_err: Optional[Exception] = None
    for attempt in range(20):
        conn: Optional[sqlite3.Connection] = None
        try:
            conn = sqlite3.connect(db_path, timeout=120.0, isolation_level=None)
            conn.execute("PRAGMA busy_timeout=120000")
            conn.execute("BEGIN IMMEDIATE")
            conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            conn.execute("COMMIT")
            return
        except sqlite3.OperationalError as exc:
            last_err = exc
            if "locked" not in str(exc).lower():
                raise
            time.sleep(min(3.0, 0.3 * (attempt + 1)))
        finally:
            if conn is not None:
                try:
                    conn.close()
                except OSError:
                    pass
    if last_err is not None:
        raise last_err


def _conn(registry: str) -> sqlite3.Connection:
    reg = tickets_db.normalize_registry(registry)
    deadline = time.monotonic() + 130.0
    while True:
        with _registry_db_locks_guard:
            clearing = reg in _registry_clearing
        if not clearing:
            break
        if time.monotonic() >= deadline:
            raise sqlite3.OperationalError(f"Реестр {reg} очищается — повторите через несколько секунд")
        time.sleep(0.05)

    lock = _db_lock(reg)
    lock.acquire()
    try:
        paths = _paths(reg)
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        raw = sqlite3.connect(paths["db"], timeout=120.0)
        raw.row_factory = sqlite3.Row
        raw.execute("PRAGMA journal_mode=WAL")
        raw.execute("PRAGMA synchronous=NORMAL")
        raw.execute("PRAGMA busy_timeout=120000")
        _ensure_schema(raw)
        return _LockedSqliteConnection(raw, lock, reg)  # type: ignore[return-value]
    except Exception:
        lock.release()
        raise


def _snapshot_table(run_id: str) -> str:
    safe = re.sub(r"[^a-zA-Z0-9_]", "_", run_id)
    return f"snapshot_{safe}"


def _save_processing_run(
    registry: str,
    run_type: str,
    df: pd.DataFrame,
    note: str = "",
) -> str:
    reg = tickets_db.normalize_registry(registry)
    run_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    label = RUN_LABELS.get(run_type, run_type)
    if note:
        label = f"{label} — {note}"

    work = df.copy()
    for col in (YELLOW_FLAG, "_dedupe_key", "_year", "_month"):
        if col in work.columns:
            work = work.drop(columns=[col])

    conn = _conn(reg)
    try:
        conn.execute("UPDATE processing_runs SET active = 0")
        _sanitize_df_for_sqlite(work).to_sql(_snapshot_table(run_id), conn, if_exists="replace", index=False)
        conn.execute(
            """
            INSERT INTO processing_runs (run_id, run_type, label, created_at, row_count, active)
            VALUES (?, ?, ?, ?, ?, 1)
            """,
            (run_id, run_type, label, datetime.now().isoformat(), int(len(work))),
        )
        conn.commit()
    finally:
        conn.close()
    return run_id


def list_processing_runs(registry: str) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    conn = _conn(reg)
    try:
        rows = conn.execute(
            """
            SELECT run_id, run_type, label, created_at, row_count, active
            FROM processing_runs
            ORDER BY created_at DESC
            """
        ).fetchall()
    except Exception:
        rows = []
    finally:
        conn.close()
    return {
        "registry": reg,
        "runs": [dict(r) for r in rows],
    }


def activate_processing_run(registry: str, run_id: str) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    table = _snapshot_table(run_id)
    conn = _conn(reg)
    try:
        df = pd.read_sql_query(f'SELECT * FROM "{table}"', conn)
        if df.empty:
            return {"error": "Снимок пустой или не найден"}
        df = _ensure_row_ids(df, reg)
        _write_table(conn, "processed", df)
        conn.execute("UPDATE processing_runs SET active = 0")
        conn.execute("UPDATE processing_runs SET active = 1 WHERE run_id = ?", (run_id,))
        conn.commit()
    except Exception as e:
        return {"error": f"Не удалось открыть снимок: {e}"}
    finally:
        conn.close()

    meta = _load_meta(reg)
    meta["processed_rows"] = int(len(df))
    meta["updated_at"] = datetime.now().isoformat()
    _save_meta(reg, meta)
    return {"success": True, "registry": reg, "run_id": run_id, "processed_rows": int(len(df))}


def delete_processing_run(registry: str, run_id: str) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    table = _snapshot_table(run_id)
    conn = _conn(reg)
    try:
        conn.execute(f'DROP TABLE IF EXISTS "{table}"')
        conn.execute("DELETE FROM processing_runs WHERE run_id = ?", (run_id,))
        conn.commit()
    finally:
        conn.close()
    return {"success": True, "registry": reg, "run_id": run_id}


def get_run_data(
    registry: str,
    run_id: str,
    offset: int = 0,
    limit: int = 0,
) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    table = _snapshot_table(run_id)
    conn = _conn(reg)
    try:
        df = pd.read_sql_query(f'SELECT * FROM "{table}"', conn)
    except Exception:
        return {"error": "Снимок не найден"}
    finally:
        conn.close()
    if df.empty:
        return {
            "columns": DISPLAY_COLUMNS,
            "data": [],
            "total": 0,
            "offset": offset,
            "limit": limit,
            "run_id": run_id,
        }
    df = _ensure_row_ids(df, reg)
    total = len(df)
    if limit and limit > 0:
        chunk = df.iloc[offset : offset + limit]
    else:
        chunk = df.iloc[offset:]
    records = chunk.replace({np.nan: None}).to_dict(orient="records")
    return {
        "columns": DISPLAY_COLUMNS,
        "data": records,
        "total": total,
        "offset": offset,
        "limit": limit,
        "run_id": run_id,
    }


def list_source_files(registry: str) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    conn = _conn(reg)
    try:
        rows = conn.execute(
            """
            SELECT file_id, original_name, stored_path, uploaded_at, row_count
            FROM source_files
            ORDER BY uploaded_at DESC
            """
        ).fetchall()
    except Exception:
        rows = []
    finally:
        conn.close()
    return {"registry": reg, "files": [dict(r) for r in rows]}


def preview_source_file(registry: str, file_id: str, limit: int = PREVIEW_ROW_LIMIT) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    conn = _conn(reg)
    try:
        row = conn.execute(
            "SELECT stored_path, original_name FROM source_files WHERE file_id = ?",
            (file_id,),
        ).fetchone()
    finally:
        conn.close()
    if not row:
        return {"error": "Файл не найден"}
    stored_path = row["stored_path"]
    if not os.path.isfile(stored_path):
        return {"error": "Файл на диске отсутствует"}
    try:
        raw = _read_raw_excel(stored_path)
    except Exception as e:
        return {"error": str(e)}
    preview = raw.head(limit)
    cols = [{"key": str(i), "title": str(c)} for i, c in enumerate(preview.columns)]
    data = []
    for _, r in preview.iterrows():
        data.append({str(i): _norm_text(r.iloc[i]) if i < len(r) else "" for i in range(len(preview.columns))})
    return {
        "file_id": file_id,
        "original_name": row["original_name"],
        "stored_path": stored_path,
        "columns": cols,
        "data": data,
        "total_rows": int(len(raw)),
        "preview_rows": int(len(preview)),
    }


def delete_source_file(registry: str, file_id: str) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    paths = _paths(reg)
    if not os.path.isfile(paths["db"]):
        sources = _sources_dir(reg)
        for name in os.listdir(sources):
            if name.startswith(file_id):
                file_path = os.path.join(sources, name)
                if not _remove_path(file_path):
                    return {"error": f"Не удалось удалить файл: {name}"}
                return {"success": True, "file_id": file_id}
        return {"error": "Файл не найден"}

    conn = _conn(reg)
    try:
        row = conn.execute(
            "SELECT stored_path FROM source_files WHERE file_id = ?",
            (file_id,),
        ).fetchone()
        if not row:
            return {"error": "Файл не найден в базе"}
        path = row["stored_path"]
        if path and os.path.isfile(path) and not _remove_path(path):
            return {"error": "Файл занят другим процессом — закройте Excel и повторите"}
        conn.execute("DELETE FROM source_files WHERE file_id = ?", (file_id,))
        conn.commit()
    finally:
        conn.close()
    return {"success": True, "file_id": file_id}


def _load_meta(registry: str) -> Dict[str, Any]:
    paths = _paths(registry)
    if os.path.exists(paths["meta"]):
        with open(paths["meta"], "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_meta(registry: str, meta: Dict[str, Any]) -> None:
    paths = _paths(registry)
    os.makedirs(os.path.dirname(paths["meta"]), exist_ok=True)
    with open(paths["meta"], "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)


def _upload_queue(meta: Dict[str, Any]) -> List[Dict[str, Any]]:
    raw = meta.get("upload_queue")
    return list(raw) if isinstance(raw, list) else []


def _migrate_meta_queue_to_db(reg: str, conn: sqlite3.Connection) -> None:
    """Один раз перенести очередь из meta.json в SQLite."""
    count = conn.execute("SELECT COUNT(*) FROM upload_queue").fetchone()[0]
    if count:
        return
    meta = _load_meta(reg)
    for q in _upload_queue(meta):
        path = str(q.get("path") or "").strip()
        if not path:
            continue
        conn.execute(
            """
            INSERT OR IGNORE INTO upload_queue
            (id, original_name, file_path, file_id, added_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                str(q.get("id") or f"q_{uuid.uuid4().hex[:10]}"),
                str(q.get("name") or os.path.basename(path)),
                path,
                q.get("file_id"),
                q.get("added_at") or datetime.now().isoformat(),
            ),
        )


def _auto_recover_upload_queue(reg: str, conn: sqlite3.Connection) -> int:
    """Вернуть в очередь Excel из upload/, если очередь пуста (после F5 / сбоя)."""
    if conn.execute("SELECT COUNT(*) FROM upload_queue").fetchone()[0] > 0:
        return 0
    try:
        source_paths = {
            (r[0] or "").lower()
            for r in conn.execute("SELECT stored_path FROM source_files").fetchall()
        }
    except Exception:
        source_paths = set()
    added = 0
    now = datetime.now().isoformat()
    for f in excel_handler.list_uploaded_files(include_sheets=False):
        path = f.get("file_path") or ""
        if not path or not excel_handler.is_excel_file(path):
            continue
        if path.lower() in source_paths:
            continue
        qid = f"q_{uuid.uuid4().hex[:10]}"
        conn.execute(
            """
            INSERT OR IGNORE INTO upload_queue
            (id, original_name, file_path, file_id, added_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                qid,
                os.path.basename(path),
                path,
                f.get("file_id"),
                f.get("modified") or now,
            ),
        )
        added += 1
    if added:
        conn.commit()
    return added


def _list_upload_queue(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT id, original_name, file_path, file_id, added_at
        FROM upload_queue
        ORDER BY added_at DESC
        """
    ).fetchall()
    out: List[Dict[str, Any]] = []
    for r in rows:
        path = r["file_path"]
        fid = r["file_id"]
        resolved = _resolve_input_path(path) or (fid and _resolve_input_path(fid))
        if resolved and resolved != path:
            path = resolved
        out.append(
            {
                "id": r["id"],
                "name": r["original_name"],
                "path": path,
                "file_id": fid,
                "added_at": r["added_at"],
            },
        )
    return out


def merge_upload_queue(registry: str, items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Сохранить очередь файлов до «Загрузить в реестр» (SQLite, переживает F5)."""
    reg = tickets_db.normalize_registry(registry)
    conn = _conn(reg)
    try:
        _migrate_meta_queue_to_db(reg, conn)
        now = datetime.now().isoformat()
        for item in items or []:
            path = str(item.get("path") or "").strip()
            if not path:
                continue
            fid = item.get("file_id")
            resolved = _resolve_input_path(path) or path
            qid = str(item.get("id") or f"q_{uuid.uuid4().hex[:10]}")
            conn.execute(
                """
                INSERT OR REPLACE INTO upload_queue
                (id, original_name, file_path, file_id, added_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    qid,
                    str(item.get("name") or os.path.basename(resolved)),
                    resolved,
                    fid,
                    item.get("added_at") or now,
                ),
            )
        conn.commit()
        queue = _list_upload_queue(conn)
    finally:
        conn.close()
    return {"registry": reg, "upload_queue": queue}


def remove_upload_queue_item(registry: str, queue_id: str) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    conn = _conn(reg)
    try:
        conn.execute("DELETE FROM upload_queue WHERE id = ?", (queue_id,))
        conn.commit()
        queue = _list_upload_queue(conn)
    finally:
        conn.close()
    return {"registry": reg, "upload_queue": queue}


def _prune_upload_queue_db(
    registry: str,
    file_paths: List[str],
    stored_entries: List[Tuple[str, str, str, int]],
) -> None:
    path_keys: set[str] = set()
    for fp in file_paths:
        resolved = _resolve_input_path(fp)
        if resolved:
            path_keys.add(resolved.lower())
            path_keys.add(os.path.basename(resolved).lower())
        if fp:
            path_keys.add(str(fp).lower())
            path_keys.add(os.path.basename(str(fp)).lower())
    for _fid, base, dest, _rc in stored_entries:
        path_keys.add(dest.lower())
        path_keys.add(base.lower())
    file_ids = {e[0] for e in stored_entries}
    conn = _conn(registry)
    try:
        rows = conn.execute(
            "SELECT id, file_path, file_id FROM upload_queue",
        ).fetchall()
        for row in rows:
            path = (row["file_path"] or "").lower()
            fid = row["file_id"]
            if path in path_keys or (fid and fid in file_ids):
                conn.execute("DELETE FROM upload_queue WHERE id = ?", (row["id"],))
        conn.commit()
    finally:
        conn.close()


def _norm_text(v: Any) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    return re.sub(r"\s+", " ", s)


def _as_source_text(v: Any) -> str:
    """Текст как в исходнике: без потери ведущих нулей у строковых ячеек."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    if isinstance(v, str):
        s = v.strip()
        return "" if s.lower() in ("nan", "none", "nat") else s
    if isinstance(v, (int, np.integer)):
        return str(v)
    if isinstance(v, (float, np.floating)):
        if np.isnan(v):
            return ""
        if float(v) == int(v):
            return str(int(v))
        return str(v).strip()
    s = str(v).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


def _norm_ticket_num(v: Any) -> str:
    """Ключ дедупликации (нормализованный номер билета)."""
    s = _as_source_text(v)
    return s


def _openpyxl_cell_text(cell: Any) -> str:
    if cell is None:
        return ""
    v = getattr(cell, "value", cell)
    if v is None:
        return ""
    if getattr(cell, "data_type", None) == "s" or isinstance(v, str):
        return str(v).strip()
    return _as_source_text(v)


def _overlay_identifier_columns_from_workbook(
    df: pd.DataFrame,
    file_path: str,
    sheet_name: str,
) -> pd.DataFrame:
    """Колонки J (паспорт) и P (номер билета) — как текст из Excel."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in (".xlsx", ".xlsm") or df.empty:
        return df
    nrows = len(df)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return df
        ws = wb[sheet_name]
        passports: List[str] = []
        tickets: List[str] = []
        for row in ws.iter_rows(min_row=HEADER_ROW_1BASED + 1, values_only=False):
            cells = list(row)
            passports.append(_openpyxl_cell_text(cells[9]) if len(cells) > 9 else "")
            tickets.append(_openpyxl_cell_text(cells[15]) if len(cells) > 15 else "")
            if len(passports) >= nrows:
                break
        out = df.copy()
        if len(passports) >= nrows:
            out["pasport"] = passports[:nrows]
        if len(tickets) >= nrows:
            out["nomer_bileta"] = tickets[:nrows]
        return out
    finally:
        wb.close()


def _parse_number(v: Any) -> float:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return 0.0
    if isinstance(v, (int, float, np.integer, np.floating)):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0
    s = _norm_text(v).replace(" ", "").replace("\u00a0", "")
    if not s:
        return 0.0
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_date_series(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce", dayfirst=True)


def _is_missing_date(v: Any) -> bool:
    if v is None:
        return True
    if v is pd.NaT:
        return True
    if type(v).__name__ == "NaTType":
        return True
    if isinstance(v, float) and np.isnan(v):
        return True
    try:
        return bool(pd.isna(v))
    except (TypeError, ValueError):
        return False


def _format_date_series(series: pd.Series) -> pd.Series:
    """Формат DD.MM.YYYY без вызова strftime на скалярном NaT."""
    if series.empty:
        return series.astype(str)
    parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)
    formatted = parsed.dt.strftime("%d.%m.%Y")
    return formatted.where(parsed.notna(), "").astype(str)


def _fmt_date_ddmmyyyy(v: Any) -> str:
    if _is_missing_date(v):
        return ""
    try:
        ts = pd.to_datetime(v, errors="coerce", dayfirst=True)
    except (TypeError, ValueError):
        return _norm_text(v)
    if _is_missing_date(ts):
        return _norm_text(v)
    try:
        return pd.Timestamp(ts).strftime("%d.%m.%Y")
    except (ValueError, AttributeError, OSError):
        return _norm_text(v)


def _sanitize_cell_for_sqlite(v: Any) -> Any:
    """SQLite не принимает pd.Timestamp/numpy-типы напрямую."""
    if v is None:
        return None
    if _is_missing_date(v):
        return None
    if isinstance(v, (pd.Timestamp, datetime)) or type(v).__name__ == "NaTType":
        formatted = _fmt_date_ddmmyyyy(v)
        return formatted if formatted else None
    if isinstance(v, float) and np.isnan(v):
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        if np.isnan(v):
            return None
        return float(v)
    if isinstance(v, np.bool_):
        return bool(v)
    if isinstance(v, (bytes, bytearray)):
        return v.decode("utf-8", errors="replace")
    return v


def _sanitize_df_for_sqlite(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        if col in DATE_COLS or pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = _format_date_series(out[col])
        else:
            out[col] = out[col].map(_sanitize_cell_for_sqlite)
    return out


def _normalize_operation_value(op: Any) -> str:
    s = _norm_text(op)
    if not s:
        return ""
    low = s.lower().strip()
    if low in RETURN_OPS or (low == "возврат") or (
        "сбор" in low and "поставщика" in low and "возврат" in low
    ):
        return "Возврат+Сбор"
    return s


def _raw_df_is_named(raw_df: pd.DataFrame) -> bool:
    return "operaciya" in raw_df.columns or "vid_uslugi" in raw_df.columns


def _normalize_raw_operations_df(raw_df: pd.DataFrame) -> pd.DataFrame:
    """При объединении файлов: G «Операция» — Возврат / Сбор поставщика → Возврат+Сбор."""
    if raw_df.empty:
        return raw_df
    out = raw_df.copy()
    if _raw_df_is_named(out):
        if "operaciya" in out.columns:
            out["operaciya"] = out["operaciya"].map(_normalize_operation_value)
        return out
    if raw_df.shape[1] <= OPERATION_COL_IDX:
        return out
    col = out.iloc[:, OPERATION_COL_IDX]
    out.iloc[:, OPERATION_COL_IDX] = col.map(_normalize_operation_value)
    return out


def _operation_kind(op: Any) -> str:
    s = _norm_text(op).lower()
    if not s:
        return ""
    if "продаж" in s or "покуп" in s:
        return "pokupka"
    if "обмен" in s:
        return "obmen"
    if "возврат" in s.replace(" ", ""):
        return "vozvrat"
    return "other"


def _make_marshrut_data_vyleta(vylet: Any, marshrut: Any) -> str:
    d = _fmt_date_ddmmyyyy(vylet)
    m = _norm_text(marshrut)
    if d and m:
        return f"{d}; {m}"
    return d or m


def _alloc_operation_sums(op: Any, amount: float) -> Tuple[float, float, float, float]:
    kind = _operation_kind(op)
    r = s = t = 0.0
    if kind == "pokupka":
        r = amount
    elif kind == "obmen":
        s = amount
    elif kind == "vozvrat":
        t = amount
    return r, s, t, r + s + t


def _resolve_input_path(file_path: str) -> Optional[str]:
    """Найти файл: абсолютный путь, upload/, UUID или имя stored_filename."""
    raw = str(file_path or "").strip().strip('"').strip("'")
    if not raw:
        return None
    raw = raw.replace("/", os.sep)
    if os.path.isfile(raw):
        return os.path.abspath(raw)
    if not os.path.isabs(raw):
        by_id = excel_handler.find_file_by_id(raw)
        if by_id and os.path.isfile(by_id):
            return by_id
        in_upload = os.path.join(UPLOAD_DIR, raw)
        if os.path.isfile(in_upload):
            return in_upload
        base = os.path.basename(raw)
        if base != raw:
            in_upload = os.path.join(UPLOAD_DIR, base)
            if os.path.isfile(in_upload):
                return in_upload
        if os.path.isdir(UPLOAD_DIR):
            for name in os.listdir(UPLOAD_DIR):
                if name == base or name.startswith(f"{raw}_") or name.endswith(base):
                    candidate = os.path.join(UPLOAD_DIR, name)
                    if os.path.isfile(candidate):
                        return candidate
    return None


def _looks_like_tickets_header(cells: List[str]) -> bool:
    if len(cells) < 7:
        return False
    joined = " ".join(cells).upper()
    if "QWER" in joined:
        return True
    g = cells[6].lower()
    b = cells[1].lower() if len(cells) > 1 else ""
    if "операц" in g:
        return True
    if "вид" in b and "услуг" in b:
        return True
    return False


def _peek_header_row(file_path: str, sheet_name: str) -> List[str]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
            for row in ws.iter_rows(
                min_row=HEADER_ROW_1BASED,
                max_row=HEADER_ROW_1BASED,
                min_col=1,
                max_col=24,
                values_only=True,
            ):
                return [str(c).strip() if c is not None else "" for c in (row or ())]
        finally:
            wb.close()
        return []
    try:
        engine: Any = "calamine" if ext in (".xlsx", ".xlsm") else None
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            nrows=HEADER_ROW_1BASED,
            engine=engine,
        )
        if df.shape[0] <= HEADER_ROW_INDEX:
            return []
        return [str(c).strip() if pd.notna(c) else "" for c in df.iloc[HEADER_ROW_INDEX].tolist()]
    except Exception:
        return []


def _sheet_has_enough_rows(file_path: str, sheet_name: str) -> bool:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return False
            return int(wb[sheet_name].max_row or 0) > HEADER_ROW_1BASED
        finally:
            wb.close()
    try:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            nrows=HEADER_ROW_1BASED + 1,
        )
        return df.shape[0] > HEADER_ROW_INDEX
    except Exception:
        return False


def _pick_tickets_sheet(file_path: str) -> Any:
    try:
        names = get_sheet_names_universal(file_path)
    except Exception:
        return 0
    if not names:
        return 0
    candidates = [n for n in names if n not in TICKETS_SHEET_BLOCKLIST]
    search = candidates if candidates else list(names)
    for name in search:
        if _looks_like_tickets_header(_peek_header_row(file_path, name)):
            return name
    for name in search:
        if _sheet_has_enough_rows(file_path, name):
            return name
    shown = ", ".join(names[:6]) + ("…" if len(names) > 6 else "")
    raise ValueError(
        f"Нет листа с данными билетов (заголовок — строка 4). Листы: {shown}"
    )


def _is_likely_tickets_source(resolved: str) -> bool:
    base = os.path.basename(resolved).lower()
    if base.endswith("_bench.xlsx") or base == "bench.xlsx":
        return False
    if "gelendzhik" in base or "геленджик" in base:
        return False
    if "путь_сотрудника" in base or "путь-сотрудника" in base:
        return False
    return True


def _read_raw_excel(file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """Быстрое чтение: calamine + только колонки QWER, без лишнего открытия книги."""
    resolved = _resolve_input_path(file_path) or file_path
    if not os.path.isfile(resolved):
        raise ValueError(f"Файл не найден: {file_path}")

    if not _is_likely_tickets_source(resolved):
        raise ValueError(
            f"Похоже на отчёт, а не исходник билетов — пропуск ({os.path.basename(resolved)})"
        )

    sheet: Any = sheet_name if sheet_name is not None else _pick_tickets_sheet(resolved)

    usecols = list(RAW_COL_INDICES)
    last_err: Optional[Exception] = None
    for engine in ("calamine", "openpyxl"):
        try:
            df = pd.read_excel(
                resolved,
                sheet_name=sheet,
                header=HEADER_ROW_INDEX,
                usecols=usecols,
                engine=engine,
            )
            if df.shape[1] != len(usecols):
                raise ValueError(f"{engine}: unexpected column count {df.shape[1]}")
            df.columns = list(RAW_COL_NAMES)
            if df.empty:
                raise ValueError("Файл пустой или нет данных после строки заголовков")
            df = _overlay_identifier_columns_from_workbook(df, resolved, str(sheet))
            return _normalize_raw_operations_df(df)
        except Exception as e:
            last_err = e
            continue

    raise ValueError(
        f"Не удалось прочитать {os.path.basename(resolved)}: {last_err or 'unknown'}"
    )


def _loaded_source_paths(reg: str) -> set[str]:
    """Пути уже загруженных в реестр исходников — не читать xlsm повторно."""
    paths: set[str] = set()
    conn = _conn(reg)
    try:
        rows = conn.execute(
            "SELECT stored_path, original_name FROM source_files",
        ).fetchall()
    except Exception:
        return paths
    finally:
        conn.close()
    for stored_path, original_name in rows:
        for candidate in (stored_path, original_name):
            if not candidate:
                continue
            resolved = _resolve_input_path(str(candidate))
            if resolved:
                paths.add(os.path.normcase(os.path.abspath(resolved)))
    return paths


def _read_raw_import_df(conn: sqlite3.Connection) -> pd.DataFrame:
    """Чтение raw_import: только нужные колонки, не SELECT *."""
    try:
        info = conn.execute("PRAGMA table_info(raw_import)").fetchall()
    except Exception:
        return pd.DataFrame()
    existing = {row[1] for row in info}
    if "operaciya" in existing or "vid_uslugi" in existing:
        cols = [c for c in RAW_COL_NAMES if c in existing]
        if not cols:
            return pd.DataFrame()
        quoted = ", ".join(f'"{c}"' for c in cols)
        return pd.read_sql_query(f"SELECT {quoted} FROM raw_import", conn)
    return pd.read_sql_query("SELECT * FROM raw_import", conn)


def _extract_raw_row(row: pd.Series) -> Dict[str, Any]:
    """Собрать сырую строку в словарь по фиксированным полям."""
    cols = list(row.index)

    def by_idx(idx: int) -> Any:
        if idx < len(cols):
            return row.iloc[idx]
        return None

    out: Dict[str, Any] = {}
    for idx, key in RAW_COL_MAP.items():
        out[key] = by_idx(idx)
    out["operaciya"] = _normalize_operation_value(out.get("operaciya"))
    return out


def _raw_df_to_proc(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Позиционное или именованное (после быстрого чтения) преобразование сырых колонок."""
    n = len(raw_df)
    if n == 0:
        return pd.DataFrame(columns=list(RAW_COL_NAMES))

    if _raw_df_is_named(raw_df):
        proc = pd.DataFrame()
        for name in RAW_COL_NAMES:
            proc[name] = raw_df[name] if name in raw_df.columns else pd.Series([None] * n)
        proc["operaciya"] = proc["operaciya"].map(_normalize_operation_value)
        return proc

    proc = pd.DataFrame({key: raw_df.iloc[:, idx] for idx, key in RAW_COL_MAP.items() if idx < raw_df.shape[1]})
    for idx, key in RAW_COL_MAP.items():
        if key not in proc.columns:
            proc[key] = pd.Series([None] * n)
    proc["operaciya"] = proc["operaciya"].map(_normalize_operation_value)
    return proc


def _raw_to_processed_df(raw_df: pd.DataFrame) -> pd.DataFrame:
    proc = _raw_df_to_proc(raw_df)
    if proc.empty:
        return pd.DataFrame(columns=[c[0] for c in PROCESSED_COLUMNS])

    result = pd.DataFrame()
    result["nomer_bileta"] = proc["nomer_bileta"].map(_as_source_text)
    result["vid_uslugi"] = proc["vid_uslugi"].map(_norm_text)
    result["tabelyny_1c"] = ""
    result["fio_1c"] = ""
    result["ploshchadka"] = ""
    for src, dst in [
        ("podrazdelenie", "podrazdelenie"),
        ("obosnovanie_pereleta", "obosnovanie_pereleta"),
        ("organizaciya", "organizaciya"),
        ("klassifikaciya", "klassifikaciya"),
        ("operaciya", "operaciya"),
        ("fio", "fio"),
        ("tab_nomer", "tab_nomer"),
        ("pasport", "pasport"),
        ("marshrut", "marshrut"),
        ("napravlenie_gorod", "napravlenie_gorod"),
        ("data_vypiski", "data_vypiski"),
        ("data_vyleta_plan", "data_vyleta_plan"),
        ("data_prileta_plan", "data_prileta_plan"),
        ("aviaperevozchik", "aviaperevozchik"),
    ]:
        result[dst] = proc[src].map(_norm_text)
    result["pasport"] = proc["pasport"].map(_as_source_text)

    amounts = proc["summa_bilet_ag"].map(_parse_number).fillna(0.0)
    kinds = proc["operaciya"].map(_operation_kind)
    result["summa_pokupka"] = amounts.where(kinds == "pokupka", 0.0)
    result["summa_obmen"] = amounts.where(kinds == "obmen", 0.0)
    result["summa_vozvrat_sbor"] = amounts.where(kinds == "vozvrat", 0.0)
    result["summa_total"] = (
        result["summa_pokupka"] + result["summa_obmen"] + result["summa_vozvrat_sbor"]
    )
    dep_dates = proc["data_vyleta_plan"].map(_fmt_date_ddmmyyyy)
    dep_route = proc["marshrut"].map(_norm_text)
    marsh = pd.Series("", index=proc.index, dtype=object)
    both = dep_dates.ne("") & dep_route.ne("")
    only_d = dep_dates.ne("") & dep_route.eq("")
    only_m = dep_dates.eq("") & dep_route.ne("")
    marsh.loc[both] = dep_dates[both] + "; " + dep_route[both]
    marsh.loc[only_d] = dep_dates[only_d]
    marsh.loc[only_m] = dep_route[only_m]
    result["marshrut_data_vyleta"] = marsh

    for c in DATE_COLS:
        if c in result.columns:
            result[c] = _format_date_series(_parse_date_series(result[c]))

    result[YELLOW_FLAG] = False
    return result


def _merge_text_values(values: List[str]) -> Tuple[str, bool]:
    uniq = []
    seen = set()
    for v in values:
        t = _norm_text(v)
        if not t:
            continue
        if t not in seen:
            seen.add(t)
            uniq.append(t)
    if not uniq:
        return "", False
    if len(uniq) == 1:
        return uniq[0], False
    return "; ".join(uniq), True


def _dedupe_processed(df: pd.DataFrame) -> pd.DataFrame:
    """Слить строки с одинаковым номером билета (суммы складываются, тексты объединяются)."""
    if df.empty:
        return df
    work = df.copy()
    work["nomer_bileta"] = work["nomer_bileta"].map(_as_source_text)
    work["operaciya"] = work["operaciya"].map(_normalize_operation_value)
    work["_dedupe_key"] = work["nomer_bileta"].map(_norm_ticket_num)
    empty = work["_dedupe_key"] == ""
    if empty.any():
        work.loc[empty, "_dedupe_key"] = "__row__" + work.index[empty].astype(str)

    rows_out: List[Dict[str, Any]] = []
    for _key, grp in work.groupby("_dedupe_key", dropna=False):
        row: Dict[str, Any] = {}
        conflict = False
        for col, _title in PROCESSED_COLUMNS:
            if col not in grp.columns:
                row[col] = ""
                continue
            if col == OPERATION_COL:
                parts = [_norm_text(v) for v in grp[col].tolist() if _norm_text(v)]
                row[col] = " + ".join(dict.fromkeys(parts))
            elif col in SUM_COLS:
                row[col] = float(grp[col].map(_parse_number).sum())
            elif col == "summa_total":
                row[col] = float(
                    grp["summa_pokupka"].map(_parse_number).sum()
                    + grp["summa_obmen"].map(_parse_number).sum()
                    + grp["summa_vozvrat_sbor"].map(_parse_number).sum()
                )
            elif col == "nomer_bileta":
                row[col] = _as_source_text(grp[col].iloc[0])
            elif col in DATE_COLS:
                vals = grp[col].dropna()
                row[col] = _fmt_date_ddmmyyyy(vals.iloc[0]) if len(vals) else ""
            elif col in TEXT_MERGE_COLS:
                val, flag = _merge_text_values(grp[col].tolist())
                row[col] = val
                conflict = conflict or flag
            else:
                row[col] = grp[col].iloc[0] if col in grp else ""
        row[YELLOW_FLAG] = conflict
        rows_out.append(row)

    out = pd.DataFrame(rows_out)
    for col, _ in PROCESSED_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    return out[[c[0] for c in PROCESSED_COLUMNS] + [YELLOW_FLAG]]


def _cell_empty(val: Any) -> bool:
    return not _norm_text(val)


def _fill_1c_cells(out: pd.DataFrame, idx: Any, hit: Dict[str, Any]) -> None:
    if _cell_empty(out.at[idx, "tabelyny_1c"]):
        out.at[idx, "tabelyny_1c"] = _norm_text(hit.get("tab_num"))
    if _cell_empty(out.at[idx, "fio_1c"]):
        out.at[idx, "fio_1c"] = _norm_text(hit.get("fio"))


def _enrich_by_tab_nomer(df: pd.DataFrame, main_df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    """Заполнить пустые D/E по табельному номеру (L) из Основной Базы."""
    if df.empty:
        return df
    main_df = _ensure_main_employee_columns(main_df) if main_df is not None else _load_main_employees()
    if main_df.empty or "tab_norm" not in main_df.columns:
        return df

    subset = main_df.loc[main_df["tab_norm"].astype(str).str.strip() != ""]
    tab_to_tab = subset.drop_duplicates(subset=["tab_norm"]).set_index("tab_norm")["tab_num"].map(_norm_text).to_dict()
    tab_to_fio = subset.drop_duplicates(subset=["tab_norm"]).set_index("tab_norm")["fio"].map(_norm_text).to_dict()

    out = df.copy()
    need = out["tabelyny_1c"].map(_cell_empty) | out["fio_1c"].map(_cell_empty)
    if not need.any():
        return out
    need_idx = out.index[need]
    tabs = out.loc[need_idx, "tab_nomer"].map(_norm_text)
    new_tab = tabs.map(tab_to_tab)
    new_fio = tabs.map(tab_to_fio)
    fill_tab = out.loc[need_idx, "tabelyny_1c"].map(_cell_empty) & new_tab.notna() & new_tab.ne("")
    fill_fio = out.loc[need_idx, "fio_1c"].map(_cell_empty) & new_fio.notna() & new_fio.ne("")
    if fill_tab.any():
        idx = fill_tab.index[fill_tab.to_numpy()]
        out.loc[idx, "tabelyny_1c"] = new_tab.loc[idx].values
    if fill_fio.any():
        idx = fill_fio.index[fill_fio.to_numpy()]
        out.loc[idx, "fio_1c"] = new_fio.loc[idx].values
    return out


def _enrich_passport_only(
    df: pd.DataFrame,
    main_df: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, int]:
    filled = 0
    if df.empty:
        return df, filled
    main_df = _ensure_main_employee_columns(main_df) if main_df is not None else _load_main_employees()
    if main_df.empty or "passport_key" not in main_df.columns:
        return df, filled

    passport_rows = main_df.loc[main_df["passport_key"] != ""].drop_duplicates(subset=["passport_key"])
    main_map = passport_rows.set_index("passport_key")[["tab_num", "fio"]].to_dict(orient="index")

    out = df.copy()
    pass_keys = out["pasport"].map(integration_ops._normalize_passport)
    tab_empty = out["tabelyny_1c"].map(_cell_empty)
    fio_empty = out["fio_1c"].map(_cell_empty)
    need = tab_empty & fio_empty & pass_keys.astype(str).str.len().gt(0)
    if not need.any() or not main_map:
        return out, filled

    hits = pass_keys.map(main_map.get)
    need_idx = out.index[need]
    hit_tabs = hits.loc[need_idx].map(
        lambda h: _norm_text(h.get("tab_num")) if isinstance(h, dict) else "",
    )
    hit_fios = hits.loc[need_idx].map(
        lambda h: _norm_text(h.get("fio")) if isinstance(h, dict) else "",
    )
    before_empty = out.loc[need_idx, "tabelyny_1c"].map(_cell_empty) & out.loc[need_idx, "fio_1c"].map(_cell_empty)
    fill_tab = out.loc[need_idx, "tabelyny_1c"].map(_cell_empty) & hit_tabs.ne("")
    fill_fio = out.loc[need_idx, "fio_1c"].map(_cell_empty) & hit_fios.ne("")
    if fill_tab.any():
        idx = fill_tab.index[fill_tab.to_numpy()]
        out.loc[idx, "tabelyny_1c"] = hit_tabs.loc[idx].values
    if fill_fio.any():
        idx = fill_fio.index[fill_fio.to_numpy()]
        out.loc[idx, "fio_1c"] = hit_fios.loc[idx].values
    after_filled = (
        out.loc[need_idx, "tabelyny_1c"].map(_cell_empty).eq(False)
        | out.loc[need_idx, "fio_1c"].map(_cell_empty).eq(False)
    )
    filled = int((before_empty & after_filled).sum())
    return out, filled


def _enrich_exact_fio(
    df: pd.DataFrame,
    main_df: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, int]:
    filled = 0
    if df.empty:
        return df, filled
    main_df = main_df if main_df is not None else _load_main_employees()
    if main_df.empty:
        return df, filled

    fio_rows = main_df.loc[main_df["fio_norm"] != ""].drop_duplicates(subset=["fio_norm"])
    fio_map = fio_rows.set_index("fio_norm")[["tab_num", "fio"]].to_dict(orient="index")

    out = df.copy()
    need = out["tabelyny_1c"].map(_cell_empty) | out["fio_1c"].map(_cell_empty)
    if not need.any():
        return out, filled
    need_idx = out.index[need]
    norms = out.loc[need_idx, "fio"].map(integration_ops._normalize_fio)
    hits = norms.map(fio_map.get)
    matched = hits.notna()
    if not matched.any():
        return out, filled

    hit_tabs = hits.loc[matched].map(
        lambda h: _norm_text(h.get("tab_num")) if isinstance(h, dict) else "",
    )
    hit_fios = hits.loc[matched].map(
        lambda h: _norm_text(h.get("fio")) if isinstance(h, dict) else "",
    )
    m_idx = hit_tabs.index
    before = out.loc[m_idx, "tabelyny_1c"].map(_cell_empty) & out.loc[m_idx, "fio_1c"].map(_cell_empty)
    fill_tab = out.loc[m_idx, "tabelyny_1c"].map(_cell_empty) & hit_tabs.ne("")
    fill_fio = out.loc[m_idx, "fio_1c"].map(_cell_empty) & hit_fios.ne("")
    if fill_tab.any():
        idx = fill_tab.index[fill_tab.to_numpy()]
        out.loc[idx, "tabelyny_1c"] = hit_tabs.loc[idx].values
    if fill_fio.any():
        idx = fill_fio.index[fill_fio.to_numpy()]
        out.loc[idx, "fio_1c"] = hit_fios.loc[idx].values
    after = out.loc[m_idx, "tabelyny_1c"].map(_cell_empty).eq(False) | out.loc[m_idx, "fio_1c"].map(_cell_empty).eq(False)
    filled = int((before & after).sum())
    return out, filled


def _enrich_from_main_db(
    df: pd.DataFrame,
    progress: Optional[Callable[[str], None]] = None,
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """Табельный → паспорт → точное ФИО (с транслитом и *uulu)."""
    stats = {"filled_tab": 0, "filled_passport": 0, "filled_fio_exact": 0, "rows_need": 0}
    if df.empty:
        return df, stats
    main_df = _load_main_employees()
    if progress:
        progress("Табельный → База…")
    before_tab = df["tabelyny_1c"].map(_cell_empty) & df["fio_1c"].map(_cell_empty)
    out = _enrich_by_tab_nomer(df, main_df)
    after_tab = out["tabelyny_1c"].map(_cell_empty) & out["fio_1c"].map(_cell_empty)
    stats["filled_tab"] = int((before_tab & ~after_tab).sum())
    if progress:
        progress("Паспорт → База…")
    out, pass_filled = _enrich_passport_only(out, main_df)
    stats["filled_passport"] = pass_filled
    if progress:
        progress("Точное ФИО → База…")
    out, fio_filled = _enrich_exact_fio(out, main_df)
    stats["filled_fio_exact"] = fio_filled
    need = out["tabelyny_1c"].map(_cell_empty) | out["fio_1c"].map(_cell_empty)
    stats["rows_need"] = int(need.sum())
    return out, stats


def _build_fio_hit_lookup(
    unique_norms: List[str],
    main_fio_map: Dict[str, Dict[str, Any]],
    by_first_char: Dict[str, List[str]],
    by_prefix3: Dict[str, List[str]],
    by_surname: Dict[str, List[str]],
    cutoff: int,
    progress: Optional[Callable[[str], None]] = None,
) -> Dict[str, Dict[str, Any]]:
    lookup: Dict[str, Dict[str, Any]] = {}
    total = len(unique_norms)
    for i, nfio in enumerate(unique_norms):
        if not nfio or len(nfio) < 4:
            continue
        if progress and i > 0 and i % 250 == 0:
            progress(f"Fuzzy ФИО: {i:,}/{total:,} уникальных".replace(",", " "))
        if nfio in main_fio_map:
            lookup[nfio] = main_fio_map[nfio]
            continue
        hit, _score = integration_ops._fuzzy_match_fio(
            nfio,
            main_fio_map,
            by_first_char,
            score_cutoff=cutoff,
            by_prefix3=by_prefix3,
            by_surname=by_surname,
        )
        if hit:
            lookup[nfio] = hit
    return lookup


def _enrich_passport_and_fio_fuzzy(
    df: pd.DataFrame,
    fio_score_cutoff: int = 86,
    progress: Optional[Callable[[str], None]] = None,
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """
    Только пустые D/E: fuzzy по Ф.И.О. (K) с B Базы, с транслитом и *uulu.
    Fuzzy считается один раз на каждое уникальное ФИО (не на каждую строку).
    """
    stats = {"filled_passport": 0, "filled_fuzzy": 0, "rows_need": 0}
    if df.empty:
        return df, stats

    main_df = _load_main_employees()
    if main_df.empty:
        return df, stats

    main_fio_map, by_first_char, by_prefix3, by_surname = integration_ops._build_fio_indexes(
        main_df.rename(columns={"tab_num": "tab_num", "fio": "fio"}),
    )
    cutoff = max(50, min(100, int(fio_score_cutoff)))

    out = df.copy()
    tab_empty = out["tabelyny_1c"].map(_cell_empty)
    fio_empty = out["fio_1c"].map(_cell_empty)
    need = tab_empty & fio_empty
    if not need.any():
        return out, stats

    stats["rows_need"] = int(need.sum())
    still = need & out["fio"].map(lambda v: bool(_norm_text(v)))
    if not still.any():
        return out, stats

    norms = out.loc[still, "fio"].map(integration_ops._normalize_fio)
    unique_norms = [n for n in norms.unique().tolist() if n and len(n) >= 4]
    if progress:
        progress(
            f"Fuzzy по ФИО: {len(unique_norms):,} уникальных из {int(still.sum()):,} строк…".replace(",", " "),
        )
    lookup = _build_fio_hit_lookup(
        unique_norms,
        main_fio_map,
        by_first_char,
        by_prefix3,
        by_surname,
        cutoff,
        progress=progress,
    )
    if not lookup:
        return out, stats

    still_idx = out.index[still]
    hit_norms = norms.loc[still_idx].map(lookup.get)
    matched = hit_norms.notna()
    if not matched.any():
        return out, stats

    m_idx = hit_norms.index[matched]
    hit_tabs = hit_norms.loc[matched].map(
        lambda h: _norm_text(h.get("tab_num")) if isinstance(h, dict) else "",
    )
    hit_fios = hit_norms.loc[matched].map(
        lambda h: _norm_text(h.get("fio")) if isinstance(h, dict) else "",
    )
    before = out.loc[m_idx, "tabelyny_1c"].map(_cell_empty) & out.loc[m_idx, "fio_1c"].map(_cell_empty)
    fill_tab = out.loc[m_idx, "tabelyny_1c"].map(_cell_empty) & hit_tabs.ne("")
    fill_fio = out.loc[m_idx, "fio_1c"].map(_cell_empty) & hit_fios.ne("")
    if fill_tab.any():
        idx = fill_tab.index[fill_tab.to_numpy()]
        out.loc[idx, "tabelyny_1c"] = hit_tabs.loc[idx].values
    if fill_fio.any():
        idx = fill_fio.index[fill_fio.to_numpy()]
        out.loc[idx, "fio_1c"] = hit_fios.loc[idx].values
    after = out.loc[m_idx, "tabelyny_1c"].map(_cell_empty).eq(False) | out.loc[m_idx, "fio_1c"].map(_cell_empty).eq(False)
    stats["filled_fuzzy"] = int((before & after).sum())
    return out, stats


_TAB_JUNK_CHARS = "=.,-/\\"

_CYR_TO_LAT_PASSPORT = str.maketrans(
    {
        "а": "a",
        "А": "A",
        "е": "e",
        "Е": "E",
        "о": "o",
        "О": "O",
        "р": "p",
        "Р": "P",
        "с": "c",
        "С": "C",
        "х": "x",
        "Х": "X",
        "у": "y",
        "У": "Y",
        "В": "B",
        "К": "K",
        "М": "M",
        "Т": "T",
        "Н": "H",
    }
)


def _clean_tab_nomer_value(value: Any) -> str:
    """Оставить только формат ВМ-******* (цифры после префикса)."""
    s = _as_source_text(value)
    if not s:
        return ""
    for ch in _TAB_JUNK_CHARS:
        s = s.replace(ch, "")
    s = s.upper().replace("VM", "ВМ")
    m = re.search(r"ВМ\s*(\d+)", s)
    if m:
        return f"ВМ-{m.group(1)}"
    digits = re.sub(r"\D", "", s)
    if digits:
        return f"ВМ-{digits}"
    return s.strip()


def _clean_passport_value(value: Any) -> str:
    s = _as_source_text(value)
    if not s:
        return ""
    return s.translate(_CYR_TO_LAT_PASSPORT)


def _enrich_fio_en_translit(
    df: pd.DataFrame,
    main_df: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, int]:
    """Точное совпадение ФИО после транслитерации латиницы (столбец Ф.И.О.) → База."""
    filled = 0
    if df.empty:
        return df, filled
    main_df = _ensure_main_employee_columns(main_df) if main_df is not None else _load_main_employees()
    if main_df.empty or "fio_norm" not in main_df.columns:
        return df, filled

    fio_rows = main_df.loc[main_df["fio_norm"].astype(str).str.strip() != ""].drop_duplicates(subset=["fio_norm"])
    fio_map = fio_rows.set_index("fio_norm")[["tab_num", "fio"]].to_dict(orient="index")

    out = df.copy()
    tab_empty = out["tabelyny_1c"].map(_cell_empty)
    fio_empty = out["fio_1c"].map(_cell_empty)
    need = tab_empty & fio_empty
    if not need.any():
        return out, filled

    def _norm_en(v: Any) -> str:
        s = _as_source_text(v)
        if not s:
            return ""
        if integration_ops._has_latin_letters(s):
            s = integration_ops._translit_latin_to_russian(s)
        return integration_ops._normalize_fio(s)

    need_idx = out.index[need]
    norms = out.loc[need_idx, "fio"].map(_norm_en)
    hits = norms.map(fio_map.get)
    matched = hits.notna() & norms.astype(str).str.len().gt(0)
    if not matched.any():
        return out, filled

    m_idx = hits.index[matched]
    hit_tabs = hits.loc[matched].map(
        lambda h: _norm_text(h.get("tab_num")) if isinstance(h, dict) else "",
    )
    hit_fios = hits.loc[matched].map(
        lambda h: _norm_text(h.get("fio")) if isinstance(h, dict) else "",
    )
    before = out.loc[m_idx, "tabelyny_1c"].map(_cell_empty) & out.loc[m_idx, "fio_1c"].map(_cell_empty)
    fill_tab = out.loc[m_idx, "tabelyny_1c"].map(_cell_empty) & hit_tabs.ne("")
    fill_fio = out.loc[m_idx, "fio_1c"].map(_cell_empty) & hit_fios.ne("")
    if fill_tab.any():
        idx = fill_tab.index[fill_tab.to_numpy()]
        out.loc[idx, "tabelyny_1c"] = hit_tabs.loc[idx].values
    if fill_fio.any():
        idx = fill_fio.index[fill_fio.to_numpy()]
        out.loc[idx, "fio_1c"] = hit_fios.loc[idx].values
    after = out.loc[m_idx, "tabelyny_1c"].map(_cell_empty).eq(False) | out.loc[m_idx, "fio_1c"].map(_cell_empty).eq(False)
    filled = int((before & after).sum())
    return out, filled


def _norm_lookup_key(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    return re.sub(r"\s+", " ", s)


def _ensure_ploshchadka_column(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "ploshchadka" in df.columns:
        return df
    out = df.copy()
    out["ploshchadka"] = ""
    return out


def _load_processed_df(registry: str) -> pd.DataFrame:
    reg = tickets_db.normalize_registry(registry)
    conn = _conn(reg)
    try:
        df = pd.read_sql_query("SELECT * FROM processed", conn)
    except Exception:
        return pd.DataFrame()
    finally:
        conn.close()
    return _ensure_ploshchadka_column(df)


def _persist_processed_df(registry: str, df: pd.DataFrame, run_type: str) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    enriched = _ensure_row_ids(df, reg)
    conn = _conn(reg)
    try:
        _write_table(conn, "processed", enriched)
        conn.commit()
    finally:
        conn.close()
    meta = _load_meta(reg)
    meta["processed_rows"] = int(len(enriched))
    if YELLOW_FLAG in enriched.columns:
        meta["conflict_rows"] = int(enriched[YELLOW_FLAG].sum())
    meta["updated_at"] = datetime.now().isoformat()
    _save_meta(reg, meta)
    run_id = _save_processing_run(reg, run_type, enriched)
    return {
        "success": True,
        "registry": reg,
        "rows": int(len(enriched)),
        "run_id": run_id,
        "conflict_rows": meta.get("conflict_rows", 0),
    }


def fill_ploshchadka_from_reference(registry: str) -> Dict[str, Any]:
    import references

    mapping = references.get_podr_to_site()
    if not mapping:
        return {
            "error": "Справочник «Подр_Площадка_Затраты» не загружен. "
            "Настройки → Справочники → загрузите файл и нажмите «Применить справочники».",
        }
    df = _load_processed_df(registry)
    if df.empty:
        return {"error": "Нет обработанных данных — сначала «Обработать и отобразить»"}
    out = df.copy()
    filled = 0
    for idx in out.index:
        key = _norm_lookup_key(out.at[idx, "podrazdelenie"] if "podrazdelenie" in out.columns else "")
        if key and key in mapping:
            out.at[idx, "ploshchadka"] = mapping[key]
            filled += 1
    base = _persist_processed_df(registry, out, "fill_ploshchadka")
    base["filled_ploshchadka"] = filled
    return base


def clean_tab_passport_columns(registry: str) -> Dict[str, Any]:
    df = _load_processed_df(registry)
    if df.empty:
        return {"error": "Нет обработанных данных — сначала «Обработать и отобразить»"}
    out = df.copy()
    tab_changed = 0
    pas_changed = 0
    if "tab_nomer" in out.columns:
        new_tab = out["tab_nomer"].map(_clean_tab_nomer_value)
        old_tab = out["tab_nomer"].map(_as_source_text)
        tab_changed = int((new_tab != old_tab).sum())
        out["tab_nomer"] = new_tab
    if "pasport" in out.columns:
        new_pas = out["pasport"].map(_clean_passport_value)
        old_pas = out["pasport"].map(_as_source_text)
        pas_changed = int((new_pas != old_pas).sum())
        out["pasport"] = new_pas
    base = _persist_processed_df(registry, out, "clean_tab_passport")
    base["tab_changed"] = tab_changed
    base["passport_changed"] = pas_changed
    return base


def enrich_processed_by_passport(registry: str) -> Dict[str, Any]:
    df = _load_processed_df(registry)
    if df.empty:
        return {"error": "Нет обработанных данных"}
    out, filled = _enrich_passport_only(df)
    base = _persist_processed_df(registry, out, "enrich_passport")
    base["filled_passport"] = filled
    return base


def enrich_processed_by_fio_en(registry: str) -> Dict[str, Any]:
    df = _load_processed_df(registry)
    if df.empty:
        return {"error": "Нет обработанных данных"}
    out, filled = _enrich_fio_en_translit(df)
    base = _persist_processed_df(registry, out, "enrich_fio_en")
    base["filled_fio_en"] = filled
    return base


def enrich_processed_by_fio_fuzzy(registry: str, fuzzy_fio_cutoff: int = 90) -> Dict[str, Any]:
    df = _load_processed_df(registry)
    if df.empty:
        return {"error": "Нет обработанных данных"}
    out, stats = _enrich_passport_and_fio_fuzzy(df, fio_score_cutoff=fuzzy_fio_cutoff)
    base = _persist_processed_df(registry, out, "enrich_fio_fuzzy")
    base["filled_fuzzy"] = stats.get("filled_fuzzy", 0)
    base["rows_need"] = stats.get("rows_need", 0)
    base["fuzzy_fio_cutoff"] = max(50, min(100, int(fuzzy_fio_cutoff)))
    return base


def apply_processed_table_action(
    registry: str,
    action: str,
    fuzzy_fio_cutoff: int = 90,
) -> Dict[str, Any]:
    act = (action or "").strip().lower()
    if act in ("clean_tab_passport", "clean-tab-passport", "clean"):
        return clean_tab_passport_columns(registry)
    if act in ("enrich_passport", "enrich-passport", "passport"):
        return enrich_processed_by_passport(registry)
    if act in ("enrich_fio_en", "enrich-fio-en", "fio_en"):
        return enrich_processed_by_fio_en(registry)
    if act in ("enrich_fio_fuzzy", "enrich-fio-fuzzy", "fio_fuzzy", "fuzzy"):
        return enrich_processed_by_fio_fuzzy(registry, fuzzy_fio_cutoff)
    if act in ("fill_ploshchadka", "fill-ploshchadka", "fill_ploshchadki"):
        return fill_ploshchadka_from_reference(registry)
    return {"error": f"Неизвестное действие: {action}"}


def clear_all_source_files(registry: str) -> Dict[str, Any]:
    """Удалить только записи source_files и файлы на диске (без raw/processed)."""
    reg = tickets_db.normalize_registry(registry)
    listing = list_source_files(reg)
    files = listing.get("files") or []
    if not files:
        return {"success": True, "registry": reg, "deleted": 0, "message": "Нет загруженных исходных файлов"}
    deleted = 0
    errors: List[str] = []
    for item in files:
        fid = str(item.get("file_id") or "")
        if not fid:
            continue
        res = delete_source_file(reg, fid)
        if res.get("error"):
            errors.append(f"{item.get('original_name') or fid}: {res['error']}")
        else:
            deleted += 1
    out: Dict[str, Any] = {
        "success": not errors or deleted > 0,
        "registry": reg,
        "deleted": deleted,
        "total": len(files),
    }
    if errors:
        out["errors"] = errors
        out["message"] = f"Удалено {deleted} из {len(files)}; ошибки: {len(errors)}"
    else:
        out["message"] = f"Удалено исходных файлов: {deleted}"
    return out


def _apply_enrichment(
    df: pd.DataFrame,
    fuzzy_fio_cutoff: int = 86,
    progress: Optional[Callable[[str], None]] = None,
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """Табельный → паспорт → точное ФИО → fuzzy ФИО для оставшихся пустых D/E."""
    enriched, base_stats = _enrich_from_main_db(df, progress=progress)
    enriched, fuzzy_stats = _enrich_passport_and_fio_fuzzy(
        enriched,
        fio_score_cutoff=fuzzy_fio_cutoff,
        progress=progress,
    )
    stats: Dict[str, Any] = {**base_stats, **fuzzy_stats}
    return enriched, stats


def _write_table(conn: sqlite3.Connection, table: str, df: pd.DataFrame) -> None:
    _sanitize_df_for_sqlite(df).to_sql(table, conn, if_exists="replace", index=False)


def _distinct_values(conn: sqlite3.Connection, column: str, limit: int = 200) -> List[str]:
    try:
        rows = conn.execute(
            f"""
            SELECT DISTINCT TRIM(CAST({column} AS TEXT)) AS v
            FROM processed
            WHERE {column} IS NOT NULL AND TRIM(CAST({column} AS TEXT)) != ''
            ORDER BY v
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [str(r[0]) for r in rows if r[0]]
    except Exception:
        return []


def registry_filter_options(
    registry: str,
    conn: Optional[sqlite3.Connection] = None,
) -> Dict[str, Any]:
    """Списки для фильтров — DISTINCT в SQL, без чтения всей таблицы."""
    reg = tickets_db.normalize_registry(registry)
    own_conn = False
    if conn is None:
        conn = _conn(reg)
        own_conn = True
    try:
        years: set[int] = set()
        months: set[int] = set()
        try:
            date_rows = conn.execute(
                """
                SELECT DISTINCT data_vypiski FROM processed
                WHERE data_vypiski IS NOT NULL AND TRIM(CAST(data_vypiski AS TEXT)) != ''
                LIMIT 5000
                """,
            ).fetchall()
            for (raw,) in date_rows:
                dt = pd.to_datetime(raw, errors="coerce", dayfirst=True)
                if pd.notna(dt):
                    years.add(int(dt.year))
                    months.add(int(dt.month))
        except Exception:
            pass
        return {
            "years": sorted(years),
            "months": sorted(months),
            "ploshchadki": _distinct_values(conn, "ploshchadka", 200) or _distinct_values(conn, "podrazdelenie", 200),
            "podrazdeleniya": _distinct_values(conn, "ploshchadka", 200) or _distinct_values(conn, "podrazdelenie", 200),
            "obosnovaniya": _distinct_values(conn, "obosnovanie_pereleta", 200),
            "organizacii": _distinct_values(conn, "organizaciya", 100),
            "klassifikacii": _distinct_values(conn, "klassifikaciya", 100),
            "aviaperevozchiki": _distinct_values(conn, "aviaperevozchik", 100),
        }
    except Exception:
        return {
            "years": [], "months": [], "ploshchadki": [], "podrazdeleniya": [], "obosnovaniya": [],
            "organizacii": [], "klassifikacii": [], "aviaperevozchiki": [],
        }
    finally:
        if own_conn and conn is not None:
            conn.close()


def _status_for_registry(reg: str, *, light: bool = False) -> Dict[str, Any]:
    meta = _load_meta(reg)
    paths = _paths(reg)
    empty_filters = {
        "years": [],
        "months": [],
        "podrazdeleniya": [],
        "obosnovaniya": [],
        "organizacii": [],
        "klassifikacii": [],
        "aviaperevozchiki": [],
    }
    stored: List[Dict[str, Any]] = []
    runs: List[Dict[str, Any]] = []
    upload_q: List[Dict[str, Any]] = []
    filters = empty_filters
    conn = _conn(reg)
    try:
        _migrate_meta_queue_to_db(reg, conn)
        if not light:
            try:
                _auto_recover_upload_queue(reg, conn)
            except Exception:
                pass
        try:
            upload_q = _list_upload_queue(conn)
        except Exception:
            upload_q = []
        try:
            stored = [
                dict(r)
                for r in conn.execute(
                    """
                    SELECT file_id, original_name, stored_path, uploaded_at, row_count
                    FROM source_files
                    ORDER BY uploaded_at DESC
                    """,
                ).fetchall()
            ]
        except Exception:
            stored = []
        try:
            runs = [
                dict(r)
                for r in conn.execute(
                    """
                    SELECT run_id, run_type, label, created_at, row_count, active
                    FROM processing_runs
                    ORDER BY created_at DESC
                    """,
                ).fetchall()
            ]
        except Exception:
            runs = []
        if meta.get("processed_rows", 0) and not light:
            filters = registry_filter_options(reg, conn)
    finally:
        conn.close()
    return {
        "registry": reg,
        "label": REGISTRY_LABELS[reg],
        "db_path": paths["db"],
        "table_processed": "processed",
        "raw_files": meta.get("raw_files", []),
        "raw_rows": int(meta.get("raw_rows", 0)),
        "stored_files": stored,
        "upload_queue": upload_q,
        "processing_runs": runs,
        "processed_loaded": bool(meta.get("processed_rows", 0)),
        "processed_rows": meta.get("processed_rows", 0),
        "deduped": bool(meta.get("deduped", False)),
        "filters": filters,
    }


def get_status(registry: Optional[str] = None, *, light: bool = False) -> Dict[str, Any]:
    if registry:
        reg = tickets_db.normalize_registry(registry)
        return _status_for_registry(reg, light=light)
    out = {}
    for reg in (REGISTRY_VSM, REGISTRY_SK):
        out[reg] = _status_for_registry(reg, light=light)
    return {"registries": out}


def load_raw_files(
    file_paths: List[str],
    registry: str,
    sheet_name: Optional[str] = None,
    append: bool = False,
    progress: Optional[Callable[[int, int, str], None]] = None,
) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    if not file_paths:
        return {"error": "Не указаны файлы"}

    sources = _sources_dir(reg)
    total = len(file_paths)
    already_loaded = _loaded_source_paths(reg) if append else set()
    paths_to_load: List[str] = []
    skipped_loaded = 0
    for fp in file_paths:
        resolved = _resolve_input_path(fp)
        key = os.path.normcase(os.path.abspath(resolved)) if resolved else ""
        if append and key and key in already_loaded:
            skipped_loaded += 1
            if progress:
                progress(len(paths_to_load) + skipped_loaded, total, f"пропуск (уже в реестре): {os.path.basename(str(fp))}")
            continue
        paths_to_load.append(fp)

    if not paths_to_load and skipped_loaded:
        meta = _load_meta(reg)
        return {
            "success": True,
            "registry": reg,
            "label": REGISTRY_LABELS[reg],
            "files_loaded": 0,
            "files_skipped": skipped_loaded,
            "raw_rows": int(meta.get("raw_rows", 0)),
            "appended": append,
            "already_in_registry": True,
        }

    if not paths_to_load:
        return {"error": "Не указаны файлы для загрузки"}

    total = len(paths_to_load)
    frames: List[pd.DataFrame] = []
    loaded_names: List[str] = []
    skipped: List[str] = []
    failed: List[str] = []
    stored_entries: List[Tuple[str, str, str, int]] = []

    def _load_one(fp: str) -> Dict[str, Any]:
        resolved = _resolve_input_path(fp)
        if not resolved or not os.path.isfile(resolved):
            return {"kind": "skip", "path": fp}
        try:
            raw = _read_raw_excel(resolved, sheet_name)
            base = os.path.basename(resolved)
            file_id = f"{excel_handler.generate_file_id()}_{uuid.uuid4().hex[:6]}"
            dest = os.path.join(sources, f"{file_id}_{base}")
            shutil.copy2(resolved, dest)
            return {
                "kind": "ok",
                "raw": raw,
                "base": base,
                "entry": (file_id, base, dest, int(len(raw))),
            }
        except Exception as e:
            return {"kind": "fail", "message": f"{os.path.basename(str(fp))}: {e}"}

    from concurrent.futures import ThreadPoolExecutor, as_completed

    workers = min(8, max(1, total))
    done = 0
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [pool.submit(_load_one, fp) for fp in paths_to_load]
        for fut in as_completed(futures):
            result = fut.result()
            done += 1
            kind = result.get("kind")
            if kind == "ok":
                frames.append(result["raw"])
                loaded_names.append(result["base"])
                stored_entries.append(result["entry"])
                if progress:
                    progress(done, total, result["base"])
            elif kind == "skip":
                skipped.append(result["path"])
                if progress:
                    progress(done, total, os.path.basename(str(result["path"])))
            else:
                failed.append(result.get("message", "Ошибка загрузки файла"))
                if progress:
                    progress(done, total, result.get("message", "ошибка"))

    if not frames:
        msg = "Не удалось загрузить ни одного файла"
        if failed:
            msg += f". {failed[0]}"
            if len(failed) > 1:
                msg += f" (+ещё {len(failed) - 1})"
        if skipped:
            shown = ", ".join(os.path.basename(str(s)) for s in skipped[:5])
            extra = f" (+{len(skipped) - 5})" if len(skipped) > 5 else ""
            msg += f". Файлы не найдены на сервере: {shown}{extra}"
        return {"error": msg}

    combined = pd.concat(frames, ignore_index=True)
    combined = _normalize_raw_operations_df(combined)
    conn = _conn(reg)
    try:
        if append:
            try:
                existing = _read_raw_import_df(conn)
                if not existing.empty:
                    combined = pd.concat([existing, combined], ignore_index=True)
            except Exception:
                pass
        _sanitize_df_for_sqlite(combined).to_sql("raw_import", conn, if_exists="replace", index=False)
        now = datetime.now().isoformat()
        for file_id, base, dest, row_count in stored_entries:
            conn.execute(
                """
                INSERT OR REPLACE INTO source_files
                (file_id, original_name, stored_path, uploaded_at, row_count)
                VALUES (?, ?, ?, ?, ?)
                """,
                (file_id, base, dest, now, row_count),
            )
        conn.commit()
    finally:
        conn.close()

    meta = _load_meta(reg)
    prev_files = meta.get("raw_files", []) if append else []
    meta["raw_files"] = list(dict.fromkeys(prev_files + loaded_names))
    meta["raw_rows"] = int(len(combined))
    if not append:
        meta["processed_rows"] = 0
        meta["deduped"] = False
    meta["updated_at"] = datetime.now().isoformat()
    _prune_upload_queue_db(reg, file_paths, stored_entries)
    _save_meta(reg, meta)

    out: Dict[str, Any] = {
        "success": True,
        "registry": reg,
        "label": REGISTRY_LABELS[reg],
        "files_loaded": len(loaded_names),
        "files_skipped": len(skipped) + skipped_loaded,
        "raw_rows": int(len(combined)),
        "appended": append,
    }
    if failed:
        out["files_failed"] = len(failed)
        out["warnings"] = failed
    return out


def process_and_display(
    registry: str,
    fuzzy_fio_cutoff: int = 86,
    progress: Optional[Callable[[str, str], None]] = None,
) -> Dict[str, Any]:
    def _p(phase: str, detail: str = "") -> None:
        if progress:
            progress(phase, detail)

    reg = tickets_db.normalize_registry(registry)
    _p("read_raw", "Чтение raw_import из SQLite…")
    conn = _conn(reg)
    try:
        raw = _read_raw_import_df(conn)
    except Exception:
        return {"error": "Сначала загрузите файлы Excel"}
    finally:
        conn.close()

    if raw.empty:
        return {"error": "Нет сырых данных для обработки"}

    _p("transform", f"Преобразование и дедупликация {len(raw):,} строк…".replace(",", " "))
    transformed = _raw_to_processed_df(raw)
    deduped = _dedupe_processed(transformed)
    _p("enrich", f"Дополнение из Базы ({len(deduped):,} строк) — обычно самый долгий этап…".replace(",", " "))

    def _enrich_detail(detail: str) -> None:
        _p("enrich", detail)

    enriched, enrich_stats = _apply_enrichment(
        deduped,
        fuzzy_fio_cutoff=fuzzy_fio_cutoff,
        progress=_enrich_detail,
    )
    processed = _assign_row_ids(enriched, reg)

    _p("save", f"Сохранение {len(processed):,} строк…".replace(",", " "))
    conn = _conn(reg)
    try:
        _write_table(conn, "processed", processed)
        conn.commit()
    finally:
        conn.close()

    meta = _load_meta(reg)
    meta["processed_rows"] = int(len(processed))
    meta["deduped"] = True
    meta["conflict_rows"] = int(processed[YELLOW_FLAG].sum()) if YELLOW_FLAG in processed.columns else 0
    meta["updated_at"] = datetime.now().isoformat()
    _save_meta(reg, meta)

    run_id = _save_processing_run(reg, "process", processed)

    return {
        "success": True,
        "registry": reg,
        "processed_rows": int(len(processed)),
        "raw_rows": int(len(raw)),
        "rows_after_dedupe": int(len(deduped)),
        "enrich_stats": enrich_stats,
        "run_id": run_id,
    }


def dedupe_and_enrich(
    registry: str,
    fuzzy: bool = False,
    fuzzy_fio_cutoff: int = 86,
    run_dedupe: bool = True,
) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    conn = _conn(reg)
    try:
        df = pd.read_sql_query("SELECT * FROM processed", conn)
    except Exception:
        return {"error": "Сначала нажмите «Обработать и Отобразить»"}
    finally:
        conn.close()

    if df.empty:
        return {"error": "Нет обработанных данных"}

    work = _dedupe_processed(df) if run_dedupe else df.copy()
    enrich_stats: Dict[str, Any] = {}
    if fuzzy:
        enriched, fuzzy_only = _enrich_passport_and_fio_fuzzy(work, fio_score_cutoff=fuzzy_fio_cutoff)
        enrich_stats = fuzzy_only
    else:
        enriched, enrich_stats = _apply_enrichment(work, fuzzy_fio_cutoff=fuzzy_fio_cutoff)

    enriched = _assign_row_ids(enriched, reg)
    conn = _conn(reg)
    try:
        _write_table(conn, "processed", enriched)
        conn.commit()
    finally:
        conn.close()

    meta = _load_meta(reg)
    meta["processed_rows"] = int(len(enriched))
    if run_dedupe:
        meta["deduped"] = True
    meta["conflict_rows"] = int(enriched[YELLOW_FLAG].sum()) if YELLOW_FLAG in enriched.columns else 0
    meta["updated_at"] = datetime.now().isoformat()
    _save_meta(reg, meta)

    skip_export = fuzzy and not run_dedupe
    stored = ""
    file_id = None
    stats_write: Dict[str, Any] = {}
    if not skip_export:
        excel_handler.ensure_upload_dir()
        out_name = f"Затраты_{REGISTRY_LABELS[reg]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        stored = f"{excel_handler.generate_file_id()}_{out_name}"
        out_path = os.path.join(UPLOAD_DIR, stored)
        export_df = enriched.copy()
        if YELLOW_FLAG in export_df.columns:
            conflict_mask = export_df[YELLOW_FLAG].astype(bool)
            export_df = export_df.drop(columns=[YELLOW_FLAG])
        else:
            conflict_mask = pd.Series([False] * len(export_df))
        rename = {c[0]: c[1] for c in PROCESSED_COLUMNS}
        export_df = export_df.rename(columns=rename)
        if len(export_df) >= excel_libs.PYEXCELERATE_ROW_THRESHOLD:
            stats_write = write_dataframe(export_df, out_path, sheet_name="Данные")
        else:
            _export_excel_with_conflicts(export_df, out_path, conflict_mask)
            stats_write = {"engine": "openpyxl"}
        file_id = os.path.splitext(stored)[0]

    run_type = "enrich_fuzzy" if fuzzy and not run_dedupe else "dedupe"
    run_id = _save_processing_run(reg, run_type, enriched)

    return {
        "success": True,
        "registry": reg,
        "run_id": run_id,
        "rows": int(len(enriched)),
        "conflict_rows": meta.get("conflict_rows", 0),
        "file_id": file_id,
        "stored_filename": stored or None,
        "fuzzy_applied": fuzzy,
        "fuzzy_fio_cutoff": fuzzy_fio_cutoff if fuzzy else None,
        "dedupe_ran": run_dedupe,
        "write_engine": stats_write.get("engine"),
        "filled_passport": enrich_stats.get("filled_passport", 0),
        "filled_fuzzy": enrich_stats.get("filled_fuzzy", 0),
        "rows_need": enrich_stats.get("rows_need", 0),
    }


def _load_main_passport_map() -> Tuple[Dict[str, Dict[str, Any]], Any, Any]:
    main_df = _load_main_employees()
    if main_df.empty:
        return {}, {}, {}
    passport_rows = main_df.loc[main_df["passport_key"] != ""].drop_duplicates(subset=["passport_key"])
    main_map = passport_rows.set_index("passport_key")[["tab_num", "fio"]].to_dict(orient="index")
    by_length, by_suffix = integration_ops._build_passport_indexes(main_map)
    return main_map, by_length, by_suffix


def get_data(
    registry: str,
    search: Optional[str] = None,
    podrazdelenie: Optional[str] = None,
    ploshchadka: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    obosnovanie: Optional[str] = None,
    organizaciya: Optional[str] = None,
    klassifikaciya: Optional[str] = None,
    aviaperevozchik: Optional[str] = None,
    offset: int = 0,
    limit: int = 200,
) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    conn = _conn(reg)
    try:
        df = pd.read_sql_query("SELECT * FROM processed", conn)
    except Exception:
        return {
            "columns": DISPLAY_COLUMNS,
            "data": [],
            "total": 0,
            "offset": offset,
            "limit": limit,
            "registry": reg,
        }
    finally:
        conn.close()

    if df.empty:
        return {
            "columns": DISPLAY_COLUMNS,
            "data": [],
            "total": 0,
            "offset": offset,
            "limit": limit,
            "registry": reg,
        }

    df = _apply_dashboard_filters(
        df,
        year=year,
        month=month,
        podrazdelenie=podrazdelenie,
        ploshchadka=ploshchadka,
        obosnovanie=obosnovanie,
        organizaciya=organizaciya,
        klassifikaciya=klassifikaciya,
        aviaperevozchik=aviaperevozchik,
    )

    if search:
        mask = False
        for col in df.columns:
            mask = mask | df[col].astype(str).str.contains(search, case=False, na=False)
        df = df[mask]

    df = _ensure_row_ids(df, reg)
    drop_internal = [c for c in df.columns if c.startswith("_") and c != ROW_ID_COL]
    if drop_internal:
        df = df.drop(columns=drop_internal, errors="ignore")
    total = len(df)
    if limit and limit > 0:
        chunk = df.iloc[offset : offset + limit]
    else:
        chunk = df.iloc[offset:]
    records = chunk.replace({np.nan: None}).to_dict(orient="records")
    col_defs: List[Dict[str, str]] = []
    for key, title in PROCESSED_COLUMNS:
        fmt = "money" if key in SUM_COLS else ("id" if key in ("nomer_bileta", "pasport") else "text")
        col_defs.append({"key": key, "title": title, "format": fmt})
    return {
        "columns": col_defs,
        "data": records,
        "total": total,
        "offset": offset,
        "limit": limit,
        "registry": reg,
    }


def export_processed_to_excel(
    registry: str,
    search: Optional[str] = None,
    podrazdelenie: Optional[str] = None,
    ploshchadka: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    obosnovanie: Optional[str] = None,
    organizaciya: Optional[str] = None,
    klassifikaciya: Optional[str] = None,
    aviaperevozchik: Optional[str] = None,
) -> Dict[str, Any]:
    """Выгрузить все строки processed в .xlsx (без лимита браузера)."""
    reg = tickets_db.normalize_registry(registry)
    conn = _conn(reg)
    try:
        df = pd.read_sql_query("SELECT * FROM processed", conn)
    finally:
        conn.close()

    if df.empty:
        return {"error": "Нет обработанных данных"}

    df = _apply_dashboard_filters(
        df,
        year=year,
        month=month,
        podrazdelenie=podrazdelenie,
        ploshchadka=ploshchadka,
        obosnovanie=obosnovanie,
        organizaciya=organizaciya,
        klassifikaciya=klassifikaciya,
        aviaperevozchik=aviaperevozchik,
    )

    if search:
        mask = False
        for col in df.columns:
            mask = mask | df[col].astype(str).str.contains(search, case=False, na=False)
        df = df[mask]

    if df.empty:
        return {"error": "Нет строк по выбранным фильтрам"}

    drop_cols = [c for c in df.columns if c.startswith("_") or c == YELLOW_FLAG]
    if drop_cols:
        df = df.drop(columns=drop_cols, errors="ignore")

    keys = [c[0] for c in PROCESSED_COLUMNS if c[0] in df.columns]
    export_df = df[keys].copy()
    export_df = export_df.rename(columns={c[0]: c[1] for c in PROCESSED_COLUMNS})

    excel_handler.ensure_upload_dir()
    out_name = f"Затраты_{REGISTRY_LABELS[reg]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    stored = f"{excel_handler.generate_file_id()}_{out_name}"
    out_path = os.path.join(UPLOAD_DIR, stored)
    stats_write = write_dataframe(export_df, out_path, sheet_name="Данные")
    file_id = os.path.splitext(stored)[0]

    return {
        "success": True,
        "registry": reg,
        "file_id": file_id,
        "stored_filename": stored,
        "row_count": int(len(export_df)),
        "write_engine": stats_write.get("engine"),
    }


def update_processed_rows(registry: str, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Обновить строки обработанного реестра по _row_id."""
    reg = tickets_db.normalize_registry(registry)
    if not rows:
        return {"error": "Нет данных для сохранения"}

    conn = _conn(reg)
    try:
        df = pd.read_sql_query("SELECT * FROM processed", conn)
    except Exception:
        return {"error": "Нет обработанных данных — сначала «Обработать и Отобразить»"}
    finally:
        conn.close()

    if df.empty:
        return {"error": "Нет обработанных данных"}

    df = _ensure_row_ids(df, reg)
    id_to_idx = {str(rid): i for i, rid in enumerate(df["_row_id"].astype(str))}
    updated = 0

    for patch in rows:
        rid = str(patch.get("_row_id", ""))
        if not rid or rid not in id_to_idx:
            continue
        idx = id_to_idx[rid]
        for key in EDITABLE_KEYS:
            if key in patch:
                val = patch[key]
                if key in SUM_COLS:
                    df.at[idx, key] = _parse_number(val)
                elif key in DATE_COLS:
                    df.at[idx, key] = _fmt_date_ddmmyyyy(val)
                else:
                    df.at[idx, key] = _norm_text(val)
        if any(k in patch for k in ("summa_pokupka", "summa_obmen", "summa_vozvrat_sbor")):
            df.at[idx, "summa_total"] = (
                _parse_number(df.at[idx, "summa_pokupka"])
                + _parse_number(df.at[idx, "summa_obmen"])
                + _parse_number(df.at[idx, "summa_vozvrat_sbor"])
            )
        updated += 1

    conn = _conn(reg)
    try:
        _write_table(conn, "processed", df)
        conn.commit()
    finally:
        conn.close()

    meta = _load_meta(reg)
    meta["updated_at"] = datetime.now().isoformat()
    _save_meta(reg, meta)

    return {"success": True, "updated": updated, "registry": reg}


def _ensure_site_column(df: pd.DataFrame) -> pd.DataFrame:
    """Колонка _site: значение «Площадка», при пустом — подразделение (старые строки)."""
    if df.empty:
        return df
    work = _ensure_ploshchadka_column(df.copy())
    pl = (
        work["ploshchadka"].astype(str).str.strip()
        if "ploshchadka" in work.columns
        else pd.Series([""] * len(work), index=work.index)
    )
    pod = (
        work["podrazdelenie"].astype(str).str.strip()
        if "podrazdelenie" in work.columns
        else pd.Series([""] * len(work), index=work.index)
    )
    site = pl.where(pl.ne("") & ~pl.str.lower().isin(("nan", "none")), pod)
    work["_site"] = site.replace({"": "—", "nan": "—", "None": "—"})
    return work


def _apply_dashboard_filters(
    df: pd.DataFrame,
    year: Optional[int] = None,
    month: Optional[int] = None,
    podrazdelenie: Optional[str] = None,
    ploshchadka: Optional[str] = None,
    obosnovanie: Optional[str] = None,
    organizaciya: Optional[str] = None,
    klassifikaciya: Optional[str] = None,
    aviaperevozchik: Optional[str] = None,
) -> pd.DataFrame:
    if df.empty:
        return df
    site_filter = (ploshchadka or podrazdelenie or "").strip()
    if site_filter:
        df = _ensure_site_column(df)
        df = df[df["_site"].astype(str).str.contains(site_filter, case=False, na=False)]
    if obosnovanie:
        df = df[df["obosnovanie_pereleta"].astype(str).str.contains(obosnovanie, case=False, na=False)]
    if organizaciya:
        df = df[df["organizaciya"].astype(str).str.contains(organizaciya, case=False, na=False)]
    if klassifikaciya:
        df = df[df["klassifikaciya"].astype(str).str.contains(klassifikaciya, case=False, na=False)]
    if aviaperevozchik:
        df = df[df["aviaperevozchik"].astype(str).str.contains(aviaperevozchik, case=False, na=False)]
    dates = pd.to_datetime(df["data_vypiski"], errors="coerce", dayfirst=True)
    df = df.assign(_year=dates.dt.year, _month=dates.dt.month)
    if year is not None:
        df = df[df["_year"] == year]
    if month is not None:
        df = df[df["_month"] == month]
    return df


def _sum_column(df: pd.DataFrame) -> pd.Series:
    if "summa_total" in df.columns:
        return df["summa_total"].map(_parse_number)
    if "summa" in df.columns:
        return df["summa"].map(_parse_number)
    return pd.Series([0.0] * len(df))


_DASHBOARD_DB_COLS = (
    "data_vypiski",
    "nomer_bileta",
    "ploshchadka",
    "podrazdelenie",
    "obosnovanie_pereleta",
    "organizaciya",
    "klassifikaciya",
    "aviaperevozchik",
    "marshrut",
    "fio",
    "summa_total",
    "summa",
    "summa_pokupka",
    "summa_obmen",
    "summa_vozvrat_sbor",
    "operaciya",
)


def _read_processed_for_dashboard(conn) -> pd.DataFrame:
    """Только колонки дашборда — быстрее, чем SELECT *."""
    try:
        info = conn.execute("PRAGMA table_info(processed)").fetchall()
    except Exception:
        return pd.DataFrame()
    existing = {row[1] for row in info}
    cols = [c for c in _DASHBOARD_DB_COLS if c in existing]
    if not cols:
        return pd.DataFrame()
    sql = f"SELECT {', '.join(cols)} FROM processed"
    try:
        return pd.read_sql_query(sql, conn)
    except Exception:
        return pd.DataFrame()


def _filtered_df(
    registry: str,
    year: Optional[int] = None,
    month: Optional[int] = None,
    podrazdelenie: Optional[str] = None,
    ploshchadka: Optional[str] = None,
    obosnovanie: Optional[str] = None,
    organizaciya: Optional[str] = None,
    klassifikaciya: Optional[str] = None,
    aviaperevozchik: Optional[str] = None,
) -> pd.DataFrame:
    reg = tickets_db.normalize_registry(registry)
    conn = _conn(reg)
    try:
        df = _read_processed_for_dashboard(conn)
    except Exception:
        return pd.DataFrame()
    finally:
        conn.close()
    return _apply_dashboard_filters(
        df,
        year,
        month,
        podrazdelenie=podrazdelenie,
        ploshchadka=ploshchadka,
        obosnovanie=obosnovanie,
        organizaciya=organizaciya,
        klassifikaciya=klassifikaciya,
        aviaperevozchik=aviaperevozchik,
    )


def _dashboard_json_records(df: pd.DataFrame) -> List[Dict[str, Any]]:
    if df.empty:
        return []
    clean = df.replace({np.nan: None, np.inf: None, -np.inf: None})
    for col in clean.select_dtypes(include=[np.number]).columns:
        clean[col] = clean[col].where(pd.notna(clean[col]), None)
    return clean.to_dict(orient="records")


def dashboard_stats(
    registry: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    podrazdelenie: Optional[str] = None,
    ploshchadka: Optional[str] = None,
    obosnovanie: Optional[str] = None,
    organizaciya: Optional[str] = None,
    klassifikaciya: Optional[str] = None,
    aviaperevozchik: Optional[str] = None,
) -> Dict[str, Any]:
    """Агрегаты для дашборда (все реестры или один)."""
    from concurrent.futures import ThreadPoolExecutor

    site_filter = ploshchadka or podrazdelenie
    registries = [tickets_db.normalize_registry(registry)] if registry else [REGISTRY_VSM, REGISTRY_SK]
    filter_args = (
        year,
        month,
        None,
        site_filter,
        obosnovanie,
        organizaciya,
        klassifikaciya,
        aviaperevozchik,
    )

    def _load_reg(reg: str) -> pd.DataFrame:
        return _filtered_df(reg, *filter_args)

    frames: list[pd.DataFrame] = []
    if len(registries) == 1:
        dfs = [_load_reg(registries[0])]
    else:
        with ThreadPoolExecutor(max_workers=2) as pool:
            dfs = list(pool.map(_load_reg, registries))

    for reg, df in zip(registries, dfs):
        if not df.empty:
            df = df.copy()
            df["registry"] = reg
            df["registry_label"] = REGISTRY_LABELS[reg]
            frames.append(df)
    if not frames:
        return _empty_dashboard()

    all_df = pd.concat(frames, ignore_index=True)
    all_df = _ensure_site_column(all_df)
    all_df["_sum"] = _sum_column(all_df)
    dates = pd.to_datetime(all_df["data_vypiski"], errors="coerce", dayfirst=True)
    all_df["_year"] = dates.dt.year
    all_df["_month"] = dates.dt.month

    total_sum = float(all_df["_sum"].sum())
    ticket_count = int(
        all_df["nomer_bileta"].astype(str).str.strip().replace("", np.nan).dropna().nunique()
    )
    site_series = all_df["_site"].astype(str).str.strip().replace("—", np.nan)
    sites = int(site_series.dropna().nunique())

    by_site = (
        all_df.groupby("_site", dropna=False)
        .agg(summa=("_sum", "sum"), tickets=("nomer_bileta", "count"))
        .reset_index()
        .rename(columns={"_site": "ploshchadka"})
        .sort_values("summa", ascending=False)
    )
    by_site["avg_check"] = by_site.apply(
        lambda r: float(r["summa"]) / r["tickets"] if r["tickets"] else 0.0, axis=1
    )
    top_site = by_site.head(10)
    top_site["ploshchadka"] = top_site["ploshchadka"].fillna("—").replace("", "—")

    by_time = (
        all_df.dropna(subset=["_year"])
        .groupby(["_year", "_month"], dropna=False)
        .agg(summa=("_sum", "sum"), tickets=("nomer_bileta", "count"))
        .reset_index()
        .sort_values(["_year", "_month"])
    )

    if "summa_pokupka" in all_df.columns:
        operations = {
            "total": int(len(all_df)),
            "pokupka": int((all_df["summa_pokupka"].map(_parse_number) > 0).sum()),
            "obmen": int((all_df["summa_obmen"].map(_parse_number) > 0).sum()),
            "vozvrat": int((all_df["summa_vozvrat_sbor"].map(_parse_number) > 0).sum()),
        }
    else:
        ops = all_df["operaciya"].astype(str).str.strip().str.lower()
        operations = {
            "total": int(len(all_df)),
            "pokupka": int(ops.str.contains(r"продаж|покуп", na=False, regex=True).sum()),
            "obmen": int(ops.str.contains("обмен", na=False).sum()),
            "vozvrat": int(ops.str.contains("возврат", na=False).sum()),
        }

    by_obosn = (
        all_df.groupby("obosnovanie_pereleta", dropna=False)
        .agg(tickets=("nomer_bileta", "count"), summa=("_sum", "sum"))
        .reset_index()
        .sort_values("summa", ascending=False)
        .head(15)
    )

    by_carrier = (
        all_df.groupby("aviaperevozchik", dropna=False)
        .agg(summa=("_sum", "sum"), tickets=("nomer_bileta", "count"))
        .reset_index()
        .sort_values("summa", ascending=False)
    )

    top_routes = (
        all_df.groupby("marshrut", dropna=False)
        .agg(
            summa=("_sum", "sum"),
            trips=("nomer_bileta", "count"),
            avg_cost=("_sum", "mean"),
        )
        .reset_index()
        .sort_values("summa", ascending=False)
        .head(5)
    )

    top_employees = (
        all_df.groupby("fio", dropna=False)
        .agg(trips=("nomer_bileta", "count"), summa=("_sum", "sum"))
        .reset_index()
        .sort_values("summa", ascending=False)
        .head(10)
    )
    top_employees["avg_cost"] = top_employees.apply(
        lambda r: float(r["summa"]) / r["trips"] if r["trips"] else 0.0, axis=1
    )

    ploshchadki_vals = sorted(
        all_df["_site"].astype(str).str.strip().replace("—", np.nan).dropna().unique().tolist()
    )
    filters = {
        "years": sorted([int(y) for y in all_df["_year"].dropna().unique()]),
        "months": sorted([int(m) for m in all_df["_month"].dropna().unique()]),
        "ploshchadki": ploshchadki_vals,
        "podrazdeleniya": ploshchadki_vals,
        "obosnovaniya": sorted(
            all_df["obosnovanie_pereleta"].astype(str).str.strip().replace("", np.nan).dropna().unique().tolist()
        )[:200],
        "organizacii": sorted(
            all_df["organizaciya"].astype(str).str.strip().replace("", np.nan).dropna().unique().tolist()
        )[:100],
        "klassifikacii": sorted(
            all_df["klassifikaciya"].astype(str).str.strip().replace("", np.nan).dropna().unique().tolist()
        )[:100],
        "aviaperevozchiki": sorted(
            all_df["aviaperevozchik"].astype(str).str.strip().replace("", np.nan).dropna().unique().tolist()
        )[:100],
    }

    return {
        "kpi": {
            "total_sum": total_sum,
            "ticket_count": ticket_count,
            "site_count": sites,
        },
        "by_ploshchadka": _dashboard_json_records(top_site),
        "by_podrazdelenie": _dashboard_json_records(
            top_site.rename(columns={"ploshchadka": "podrazdelenie"})
        ),
        "by_time": _dashboard_json_records(by_time),
        "operations": operations,
        "by_obosnovanie": _dashboard_json_records(by_obosn),
        "by_aviaperevozchik": _dashboard_json_records(by_carrier),
        "top_marshruty": _dashboard_json_records(top_routes),
        "top_employees": _dashboard_json_records(top_employees),
        "filters": filters,
    }


def _empty_dashboard() -> Dict[str, Any]:
    return {
        "kpi": {"total_sum": 0, "ticket_count": 0, "site_count": 0},
        "by_ploshchadka": [],
        "by_podrazdelenie": [],
        "by_time": [],
        "operations": {"total": 0, "pokupka": 0, "obmen": 0, "vozvrat": 0},
        "by_obosnovanie": [],
        "by_aviaperevozchik": [],
        "top_marshruty": [],
        "top_employees": [],
        "filters": {
            "years": [], "months": [], "ploshchadki": [], "podrazdeleniya": [], "obosnovaniya": [],
            "organizacii": [], "klassifikacii": [], "aviaperevozchiki": [],
        },
    }


def _export_excel_with_conflicts(
    df: pd.DataFrame,
    out_path: str,
    conflict_rows: pd.Series,
) -> None:
    """Excel с жёлтой заливкой ячеек, где при дедупликации значения различались (сцепка через «;»)."""
    df.to_excel(out_path, index=False, sheet_name="Данные", engine="openpyxl")
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(out_path)
        ws = wb["Данные"]
        yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        money_cols = {c[1] for c in PROCESSED_COLUMNS if c[0] in SUM_COLS}
        text_cols = {c[1] for c in PROCESSED_COLUMNS if c[0] in TEXT_MERGE_COLS}
        for c_idx, header in enumerate(ws[1], start=1):
            if header.value in money_cols:
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=c_idx).number_format = '#,##0.00\\ "₽"'
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            is_conflict = bool(conflict_rows.iloc[r_idx - 2]) if r_idx - 2 < len(conflict_rows) else False
            for c_idx, val in enumerate(row, start=1):
                s = _norm_text(val)
                if is_conflict and ";" in s:
                    ws.cell(row=r_idx, column=c_idx).fill = yellow
                elif ws.cell(row=1, column=c_idx).value in text_cols and ";" in s:
                    ws.cell(row=r_idx, column=c_idx).fill = yellow
        wb.save(out_path)
    except Exception:
        pass


def clear_registry(registry: str) -> Dict[str, Any]:
    reg = tickets_db.normalize_registry(registry)
    paths = _paths(reg)
    db_path = paths["db"]
    lock = _db_lock(reg)

    with _registry_db_locks_guard:
        _registry_clearing.add(reg)
    try:
        _force_close_registry_connections(reg)
        if not lock.acquire(timeout=120):
            return {
                "error": "Не удалось заблокировать реестр за 120 с. Перезапустите START.bat и повторите.",
            }

        try:
            wipe_note: Optional[str] = None
            failed: List[str] = []
            paths_to_remove = (
                db_path,
                f"{db_path}-wal",
                f"{db_path}-shm",
                paths["meta"],
            )
            reg_dir = os.path.join(UPLOAD_DIR, f"tickets_costs_{reg}")

            for attempt in range(30):
                _force_close_registry_connections(reg)
                if os.path.isfile(db_path):
                    try:
                        _wipe_registry_db(db_path)
                        wipe_note = None
                    except Exception as exc:
                        wipe_note = str(exc)

                failed = []
                for path in paths_to_remove:
                    if os.path.exists(path) and not _remove_path(path, retries=8):
                        failed.append(os.path.basename(path))
                if os.path.isdir(reg_dir) and not _remove_path(reg_dir, retries=8):
                    failed.append(os.path.basename(reg_dir))

                if not failed:
                    break
                time.sleep(min(2.0, 0.25 * (attempt + 1)))

            try:
                tickets_db.clear_cache(reg)
            except Exception:
                pass

            if failed:
                parts: List[str] = []
                if wipe_note:
                    parts.append(f"SQLite: {wipe_note}")
                parts.append(f"не удалено на диске: {', '.join(failed)}")
                parts.append("Перезапустите START.bat и повторите очистку.")
                return {"error": ". ".join(parts)}

            return {
                "success": True,
                "registry": reg,
                "label": REGISTRY_LABELS[reg],
                "message": "Удалены сырые данные, processed, снимки и исходные файлы реестра",
            }
        finally:
            lock.release()
    finally:
        with _registry_db_locks_guard:
            _registry_clearing.discard(reg)
