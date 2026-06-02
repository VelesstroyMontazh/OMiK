"""
Macro Engine - VBA-like macro parsing and execution.
Supports a safe subset of VBA operations and Python macros.
"""

import re
import os
import traceback
from typing import Dict, Any, List, Optional

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


# =============================================================================
# VBA Parser - Parses a safe subset of VBA syntax
# =============================================================================

class VBAParser:
    """
    Parses a safe subset of VBA-like syntax.
    Supports: Sub/End Sub, Dim, For/Next, If/Then/Else/End If,
              Set, Range(), Cells(), Worksheets(), ActiveSheet,
              Value, Formula, and basic operations.
    """

    def __init__(self, code: str):
        self.code = code
        self.lines = code.strip().split('\n')
        self.variables: Dict[str, Any] = {}
        self.output: List[str] = []
        self.errors: List[str] = []

    def parse_and_execute(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """Parse and execute VBA-like code."""
        try:
            # Find Sub blocks
            sub_blocks = self._extract_sub_blocks()

            for sub_name, sub_lines in sub_blocks.items():
                self._execute_block(sub_lines, workbook)

            return {
                "success": True,
                "output": self.output,
                "errors": self.errors,
                "variables": {k: str(v) for k, v in self.variables.items()},
            }
        except Exception as e:
            return {
                "success": False,
                "output": self.output,
                "errors": [str(e), traceback.format_exc()],
                "variables": {k: str(v) for k, v in self.variables.items()},
            }

    def _extract_sub_blocks(self) -> Dict[str, List[str]]:
        """Extract Sub procedure blocks from VBA code."""
        blocks = {}
        current_sub = None
        current_lines = []

        for line in self.lines:
            stripped = line.strip()

            # Match Sub declaration
            sub_match = re.match(r'^\s*Sub\s+(\w+)\s*\(\s*\)\s*$', stripped, re.IGNORECASE)
            if sub_match:
                current_sub = sub_match.group(1)
                current_lines = []
                continue

            # Match End Sub
            if re.match(r'^\s*End\s+Sub\s*$', stripped, re.IGNORECASE):
                if current_sub:
                    blocks[current_sub] = current_lines
                    current_sub = None
                    current_lines = []
                continue

            # Collect lines inside Sub
            if current_sub is not None:
                current_lines.append(stripped)

        # If no Sub blocks found, treat entire code as a block
        if not blocks and self.lines:
            blocks["Main"] = [l.strip() for l in self.lines if l.strip()]

        return blocks

    def _execute_block(self, lines: List[str], workbook: openpyxl.Workbook):
        """Execute a block of VBA-like lines."""
        i = 0
        while i < len(lines):
            line = lines[i]

            # Skip empty lines and comments
            if not line or line.startswith("'") or line.startswith("Rem"):
                i += 1
                continue

            # Dim statement
            dim_match = re.match(r'^\s*Dim\s+(\w+)\s+(As\s+\w+)?\s*$', line, re.IGNORECASE)
            if dim_match:
                var_name = dim_match.group(1)
                self.variables[var_name] = None
                i += 1
                continue

            # For loop
            for_match = re.match(
                r'^\s*For\s+(\w+)\s*=\s*(.+?)\s+To\s+(.+?)(\s+Step\s+(.+))?\s*$',
                line, re.IGNORECASE
            )
            if for_match:
                var_name = for_match.group(1)
                start_val = self._eval_expr(for_match.group(2))
                end_val = self._eval_expr(for_match.group(3))
                step_val = int(self._eval_expr(for_match.group(5))) if for_match.group(5) else 1

                # Collect body until Next
                body_lines = []
                depth = 1
                j = i + 1
                while j < len(lines) and depth > 0:
                    inner = lines[j].strip()
                    if re.match(r'^\s*For\s+', inner, re.IGNORECASE):
                        depth += 1
                    elif re.match(r'^\s*Next\s*', inner, re.IGNORECASE):
                        depth -= 1
                        if depth == 0:
                            break
                    body_lines.append(inner)
                    j += 1

                # Execute loop
                self.variables[var_name] = int(start_val)
                current = int(start_val)
                while (step_val > 0 and current <= int(end_val)) or \
                      (step_val < 0 and current >= int(end_val)):
                    self.variables[var_name] = current
                    self._execute_block(body_lines, workbook)
                    current += step_val

                i = j + 1
                continue

            # If/Then/Else
            if_match = re.match(r'^\s*If\s+(.+?)\s+Then\s*$', line, re.IGNORECASE)
            if if_match:
                condition = if_match.group(1)
                then_lines = []
                else_lines = []
                in_else = False
                depth = 1
                j = i + 1
                while j < len(lines) and depth > 0:
                    inner = lines[j].strip()
                    if re.match(r'^\s*If\s+.+?\s+Then\s*$', inner, re.IGNORECASE):
                        depth += 1
                    elif re.match(r'^\s*End\s+If\s*$', inner, re.IGNORECASE):
                        depth -= 1
                        if depth == 0:
                            break
                    elif re.match(r'^\s*Else\s*$', inner, re.IGNORECASE) and depth == 1:
                        in_else = True
                        j += 1
                        continue

                    if in_else:
                        else_lines.append(inner)
                    else:
                        then_lines.append(inner)
                    j += 1

                if self._eval_condition(condition):
                    self._execute_block(then_lines, workbook)
                else:
                    self._execute_block(else_lines, workbook)

                i = j + 1
                continue

            # Single-line If
            if_then_match = re.match(r'^\s*If\s+(.+?)\s+Then\s+(.+)$', line, re.IGNORECASE)
            if if_then_match:
                condition = if_then_match.group(1)
                action = if_then_match.group(2)
                if self._eval_condition(condition):
                    self._execute_line(action, workbook)
                i += 1
                continue

            # Execute regular line
            self._execute_line(line, workbook)
            i += 1

    def _execute_line(self, line: str, workbook: openpyxl.Workbook):
        """Execute a single VBA-like line."""
        # Range().Value = something
        range_set = re.match(
            r'^\s*Range\(\s*"([^"]+)"\s*\)\s*\.?\s*(Value|Formula|Value2)?\s*=\s*(.+)$',
            line, re.IGNORECASE
        )
        if range_set:
            cell_ref = range_set.group(1)
            value = self._eval_expr(range_set.group(3))
            self._set_range_value(workbook, cell_ref, value)
            return

        # Cells(r, c).Value = something
        cells_set = re.match(
            r'^\s*Cells\(\s*(.+?)\s*,\s*(.+?)\s*\)\s*\.?\s*(Value|Formula|Value2)?\s*=\s*(.+)$',
            line, re.IGNORECASE
        )
        if cells_set:
            row = int(self._eval_expr(cells_set.group(1)))
            col = int(self._eval_expr(cells_set.group(2)))
            value = self._eval_expr(cells_set.group(4))
            self._set_cells_value(workbook, row, col, value)
            return

        # Variable assignment
        var_assign = re.match(r'^\s*(\w+)\s*=\s*(.+)$', line, re.IGNORECASE)
        if var_assign:
            var_name = var_assign.group(1)
            if var_name.lower() not in ['sub', 'end', 'if', 'then', 'else', 'for', 'to', 'step', 'next', 'dim']:
                self.variables[var_name] = self._eval_expr(var_assign.group(2))
                return

        # Debug.Print or Print
        print_match = re.match(r'^\s*(Debug\.Print|Print)\s+(.+)$', line, re.IGNORECASE)
        if print_match:
            value = self._eval_expr(print_match.group(2))
            self.output.append(str(value))
            return

        # MsgBox
        msgbox_match = re.match(r'^\s*MsgBox\s+(.+)$', line, re.IGNORECASE)
        if msgbox_match:
            value = self._eval_expr(msgbox_match.group(1))
            self.output.append(str(value))
            return

    def _eval_expr(self, expr: str) -> Any:
        """Evaluate a VBA expression."""
        expr = expr.strip()

        # String literal
        if expr.startswith('"') and expr.endswith('"'):
            return expr[1:-1]

        # Range().Value - read
        range_get = re.match(r'^\s*Range\(\s*"([^"]+)"\s*\)\s*\.?\s*(Value|Value2)?\s*$', expr, re.IGNORECASE)
        if range_get:
            # Return placeholder - actual reading requires workbook context
            return f"Range({range_get.group(1)})"

        # Cells(r, c).Value - read
        cells_get = re.match(
            r'^\s*Cells\(\s*(.+?)\s*,\s*(.+?)\s*\)\s*\.?\s*(Value|Value2)?\s*$',
            expr, re.IGNORECASE
        )
        if cells_get:
            return f"Cells({cells_get.group(1)},{cells_get.group(2)})"

        # Variable reference
        if expr in self.variables:
            return self.variables[expr]

        # Numeric literal
        try:
            if '.' in expr:
                return float(expr)
            return int(expr)
        except ValueError:
            pass

        # Arithmetic expressions with variables
        try:
            # Replace variable names with values for evaluation
            eval_expr = expr
            for var_name, var_val in self.variables.items():
                if var_val is not None:
                    eval_expr = re.sub(r'\b' + re.escape(var_name) + r'\b', str(var_val), eval_expr)
            # Safe eval for math
            result = eval(eval_expr, {"__builtins__": {}}, {})
            return result
        except Exception:
            return expr

    def _eval_condition(self, condition: str) -> bool:
        """Evaluate a VBA condition."""
        # Simple comparison operators
        for op in ['>=', '<=', '<>', '!=', '=', '>', '<']:
            if op in condition:
                parts = condition.split(op, 1)
                if len(parts) == 2:
                    left = self._eval_expr(parts[0].strip())
                    right = self._eval_expr(parts[1].strip())

                    # Try numeric comparison
                    try:
                        left_num = float(left) if not isinstance(left, (int, float)) else left
                        right_num = float(right) if not isinstance(right, (int, float)) else right
                        if op == '>=': return left_num >= right_num
                        if op == '<=': return left_num <= right_num
                        if op in ['<>', '!=']: return left_num != right_num
                        if op == '=': return left_num == right_num
                        if op == '>': return left_num > right_num
                        if op == '<': return left_num < right_num
                    except (ValueError, TypeError):
                        if op == '=': return str(left) == str(right)
                        if op in ['<>', '!=']: return str(left) != str(right)

        # Boolean keywords
        condition_lower = condition.lower().strip()
        if condition_lower == 'true':
            return True
        if condition_lower == 'false':
            return False

        return False

    def _set_range_value(self, workbook: openpyxl.Workbook, cell_ref: str, value: Any):
        """Set a cell value using Range notation (e.g., 'A1' or 'A1:B5')."""
        ws = workbook.active

        # Handle range (merge or set all cells)
        if ':' in cell_ref:
            from openpyxl.utils.cell import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(cell_ref)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col, value=value)
        else:
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            col_str, row = coordinate_from_string(cell_ref)
            col = column_index_from_string(col_str)
            ws.cell(row=row, column=col, value=value)

        self.output.append(f"Set Range('{cell_ref}') = {value}")

    def _set_cells_value(self, workbook: openpyxl.Workbook, row: int, col: int, value: Any):
        """Set a cell value using Cells notation."""
        ws = workbook.active
        ws.cell(row=row, column=col, value=value)
        self.output.append(f"Set Cells({row}, {col}) = {value}")


# =============================================================================
# Python Macro Executor
# =============================================================================

class PythonMacroExecutor:
    """
    Execute Python macros with access to openpyxl and pandas.
    Sandboxed execution with limited builtins.
    """

    def __init__(self):
        self.output: List[str] = []
        self.errors: List[str] = []

    def execute(self, code: str, file_path: str) -> Dict[str, Any]:
        """Execute Python macro code on an Excel file."""
        try:
            # Load the workbook
            ext = os.path.splitext(file_path)[1].lower()
            if ext in ['.xlsx', '.xlsm']:
                wb = openpyxl.load_workbook(file_path)
            else:
                return {
                    "success": False,
                    "output": [],
                    "errors": [f"Python macros only support .xlsx files. Got: {ext}"],
                }

            # Prepare namespace
            import pandas as pd
            import numpy as np
            import excel_libs
            import excel_libs
            import excel_libs
            import excel_libs
            import excel_libs

            namespace = {
                'wb': wb,
                'ws': wb.active,
                'workbook': wb,
                'worksheet': wb.active,
                'pd': pd,
                'np': np,
                'excel_libs': excel_libs,
                'openpyxl': openpyxl,
                'get_column_letter': get_column_letter,
                'column_index_from_string': column_index_from_string,
                'print': lambda *args: self.output.append(' '.join(str(a) for a in args)),
                'range': range,
                'len': len,
                'int': int,
                'float': float,
                'str': str,
                'list': list,
                'dict': dict,
                'tuple': tuple,
                'enumerate': enumerate,
                'zip': zip,
                'sorted': sorted,
                'min': min,
                'max': max,
                'sum': sum,
                'abs': abs,
                'round': round,
                'True': True,
                'False': False,
                'None': None,
            }

            # Execute code
            exec(code, {"__builtins__": {}}, namespace)

            # Save workbook
            wb.save(file_path)
            wb.close()

            return {
                "success": True,
                "output": self.output,
                "errors": self.errors,
            }
        except Exception as e:
            return {
                "success": False,
                "output": self.output,
                "errors": [str(e), traceback.format_exc()],
            }


# =============================================================================
# Main Macro Execution Entry Point
# =============================================================================

def execute_macro(
    file_path: str,
    macro_code: str,
    language: str = "vba"
) -> Dict[str, Any]:
    """
    Execute a macro on an Excel file.
    language: 'vba' or 'python'
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in ['.xlsx', '.xlsm']:
        raise ValueError(f"Macro execution only supports .xlsx/.xlsm files. Got: {ext}")

    if language.lower() == "vba":
        wb = openpyxl.load_workbook(file_path)
        parser = VBAParser(macro_code)
        result = parser.parse_and_execute(wb)

        # Save if any modifications were made
        if result["success"]:
            try:
                wb.save(file_path)
            except Exception as e:
                result["errors"].append(f"Save error: {str(e)}")
            finally:
                wb.close()
        else:
            wb.close()

        return result

    elif language.lower() == "python":
        executor = PythonMacroExecutor()
        return executor.execute(macro_code, file_path)

    else:
        raise ValueError(f"Unsupported macro language: {language}. Use 'vba' or 'python'")


def list_macros(file_path: str) -> Dict[str, Any]:
    """
    List macros/VBA modules in an Excel file.
    Note: Full VBA extraction requires additional libraries.
    This provides basic detection.
    """
    ext = os.path.splitext(file_path)[1].lower()

    macros_found = []

    if ext == '.xlsm':
        # .xlsm files may contain VBA project
        # Check for vbaProject.bin in the ZIP structure
        import zipfile
        try:
            with zipfile.ZipFile(file_path, 'r') as zf:
                for name in zf.namelist():
                    if 'vba' in name.lower():
                        macros_found.append({
                            "name": name,
                            "type": "vba_module",
                        })
        except Exception:
            pass

    return {
        "file_path": file_path,
        "macros": macros_found,
        "count": len(macros_found),
        "note": "Full VBA extraction requires oletools or similar libraries. "
                "Python macros can be executed via the execute endpoint."
    }
